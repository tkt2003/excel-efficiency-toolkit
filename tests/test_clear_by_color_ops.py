from pathlib import Path
from unittest.mock import Mock

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from src.excel_efficiency_toolkit import clear_by_color_ops
from src.excel_efficiency_toolkit.clear_by_color_ops import (
    apply_multi_workbook_backup_plan,
    build_clear_by_color_batch_backup_dir,
    build_clear_by_color_backup_path,
    build_clear_detail_log_record,
    build_clear_ranges_from_cells,
    build_sheet_clear_plans,
    build_clear_summary_log_record,
    build_unique_backup_file_path,
    clear_workbook_file_with_openpyxl,
    clear_loaded_workbook_cells,
    execute_clear_active_workbook_plan,
    get_multi_workbook_backup_candidates,
    get_load_workbook_options_for_color_clear,
    is_supported_openpyxl_color_workbook,
    resolve_multi_workbook_save_mode,
    scan_workbook_color_matches,
    should_keep_vba_for_workbook,
    write_clear_by_color_log_workbook,
)
from src.excel_efficiency_toolkit.color_sum_ops import get_openpyxl_fill_key


def _build_scan_workbook(path: Path):
    yellow = PatternFill(fill_type="solid", fgColor="FFFF00")
    green = PatternFill(fill_type="solid", fgColor="00FF00")

    workbook = Workbook()
    visible = workbook.active
    visible.title = "可见表1"
    visible["A1"] = "样本"
    visible["A1"].fill = yellow
    visible["B1"] = 1
    visible["B1"].fill = yellow
    visible["C1"] = 2
    visible["C1"].fill = green
    visible["A2"] = 3
    visible["A2"].fill = yellow

    visible2 = workbook.create_sheet("可见表2")
    visible2["D4"] = 4
    visible2["D4"].fill = yellow

    hidden = workbook.create_sheet("隐藏表")
    hidden.sheet_state = "hidden"
    hidden["A1"] = 5
    hidden["A1"].fill = yellow

    workbook.save(path)
    workbook.close()


def _build_clear_workbook(path: Path):
    yellow = PatternFill(fill_type="solid", fgColor="FFFF00")

    workbook = Workbook()
    sheet1 = workbook.active
    sheet1.title = "Sheet1"
    sheet1["A1"] = 100
    sheet1["A1"].fill = yellow
    sheet1["B1"] = "=SUM(1,2)"
    sheet1["B1"].fill = yellow
    sheet1["C1"] = "保留"
    sheet1["C1"].fill = PatternFill(fill_type="solid", fgColor="00FF00")

    sheet2 = workbook.create_sheet("Sheet2")
    sheet2["D4"] = 200
    sheet2["D4"].fill = yellow
    sheet2["E4"] = "=1+1"
    sheet2["E4"].fill = yellow
    sheet2.merge_cells("G1:H1")
    sheet2["G1"] = "合并值"
    sheet2["G1"].fill = yellow

    workbook.save(path)
    workbook.close()


def test_is_supported_openpyxl_color_workbook_accepts_xlsx_xlsm_and_rejects_xls():
    assert is_supported_openpyxl_color_workbook("book.xlsx")
    assert is_supported_openpyxl_color_workbook("book.xlsm")
    assert not is_supported_openpyxl_color_workbook("book.xls")


def test_build_clear_by_color_backup_path_does_not_overwrite(tmp_path):
    target_path = tmp_path / "目标.xlsx"
    target_path.write_text("a", encoding="utf-8")

    first = build_clear_by_color_backup_path(str(target_path), timestamp="20260612_120000")
    Path(first).write_text("b", encoding="utf-8")
    second = build_clear_by_color_backup_path(str(target_path), timestamp="20260612_120000")

    assert first != second


def test_build_clear_by_color_batch_backup_dir_uses_expected_name(tmp_path):
    backup_dir = build_clear_by_color_batch_backup_dir(str(tmp_path), timestamp="20260613_001530")
    assert backup_dir == str(tmp_path / "按颜色清空内容_备份_20260613_001530")


def test_build_unique_backup_file_path_auto_numbers_same_name(tmp_path):
    batch_dir = tmp_path / "backup"
    batch_dir.mkdir()
    first = build_unique_backup_file_path(str(batch_dir), "目标.xlsx")
    Path(first).write_text("a", encoding="utf-8")
    second = build_unique_backup_file_path(str(batch_dir), "目标.xlsx")

    assert first != second
    assert second.endswith("目标_2.xlsx")


def test_scan_workbook_color_matches_finds_same_color_and_skips_hidden_sheet(tmp_path):
    path = tmp_path / "扫描.xlsx"
    _build_scan_workbook(path)

    workbook = load_workbook(path)
    selected_color_key = get_openpyxl_fill_key(workbook["可见表1"]["A1"])
    workbook.close()

    plan = scan_workbook_color_matches(str(path), selected_color_key)

    assert plan["matched_sheet_count"] == 2
    assert plan["matched_cell_count"] == 4
    names = [item["sheet_name"] for item in plan["sheet_plans"]]
    assert "隐藏表" not in names


def test_clear_loaded_workbook_cells_clears_values_and_formulas_but_keeps_fill(tmp_path):
    path = tmp_path / "清空.xlsx"
    _build_clear_workbook(path)
    workbook = load_workbook(path)

    result = clear_loaded_workbook_cells(
        workbook,
        str(path),
        [{"sheet_name": "Sheet1", "cells": ["A1", "B1"]}],
    )
    workbook.save(path)
    workbook.close()

    reloaded = load_workbook(path)
    assert reloaded["Sheet1"]["A1"].value is None
    assert reloaded["Sheet1"]["B1"].value is None
    assert reloaded["Sheet1"]["A1"].fill.fgColor.rgb == "00FFFF00"
    assert reloaded["Sheet1"]["B1"].fill.fgColor.rgb == "00FFFF00"
    assert result["cleared_cell_count"] == 2
    reloaded.close()


def test_clear_loaded_workbook_cells_supports_multiple_sheets(tmp_path):
    path = tmp_path / "多sheet.xlsx"
    _build_clear_workbook(path)
    workbook = load_workbook(path)

    clear_loaded_workbook_cells(
        workbook,
        str(path),
        [
            {"sheet_name": "Sheet1", "cells": ["A1"]},
            {"sheet_name": "Sheet2", "cells": ["D4", "E4"]},
        ],
    )
    workbook.save(path)
    workbook.close()

    reloaded = load_workbook(path)
    assert reloaded["Sheet1"]["A1"].value is None
    assert reloaded["Sheet2"]["D4"].value is None
    assert reloaded["Sheet2"]["E4"].value is None
    reloaded.close()


def test_clear_loaded_workbook_cells_skips_non_top_left_merged_cell_without_crashing(tmp_path):
    path = tmp_path / "合并.xlsx"
    _build_clear_workbook(path)
    workbook = load_workbook(path)
    detail_records = []

    result = clear_loaded_workbook_cells(
        workbook,
        str(path),
        [{"sheet_name": "Sheet2", "cells": ["H1"]}],
        detail_records=detail_records,
    )

    assert result["cleared_cell_count"] == 0
    assert result["skipped_merged_cell_count"] == 1
    assert detail_records[0]["状态"] == "跳过"
    workbook.close()


def test_build_clear_ranges_from_cells_groups_same_row_contiguous_columns():
    ranges = build_clear_ranges_from_cells([(2, 1), (2, 2), (2, 4), (3, 1)])
    assert ranges == ["A2:B2", "D2", "A3"]


def test_build_clear_summary_log_record_structure_is_correct():
    record = build_clear_summary_log_record(
        "D:/a.xlsx",
        2,
        10,
        8,
        "com",
        "批次备份",
        "D:/backup.xlsx",
        "成功",
        "说明",
    )
    assert record["匹配工作表数量"] == 2
    assert record["清空单元格数量"] == 8
    assert record["保存方式"] == "com"
    assert record["备份方式"] == "批次备份"


def test_build_clear_detail_log_record_structure_is_correct():
    record = build_clear_detail_log_record("D:/a.xlsx", "Sheet1", "C3", "跳过", "合并单元格")
    assert record["工作表名"] == "Sheet1"
    assert record["单元格地址"] == "C3"


def test_write_clear_by_color_log_workbook_generates_file(tmp_path):
    log_path = write_clear_by_color_log_workbook(
        str(tmp_path),
        [build_clear_summary_log_record("D:/a.xlsx", 1, 2, 2, "openpyxl", "批次备份", "D:/bak.xlsx", "成功", "完成")],
        [build_clear_detail_log_record("D:/a.xlsx", "Sheet1", "A1", "成功", "")],
        timestamp="20260612_120000",
    )

    assert Path(log_path).exists()
    workbook = load_workbook(log_path)
    assert workbook.sheetnames == ["处理汇总", "处理明细"]
    assert workbook["处理汇总"]["E1"].value == "保存方式"
    assert workbook["处理汇总"]["E2"].value == "openpyxl"
    assert workbook["处理汇总"]["F1"].value == "备份方式"
    assert workbook["处理汇总"]["F2"].value == "批次备份"
    assert workbook["处理汇总"]["G1"].value == "备份路径"
    workbook.close()


def test_get_load_workbook_options_for_color_clear_uses_keep_vba_for_xlsm():
    assert should_keep_vba_for_workbook("book.xlsm") is True
    assert should_keep_vba_for_workbook("book.xlsx") is False
    assert get_load_workbook_options_for_color_clear("book.xlsm") == {"keep_vba": True}


def test_resolve_multi_workbook_save_mode_uses_com_for_external_links(monkeypatch):
    monkeypatch.setattr(clear_by_color_ops, "workbook_has_external_links", lambda path: True)
    assert resolve_multi_workbook_save_mode("book.xlsx") == "com"


def test_resolve_multi_workbook_save_mode_uses_openpyxl_for_plain_file(monkeypatch):
    monkeypatch.setattr(clear_by_color_ops, "workbook_has_external_links", lambda path: False)
    assert resolve_multi_workbook_save_mode("book.xlsx") == "openpyxl"


def test_clear_workbook_file_with_openpyxl_saves_plain_workbook(tmp_path):
    path = tmp_path / "plain.xlsx"
    _build_clear_workbook(path)

    result = clear_workbook_file_with_openpyxl(
        str(path),
        [{"sheet_name": "Sheet1", "cells": ["A1", "B1"]}],
    )

    reloaded = load_workbook(path)
    assert reloaded["Sheet1"]["A1"].value is None
    assert reloaded["Sheet1"]["B1"].value is None
    assert result["cleared_cell_count"] == 2
    reloaded.close()


def test_build_sheet_clear_plans_contains_contiguous_ranges():
    plans = build_sheet_clear_plans([
        {"sheet_name": "Sheet1", "cells": ["A2", "B2", "D2"]},
    ])

    assert plans[0]["range_addresses"] == ["A2:B2", "D2"]
    assert plans[0]["range_count"] == 2


def test_get_multi_workbook_backup_candidates_only_returns_files_that_will_be_modified():
    records = [
        {"file_path": "a.xlsx", "needs_modify": True},
        {"file_path": "b.xlsx", "needs_modify": False},
        {"file_path": "c.xls", "needs_modify": False},
    ]

    candidates = get_multi_workbook_backup_candidates(records)
    assert [item["file_path"] for item in candidates] == ["a.xlsx"]


def test_apply_multi_workbook_backup_plan_only_copies_files_that_will_be_modified(tmp_path):
    file_a = tmp_path / "A.xlsx"
    file_b = tmp_path / "B.xlsx"
    file_c = tmp_path / "C.xls"
    file_a.write_text("a", encoding="utf-8")
    file_b.write_text("b", encoding="utf-8")
    file_c.write_text("c", encoding="utf-8")
    records = [
        {"file_path": str(file_a), "file_name": "A.xlsx", "needs_modify": True, "backup_mode": "", "backup_path": ""},
        {"file_path": str(file_b), "file_name": "B.xlsx", "needs_modify": False, "backup_mode": "无需备份", "backup_path": ""},
        {"file_path": str(file_c), "file_name": "C.xls", "needs_modify": False, "backup_mode": "跳过", "backup_path": ""},
    ]

    batch_dir = apply_multi_workbook_backup_plan(records, str(tmp_path), skip_backup=False, timestamp="20260613_001530")

    assert Path(batch_dir).exists()
    assert Path(records[0]["backup_path"]).exists()
    assert records[0]["backup_mode"] == "批次备份"
    assert records[1]["backup_path"] == ""
    assert records[2]["backup_path"] == ""
    assert len(list(Path(batch_dir).iterdir())) == 1


def test_apply_multi_workbook_backup_plan_respects_user_skip_backup(tmp_path):
    file_a = tmp_path / "A.xlsx"
    file_a.write_text("a", encoding="utf-8")
    records = [
        {"file_path": str(file_a), "file_name": "A.xlsx", "needs_modify": True, "backup_mode": "", "backup_path": ""},
    ]

    batch_dir = apply_multi_workbook_backup_plan(records, str(tmp_path), skip_backup=True, timestamp="20260613_001530")

    assert batch_dir == ""
    assert records[0]["backup_mode"] == "用户选择不备份"
    assert records[0]["backup_path"] == "用户选择不备份"


def test_apply_multi_workbook_backup_plan_auto_numbers_same_file_name(tmp_path):
    first_dir = tmp_path / "one"
    second_dir = tmp_path / "two"
    first_dir.mkdir()
    second_dir.mkdir()
    file_a = first_dir / "目标.xlsx"
    file_b = second_dir / "目标.xlsx"
    file_a.write_text("a", encoding="utf-8")
    file_b.write_text("b", encoding="utf-8")
    records = [
        {"file_path": str(file_a), "file_name": file_a.name, "needs_modify": True, "backup_mode": "", "backup_path": ""},
        {"file_path": str(file_b), "file_name": file_b.name, "needs_modify": True, "backup_mode": "", "backup_path": ""},
    ]

    batch_dir = apply_multi_workbook_backup_plan(records, str(tmp_path), skip_backup=False, timestamp="20260613_001530")

    backup_names = sorted(path.name for path in Path(batch_dir).iterdir())
    assert backup_names == ["目标.xlsx", "目标_2.xlsx"]


def test_execute_clear_active_workbook_plan_does_not_call_openpyxl_save(monkeypatch):
    class FakeRange:
        def __init__(self):
            self.clear_count = 0

        def ClearContents(self):
            self.clear_count += 1

    class FakeSheet:
        def __init__(self):
            self.ranges = {}

        def Range(self, address):
            self.ranges.setdefault(address, FakeRange())
            return self.ranges[address]

    class FakeWorkbook:
        def __init__(self):
            self.Name = "目标.xlsx"
            self.FullName = r"D:\tmp\目标.xlsx"
            self.Saved = True
            self.Worksheets = []
            self.sheet = FakeSheet()
            fake_worksheet = Mock()
            fake_worksheet.Name = "Sheet1"
            fake_worksheet.Range = self.sheet.Range
            self.Worksheets.append(fake_worksheet)

    fake_workbook = FakeWorkbook()
    monkeypatch.setattr(clear_by_color_ops, "_get_active_excel", lambda *args, **kwargs: object())
    monkeypatch.setattr(clear_by_color_ops, "_find_open_workbook_by_path", lambda excel, path: fake_workbook)
    monkeypatch.setattr(clear_by_color_ops, "load_workbook", Mock(side_effect=AssertionError("不应调用 openpyxl 保存当前活动工作簿")))

    plan = {
        "workbook_path": r"D:\tmp\目标.xlsx",
        "matched_sheet_count": 1,
        "matched_cell_count": 3,
        "current_color_text": "Interior.Color=65535",
        "scan_seconds": 0.01,
        "sheet_clear_plans": [
            {
                "sheet_name": "Sheet1",
                "matched_cell_count": 3,
                "range_addresses": ["A1:B1", "D1"],
                "range_count": 2,
            }
        ],
    }

    result = execute_clear_active_workbook_plan(plan)

    assert result["cleared_cell_count"] == 3
    assert result["range_group_count"] == 2
    assert fake_workbook.sheet.ranges["A1:B1"].clear_count == 1
    assert fake_workbook.sheet.ranges["D1"].clear_count == 1


def test_clear_multiple_workbooks_by_color_skip_backup_marks_summary(monkeypatch, tmp_path):
    file_a = tmp_path / "A.xlsx"
    file_a.write_text("a", encoding="utf-8")
    monkeypatch.setattr(clear_by_color_ops, "prepare_active_fill_color_context", lambda logger=None: {
        "workbook_name": "样本.xlsx",
        "selected_sheet_name": "Sheet1",
        "selected_cell_address": "A1",
        "selected_color_key": ("solid", "rgb", "FFFF00", None, None, 0),
        "current_color_text": "Interior.Color=65535",
    })
    monkeypatch.setattr(clear_by_color_ops, "collect_multi_workbook_clear_records", lambda paths, key, logger=None: [
        {
            "file_path": str(file_a),
            "file_name": "A.xlsx",
            "is_supported": True,
            "has_external_links": False,
            "matched_sheet_count": 1,
            "matched_cell_count": 2,
            "save_mode": "openpyxl",
            "needs_modify": True,
            "backup_mode": "",
            "backup_path": "",
            "status": "待处理",
            "note": "",
            "sheet_plans": [{"sheet_name": "Sheet1", "cells": ["A1", "B1"]}],
            "sheet_clear_plans": [],
        }
    ])
    monkeypatch.setattr(clear_by_color_ops, "clear_workbook_file_with_openpyxl", lambda *args, **kwargs: {
        "cleared_cell_count": 2,
        "skipped_merged_cell_count": 0,
        "range_group_count": 0,
    })

    result = clear_by_color_ops.clear_multiple_workbooks_by_color([str(file_a)], skip_backup=True)

    assert result["batch_backup_dir"] == ""
    assert result["summary_records"][0]["备份方式"] == "用户选择不备份"
    assert result["summary_records"][0]["备份路径"] == "用户选择不备份"

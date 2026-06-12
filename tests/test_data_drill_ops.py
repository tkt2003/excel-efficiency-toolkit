from openpyxl import Workbook, load_workbook

from src.excel_efficiency_toolkit.data_drill_ops import (
    build_data_drill_range_records,
    build_unique_output_path,
    is_excel_error_value,
    is_office_temp_file,
    is_supported_openpyxl_workbook,
    read_drill_cell_from_workbook,
    read_drill_range_from_workbook,
    resolve_sheet_name,
    summarize_data_drill_records,
    write_data_drill_result_workbook,
)


def test_supported_openpyxl_workbook_accepts_xlsx_and_xlsm():
    assert is_supported_openpyxl_workbook("book.xlsx")
    assert is_supported_openpyxl_workbook("book.XLSM")


def test_supported_openpyxl_workbook_rejects_xls():
    assert not is_supported_openpyxl_workbook("book.xls")


def test_office_temp_file_detection():
    assert is_office_temp_file(r"D:\tmp\~$book.xlsx")
    assert not is_office_temp_file(r"D:\tmp\book.xlsx")


def test_build_unique_output_path_adds_sequence_without_overwrite(tmp_path):
    first_path = tmp_path / "数据穿透查询结果.xlsx"
    second_path = tmp_path / "数据穿透查询结果_2.xlsx"
    first_path.write_text("exists", encoding="utf-8")
    second_path.write_text("exists", encoding="utf-8")

    output_path = build_unique_output_path(str(tmp_path), base_name="数据穿透查询结果")

    assert output_path == str(tmp_path / "数据穿透查询结果_3.xlsx")
    assert first_path.read_text(encoding="utf-8") == "exists"
    assert second_path.read_text(encoding="utf-8") == "exists"


def test_resolve_sheet_name_matches_strip_and_casefold():
    workbook = Workbook()
    workbook.active.title = " Data "

    actual_name, available_names, message = resolve_sheet_name(workbook, " data ")

    assert actual_name == " Data "
    assert available_names == [" Data "]
    assert message == "匹配同名 Sheet"


def test_resolve_sheet_name_missing_returns_available_sheet_list_and_skip_message():
    workbook = Workbook()
    workbook.active.title = "资产负债表"
    workbook.create_sheet("利润表")

    actual_name, available_names, message = resolve_sheet_name(workbook, "货币资金")

    assert actual_name is None
    assert available_names == ["资产负债表", "利润表"]
    assert "缺少同名 Sheet" in message
    assert "资产负债表" in message
    assert "利润表" in message


def test_read_drill_cell_from_workbook_reads_plain_value(tmp_path):
    path = tmp_path / "source.xlsx"
    _save_workbook(path, "货币资金", {"C5": 123.45})

    record = read_drill_cell_from_workbook(str(path), "货币资金", "C5")

    assert record["status"] == "成功"
    assert record["value"] == 123.45
    assert record["values_by_address"] == {"C5": 123.45}
    assert record["is_empty"] is False
    assert record["is_error_value"] is False


def test_read_drill_range_from_workbook_reads_area_values(tmp_path):
    path = tmp_path / "source.xlsx"
    _save_workbook(path, "货币资金", {"B2": 1, "C2": 2, "B3": 3, "C3": 4})

    record = read_drill_range_from_workbook(str(path), "货币资金", "B2:C3")

    assert record["status"] == "成功"
    assert record["range_address"] == "B2:C3"
    assert record["cell_address"] == "B2"
    assert record["values_by_address"] == {"B2": 1, "C2": 2, "B3": 3, "C3": 4}
    assert record["value"] == 1


def test_excel_error_value_detection_and_reading(tmp_path):
    path = tmp_path / "source.xlsx"
    _save_workbook(path, "货币资金", {"C5": "#VALUE!"})

    record = read_drill_cell_from_workbook(str(path), "货币资金", "C5")

    assert is_excel_error_value("#DIV/0!")
    assert not is_excel_error_value("普通文本")
    assert record["status"] == "成功"
    assert record["value"] == "#VALUE!"
    assert record["is_error_value"] is True


def test_formula_cache_empty_writes_prompt_message(tmp_path):
    path = tmp_path / "formula.xlsx"
    _save_workbook(path, "货币资金", {"C5": "=SUM(1,2)"})

    record = read_drill_cell_from_workbook(str(path), "货币资金", "C5")

    assert record["status"] == "成功"
    assert record["value"] is None
    assert record["message"] == "公式缓存为空，取值可能需要先打开源文件计算并保存"


def test_build_records_skips_temp_xls_and_missing_sheet(tmp_path):
    normal_path = tmp_path / "normal.xlsx"
    missing_sheet_path = tmp_path / "missing_sheet.xlsx"
    xls_path = tmp_path / "old.xls"
    temp_path = tmp_path / "~$temp.xlsx"
    _save_workbook(normal_path, "货币资金", {"C5": 1})
    _save_workbook(missing_sheet_path, "资产负债表", {"C5": 2})
    xls_path.write_bytes(b"placeholder")
    temp_path.write_bytes(b"placeholder")

    records = build_data_drill_range_records(
        [str(normal_path), str(missing_sheet_path), str(xls_path), str(temp_path)],
        "货币资金",
        "C5",
    )

    assert [record["status"] for record in records] == ["成功", "跳过", "跳过", "跳过"]
    assert "缺少同名 Sheet" in records[1]["message"]
    assert records[2]["message"] == "请另存为 xlsx/xlsm 后再处理"
    assert records[3]["message"] == "临时文件已跳过"


def test_summarize_data_drill_records_counts_statuses():
    records = [
        {"status": "成功"},
        {"status": "成功"},
        {"status": "跳过"},
        {"status": "失败"},
    ]

    summary = summarize_data_drill_records(records)

    assert summary == {
        "success_count": 2,
        "skipped_count": 1,
        "failed_count": 1,
        "total_count": 4,
    }


def test_write_result_workbook_keeps_single_cell_usage(tmp_path):
    output_path = tmp_path / "数据穿透查询结果.xlsx"
    records = [
        {
            "source_file_name": "source.xlsx",
            "source_file_path": str(tmp_path / "source.xlsx"),
            "target_sheet_name": "货币资金",
            "cell_address": "C5",
            "range_address": "C5",
            "values_by_address": {"C5": "#N/A"},
            "value": "#N/A",
            "status": "成功",
            "message": "读取成功",
        }
    ]

    write_data_drill_result_workbook(records, str(output_path), "货币资金", "C5")
    workbook = load_workbook(output_path)
    sheet = workbook["数据穿透查询结果"]
    headers = [cell.value for cell in sheet[1]]

    assert headers == ["源文件名", "源文件路径", "工作表名", "C5", "状态", "说明"]
    assert sheet["D2"].value == "#N/A"
    assert sheet["E2"].value == "成功"
    workbook.close()


def test_write_result_workbook_supports_multi_file_area_output(tmp_path):
    output_path = tmp_path / "数据穿透查询结果.xlsx"
    records = [
        {
            "source_file_name": "source.xlsx",
            "source_file_path": str(tmp_path / "source.xlsx"),
            "target_sheet_name": "货币资金",
            "range_address": "B2:C3",
            "values_by_address": {"B2": 1, "C2": 2, "B3": 3, "C3": "读取异常：示例"},
            "status": "成功",
            "message": "读取成功",
        }
    ]

    write_data_drill_result_workbook(records, str(output_path), "货币资金", "B2:C3")
    workbook = load_workbook(output_path)
    sheet = workbook["数据穿透查询结果"]

    assert [cell.value for cell in sheet[1]] == [
        "源文件名",
        "源文件路径",
        "工作表名",
        "B2",
        "C2",
        "B3",
        "C3",
        "状态",
        "说明",
    ]
    assert [sheet.cell(row=2, column=index).value for index in range(4, 8)] == [1, 2, 3, "读取异常：示例"]
    workbook.close()


def _save_workbook(path, sheet_name, values_by_cell):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for cell_address, value in values_by_cell.items():
        sheet[cell_address] = value
    workbook.save(path)
    workbook.close()

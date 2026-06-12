import pytest
from openpyxl import Workbook, load_workbook

from src.excel_efficiency_toolkit.template_multi_link_ops import (
    LOG_SHEET_NAME,
    RULE_SHEET_NAME,
    SETTINGS_SHEET_NAME,
    TemplateLinkReplacement,
    TemplateMultiLinkRule,
    TemplateMultiLinkSettings,
    build_template_multi_link_actions,
    create_template_multi_link_rule_workbook,
    read_template_multi_link_rules,
    read_template_multi_link_settings,
    summarize_template_multi_link_actions,
    write_template_multi_link_results_to_workbook,
)


def test_create_rule_workbook_writes_settings_rows_and_dynamic_link_columns(tmp_path):
    template = tmp_path / "template.xlsx"
    template.write_text("placeholder", encoding="utf-8")
    single_1 = tmp_path / "single_公司1.xlsx"
    single_2 = tmp_path / "single_公司2.xlsx"
    single_1.write_text("placeholder", encoding="utf-8")
    single_2.write_text("placeholder", encoding="utf-8")
    old_single = str(tmp_path / "old_single.xlsx")
    old_group = str(tmp_path / "old_group.xlsx")

    rule_path = create_template_multi_link_rule_workbook(
        template_path=str(template),
        output_dir=str(tmp_path),
        old_links=[old_single, old_group],
        main_old_link=old_single,
        main_source_paths=[str(single_1), str(single_2)],
    )

    workbook = load_workbook(rule_path)
    try:
        assert workbook.sheetnames == ["使用说明", SETTINGS_SHEET_NAME, RULE_SHEET_NAME, LOG_SHEET_NAME]
        settings_sheet = workbook[SETTINGS_SHEET_NAME]
        assert settings_sheet["B2"].value == str(template.resolve())
        assert settings_sheet["B3"].value == str(tmp_path.resolve())
        assert settings_sheet["B4"].value == "_批量生成"
        assert settings_sheet["B5"].value == old_single

        rule_sheet = workbook[RULE_SHEET_NAME]
        assert rule_sheet.max_row == 3
        assert [rule_sheet.cell(row=1, column=column).value for column in range(1, 11)] == [
            "是否执行",
            "输出文件名",
            "旧链接 1",
            "新链接 1",
            "旧链接 2",
            "新链接 2",
            "状态",
            "说明",
            "输出文件路径",
            "主源文件路径",
        ]
        assert rule_sheet.freeze_panes == "C2"
        assert rule_sheet["B2"].value == "single_公司1_批量生成.xlsx"
        assert rule_sheet["D2"].value == str(single_1.resolve())
        assert rule_sheet["F2"].value is None
        assert rule_sheet["I2"].value.endswith("single_公司1_批量生成.xlsx")
        assert rule_sheet["J2"].value == str(single_1.resolve())
    finally:
        workbook.close()


@pytest.mark.parametrize(
    ("suffix", "template_name", "existing_name", "expected_name"),
    [
        (None, "template.xlsx", "", "single_公司1_批量生成.xlsx"),
        ("_附注", "template.xlsx", "", "single_公司1_附注.xlsx"),
        ("", "template.xlsx", "", "single_公司1.xlsx"),
        (None, "template.xlsm", "", "single_公司1_批量生成.xlsm"),
        (None, "template.xlsx", "single_公司1_批量生成.xlsx", "single_公司1_批量生成_2.xlsx"),
    ],
)
def test_rule_workbook_output_filename_rules(tmp_path, suffix, template_name, existing_name, expected_name):
    template = tmp_path / template_name
    template.write_text("placeholder", encoding="utf-8")
    source_dir = tmp_path / "source"
    output_dir = tmp_path / "output"
    source_dir.mkdir()
    output_dir.mkdir()
    single = source_dir / "single_公司1.xlsx"
    single.write_text("placeholder", encoding="utf-8")
    if existing_name:
        (output_dir / existing_name).write_text("exists", encoding="utf-8")

    rule_path = create_template_multi_link_rule_workbook(
        template_path=str(template),
        output_dir=str(output_dir),
        old_links=["old_single.xlsx"],
        main_old_link="old_single.xlsx",
        main_source_paths=[str(single)],
        output_name_suffix=suffix,
    )

    workbook = load_workbook(rule_path)
    try:
        assert workbook[RULE_SHEET_NAME]["B2"].value == expected_name
        assert workbook[RULE_SHEET_NAME]["G2"].value.endswith(expected_name)
    finally:
        workbook.close()


def test_read_rules_builds_multi_link_actions_and_skips_execute_no(tmp_path):
    template = tmp_path / "template.xlsx"
    template.write_text("placeholder", encoding="utf-8")
    single_1 = tmp_path / "single_公司1.xlsx"
    single_2 = tmp_path / "single_公司2.xlsx"
    group_1 = tmp_path / "group_公司1.xlsx"
    single_1.write_text("placeholder", encoding="utf-8")
    single_2.write_text("placeholder", encoding="utf-8")
    group_1.write_text("placeholder", encoding="utf-8")

    rule_path = create_template_multi_link_rule_workbook(
        template_path=str(template),
        output_dir=str(tmp_path),
        old_links=["old_single.xlsx", "old_group.xlsx"],
        main_old_link="old_single.xlsx",
        main_source_paths=[str(single_1), str(single_2)],
    )
    workbook = load_workbook(rule_path)
    sheet = workbook[RULE_SHEET_NAME]
    sheet["F2"] = str(group_1)
    sheet["A3"] = "否"
    workbook.save(rule_path)
    workbook.close()

    settings = read_template_multi_link_settings(rule_path)
    rules = read_template_multi_link_rules(rule_path)
    actions = build_template_multi_link_actions(rules, settings)

    assert len(rules) == 2
    assert len(actions[0].replacements) == 2
    assert actions[0].replacements[0].old_link_path == "old_single.xlsx"
    assert actions[0].replacements[0].new_link_path == str(single_1.resolve())
    assert actions[0].replacements[1].old_link_path == "old_group.xlsx"
    assert actions[0].replacements[1].new_link_path == str(group_1.resolve())
    assert actions[1].status == "跳过"
    assert actions[1].message == "是否执行为否，已跳过"


def test_empty_new_link_is_not_converted_to_replacement(tmp_path):
    template = tmp_path / "template.xlsx"
    template.write_text("placeholder", encoding="utf-8")
    single = tmp_path / "single_公司1.xlsx"
    single.write_text("placeholder", encoding="utf-8")

    rule_path = create_template_multi_link_rule_workbook(
        template_path=str(template),
        output_dir=str(tmp_path),
        old_links=["old_single.xlsx", "old_group.xlsx"],
        main_old_link="old_single.xlsx",
        main_source_paths=[str(single)],
    )

    settings = read_template_multi_link_settings(rule_path)
    actions = build_template_multi_link_actions(read_template_multi_link_rules(rule_path), settings)

    assert len(actions) == 1
    assert len(actions[0].replacements) == 1
    assert actions[0].replacements[0].old_link_path == "old_single.xlsx"


def test_summarize_actions_counts_statuses():
    settings = TemplateMultiLinkSettings("template.xlsx", ".", "_批量生成", "old.xlsx")
    rules = [
        TemplateMultiLinkRule(2, "是", "a.xlsx", "a.xlsx", "source.xlsx", [TemplateLinkReplacement("old", "new")]),
        TemplateMultiLinkRule(3, "否", "b.xlsx", "b.xlsx", "source.xlsx", [TemplateLinkReplacement("old", "new")]),
        TemplateMultiLinkRule(4, "是", "c.xlsx", "c.xlsx", "", [TemplateLinkReplacement("old", "new")]),
    ]

    actions = build_template_multi_link_actions(rules, settings)
    actions[0].status = "成功"
    actions[2].status = "失败"

    assert summarize_template_multi_link_actions(actions) == {
        "total_count": 3,
        "success_count": 1,
        "skipped_count": 1,
        "failed_count": 1,
    }


def test_temp_and_unsupported_main_source_files_are_skipped_when_creating_rule_workbook(tmp_path):
    template = tmp_path / "template.xlsx"
    template.write_text("placeholder", encoding="utf-8")
    valid = tmp_path / "single_公司1.xlsm"
    valid.write_text("placeholder", encoding="utf-8")

    rule_path = create_template_multi_link_rule_workbook(
        template_path=str(template),
        output_dir=str(tmp_path),
        old_links=["old_single.xlsx"],
        main_old_link="old_single.xlsx",
        main_source_paths=[str(tmp_path / "~$temp.xlsx"), str(tmp_path / "bad.csv"), str(valid)],
    )

    workbook = load_workbook(rule_path)
    try:
        sheet = workbook[RULE_SHEET_NAME]
        assert sheet.max_row == 2
        assert sheet["H2"].value == str(valid.resolve())
    finally:
        workbook.close()


@pytest.mark.parametrize(
    ("setting_cell", "message"),
    [
        ("B2", "规则表参数缺少模板路径。"),
        ("B3", "规则表参数缺少输出目录。"),
        ("B5", "规则表参数缺少主旧链接。"),
    ],
)
def test_read_settings_reports_required_parameter_errors(tmp_path, setting_cell, message):
    path = tmp_path / "rule.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SETTINGS_SHEET_NAME
    sheet.append(["参数项", "参数值", "说明"])
    sheet.append(["模板路径", "template.xlsx", ""])
    sheet.append(["输出目录", str(tmp_path), ""])
    sheet.append(["输出后缀", "_批量生成", ""])
    sheet.append(["主旧链接", "old.xlsx", ""])
    workbook.save(path)
    workbook.close()

    workbook = load_workbook(path)
    workbook[SETTINGS_SHEET_NAME][setting_cell] = ""
    workbook.save(path)
    workbook.close()

    with pytest.raises(ValueError, match=message):
        read_template_multi_link_settings(path)


def test_write_results_updates_rule_sheet_and_log_sheet(tmp_path):
    template = tmp_path / "template.xlsx"
    template.write_text("placeholder", encoding="utf-8")
    single = tmp_path / "single_公司1.xlsx"
    single.write_text("placeholder", encoding="utf-8")
    rule_path = create_template_multi_link_rule_workbook(
        template_path=str(template),
        output_dir=str(tmp_path),
        old_links=["old_single.xlsx"],
        main_old_link="old_single.xlsx",
        main_source_paths=[str(single)],
    )

    settings = read_template_multi_link_settings(rule_path)
    actions = build_template_multi_link_actions(read_template_multi_link_rules(rule_path), settings)
    actions[0].status = "成功"
    actions[0].message = "已生成并替换 1 个链接"
    actions[0].replacements[0].status = "成功"
    actions[0].replacements[0].message = "已替换"

    write_template_multi_link_results_to_workbook(rule_path, actions)

    workbook = load_workbook(rule_path)
    try:
        rule_sheet = workbook[RULE_SHEET_NAME]
        log_sheet = workbook[LOG_SHEET_NAME]
        assert rule_sheet["E2"].value == "成功"
        assert rule_sheet["F2"].value == "已生成并替换 1 个链接"
        assert log_sheet["F2"].value == "成功"
        assert log_sheet["G2"].value == "已替换"
    finally:
        workbook.close()


def test_read_rules_uses_headers_instead_of_fixed_columns(tmp_path):
    template = tmp_path / "template.xlsx"
    template.write_text("placeholder", encoding="utf-8")
    single = tmp_path / "single_公司1.xlsx"
    group = tmp_path / "group_公司1.xlsx"
    single.write_text("placeholder", encoding="utf-8")
    group.write_text("placeholder", encoding="utf-8")

    rule_path = create_template_multi_link_rule_workbook(
        template_path=str(template),
        output_dir=str(tmp_path),
        old_links=["old_single.xlsx", "old_group.xlsx"],
        main_old_link="old_single.xlsx",
        main_source_paths=[str(single)],
    )

    workbook = load_workbook(rule_path)
    try:
        sheet = workbook[RULE_SHEET_NAME]
        status_value = sheet["G2"].value
        message_value = sheet["H2"].value
        output_path_value = sheet["I2"].value
        main_source_value = sheet["J2"].value
        sheet["G1"] = "输出文件路径"
        sheet["G2"] = output_path_value
        sheet["H1"] = "主源文件路径"
        sheet["H2"] = main_source_value
        sheet["I1"] = "状态"
        sheet["I2"] = status_value
        sheet["J1"] = "说明"
        sheet["J2"] = message_value
        sheet["K1"] = "备注"
        sheet["K2"] = "测试列"
        workbook.save(rule_path)
    finally:
        workbook.close()

    settings = read_template_multi_link_settings(rule_path)
    rules = read_template_multi_link_rules(rule_path)
    actions = build_template_multi_link_actions(rules, settings)

    assert len(rules) == 1
    assert rules[0].output_path.endswith("single_公司1_批量生成.xlsx")
    assert rules[0].main_source_path == str(single.resolve())
    assert len(actions[0].replacements) == 1
    assert actions[0].replacements[0].new_link_path == str(single.resolve())

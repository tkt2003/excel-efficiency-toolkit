from openpyxl import load_workbook

from src.excel_efficiency_toolkit.link_replace_ops import (
    LOG_SHEET_NAME,
    RULE_SHEET_NAME,
    LinkReplaceSettings,
    build_backup_path,
    build_link_replace_actions,
    create_link_replace_rule_workbook,
    group_link_replace_actions_by_workbook_path,
    is_excel_workbook_file,
    is_office_temp_file,
    read_link_replace_rules,
    should_execute_rule,
    summarize_link_replace_actions,
    write_link_replace_results_to_workbook,
)


def test_excel_file_detection_accepts_supported_extensions_and_skips_temp():
    assert is_excel_workbook_file("book.xlsx")
    assert is_excel_workbook_file("book.XLSM")
    assert is_excel_workbook_file("book.xls")
    assert not is_excel_workbook_file("book.csv")
    assert not is_excel_workbook_file("~$book.xlsx")


def test_office_temp_file_detection():
    assert is_office_temp_file(r"D:\tmp\~$book.xlsx")
    assert not is_office_temp_file(r"D:\tmp\book.xlsx")


def test_should_execute_rule_defaults_to_yes_and_parses_no_values():
    assert should_execute_rule("")
    assert should_execute_rule(None)
    assert should_execute_rule("是")
    assert should_execute_rule("yes")
    assert not should_execute_rule("否")
    assert not should_execute_rule("NO")
    assert not should_execute_rule("0")


def test_create_and_read_rule_workbook_skips_empty_new_link(tmp_path):
    target = tmp_path / "target.xlsx"
    target.write_text("placeholder", encoding="utf-8")
    rule_path = create_link_replace_rule_workbook(
        [
            {
                "workbook_path": str(target),
                "workbook_name": "target.xlsx",
                "old_link_path": r"D:\old\old_tb.xlsx",
            }
        ],
        output_dir=str(tmp_path),
    )

    rules = read_link_replace_rules(rule_path)
    actions = build_link_replace_actions(rules, LinkReplaceSettings())

    assert len(rules) == 1
    assert actions[0].status == "跳过"
    assert actions[0].message == "未填写新链接路径"


def test_build_actions_respects_execute_no_and_groups_by_workbook(tmp_path):
    first = tmp_path / "first.xlsx"
    second = tmp_path / "second.xlsx"
    first.write_text("placeholder", encoding="utf-8")
    second.write_text("placeholder", encoding="utf-8")
    rule_path = create_link_replace_rule_workbook(
        [
            {"workbook_path": str(first), "workbook_name": "first.xlsx", "old_link_path": "old1.xlsx"},
            {"workbook_path": str(first), "workbook_name": "first.xlsx", "old_link_path": "old2.xlsx"},
            {"workbook_path": str(second), "workbook_name": "second.xlsx", "old_link_path": "old3.xlsx"},
        ],
        output_dir=str(tmp_path),
    )

    workbook = load_workbook(rule_path)
    sheet = workbook[RULE_SHEET_NAME]
    sheet["D2"] = "new1.xlsx"
    sheet["D3"] = "new2.xlsx"
    sheet["E3"] = "否"
    sheet["D4"] = "new3.xlsx"
    workbook.save(rule_path)
    workbook.close()

    actions = build_link_replace_actions(read_link_replace_rules(rule_path), LinkReplaceSettings())
    grouped = group_link_replace_actions_by_workbook_path(actions)

    assert [action.status for action in actions] == ["待执行", "跳过", "待执行"]
    assert actions[1].message == "是否执行为否，已跳过"
    assert len(grouped) == 2


def test_backup_path_adds_sequence_without_overwrite(tmp_path):
    target = tmp_path / "target.xlsx"
    target.write_text("placeholder", encoding="utf-8")
    first_backup = tmp_path / "target_批量换链接备份_20260101_010203.xlsx"
    first_backup.write_text("exists", encoding="utf-8")

    backup_path = build_backup_path(str(target), timestamp="20260101_010203")

    assert backup_path == str(tmp_path / "target_批量换链接备份_20260101_010203_2.xlsx")
    assert first_backup.read_text(encoding="utf-8") == "exists"


def test_write_results_to_rule_workbook(tmp_path):
    target = tmp_path / "target.xlsx"
    target.write_text("placeholder", encoding="utf-8")
    rule_path = create_link_replace_rule_workbook(
        [{"workbook_path": str(target), "workbook_name": "target.xlsx", "old_link_path": "old.xlsx"}],
        output_dir=str(tmp_path),
    )
    workbook = load_workbook(rule_path)
    workbook[RULE_SHEET_NAME]["D2"] = "new.xlsx"
    workbook.save(rule_path)
    workbook.close()

    actions = build_link_replace_actions(read_link_replace_rules(rule_path), LinkReplaceSettings())
    actions[0].status = "成功"
    actions[0].message = "已替换"

    write_link_replace_results_to_workbook(rule_path, actions)

    workbook = load_workbook(rule_path)
    rule_sheet = workbook[RULE_SHEET_NAME]
    log_sheet = workbook[LOG_SHEET_NAME]

    assert rule_sheet["F2"].value == "成功"
    assert rule_sheet["G2"].value == "已替换"
    assert log_sheet["E2"].value == "成功"
    assert log_sheet["F2"].value == "已替换"
    workbook.close()


def test_summarize_actions_counts_statuses(tmp_path):
    target = tmp_path / "target.xlsx"
    target.write_text("placeholder", encoding="utf-8")
    rule_path = create_link_replace_rule_workbook(
        [
            {"workbook_path": str(target), "workbook_name": "target.xlsx", "old_link_path": "old1.xlsx"},
            {"workbook_path": str(target), "workbook_name": "target.xlsx", "old_link_path": "old2.xlsx"},
            {"workbook_path": str(target), "workbook_name": "target.xlsx", "old_link_path": "old3.xlsx"},
        ],
        output_dir=str(tmp_path),
    )
    workbook = load_workbook(rule_path)
    sheet = workbook[RULE_SHEET_NAME]
    sheet["D2"] = "new1.xlsx"
    sheet["D3"] = "new2.xlsx"
    sheet["E3"] = "否"
    sheet["D4"] = "new3.xlsx"
    workbook.save(rule_path)
    workbook.close()

    actions = build_link_replace_actions(read_link_replace_rules(rule_path), LinkReplaceSettings())
    actions[0].status = "成功"
    actions[2].status = "失败"

    assert summarize_link_replace_actions(actions) == {
        "success_count": 1,
        "skipped_count": 1,
        "failed_count": 1,
        "total_count": 3,
    }

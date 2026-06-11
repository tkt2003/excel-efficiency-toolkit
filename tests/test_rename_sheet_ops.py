from pathlib import Path

from openpyxl import load_workbook

from src.excel_efficiency_toolkit.rename_sheet_ops import (
    LOG_SHEET_NAME,
    RULE_HEADERS,
    RULE_SHEET_NAME,
    SETTINGS_SHEET_NAME,
    SheetRenameRule,
    SheetRenameSettings,
    build_sheet_rename_plan,
    create_sheet_rename_rule_workbook,
    execute_sheet_rename_plan,
    group_sheet_rename_rules_by_workbook_path,
    read_sheet_rename_rules,
    read_sheet_rename_settings,
    write_sheet_rename_results_to_workbook,
)


class FakeWorksheet:
    def __init__(self, workbook, name):
        self._workbook = workbook
        self._name = name

    @property
    def Name(self):
        return self._name

    @Name.setter
    def Name(self, value):
        for sheet in self._workbook.Worksheets:
            if sheet is not self and sheet.Name.casefold() == value.casefold():
                raise ValueError(f"工作表名已存在：{value}")
        self._name = value


class FakeWorkbook:
    def __init__(self, sheet_names):
        self.Worksheets = []
        self.Worksheets.extend(FakeWorksheet(self, name) for name in sheet_names)


def test_create_sheet_rename_rule_workbook_writes_expected_sheets_and_rows(tmp_path):
    workbook_path = _create_file(tmp_path / "目标.xlsx")
    rule_path = create_sheet_rename_rule_workbook(
        workbook_infos=[
            {
                "workbook_path": str(workbook_path),
                "workbook_name": "目标.xlsx",
                "sheet_infos": [
                    {"name": "可见表", "is_hidden": False, "order": 1},
                    {"name": "隐藏表", "is_hidden": True, "order": 2},
                ],
            }
        ],
    )

    workbook = load_workbook(rule_path)
    try:
        assert workbook.sheetnames == ["使用说明", "参数设置", "重命名清单", "处理日志"]
        sheet = workbook[RULE_SHEET_NAME]
        assert [sheet.cell(row=1, column=i).value for i in range(1, 9)] == RULE_HEADERS
        assert sheet["A2"].value == str(workbook_path)
        assert sheet["B2"].value == "目标.xlsx"
        assert sheet["C2"].value == "可见表"
        assert sheet["D2"].value is None
        assert sheet["G2"].value == "否"
        assert sheet["H2"].value == 1
        assert sheet["C3"].value == "隐藏表"
        assert sheet["G3"].value == "是"
        assert sheet["H3"].value == 2
    finally:
        workbook.close()


def test_read_sheet_rename_settings_returns_defaults_and_falls_back_on_invalid_values(tmp_path):
    workbook_path = _create_file(tmp_path / "目标.xlsx")
    rule_path = create_sheet_rename_rule_workbook(
        workbook_infos=[_workbook_info(workbook_path, ["Sheet1"])],
    )

    settings = read_sheet_rename_settings(rule_path)
    assert settings.conflict_mode == "自动编号"
    assert settings.rename_hidden_sheets is False
    assert settings.clean_illegal_chars is True
    assert settings.too_long_mode == "自动截断"
    assert settings.skip_temp_files is True

    workbook = load_workbook(rule_path)
    sheet = workbook[SETTINGS_SHEET_NAME]
    sheet["B2"] = "覆盖"
    sheet["B3"] = "maybe"
    sheet["B4"] = "unknown"
    sheet["B5"] = "保留全部"
    sheet["B6"] = "随便"
    workbook.save(rule_path)
    workbook.close()

    settings = read_sheet_rename_settings(rule_path)
    assert settings.conflict_mode == "自动编号"
    assert settings.rename_hidden_sheets is False
    assert settings.clean_illegal_chars is True
    assert settings.too_long_mode == "自动截断"
    assert settings.skip_temp_files is True
    assert len(settings.warnings) == 5


def test_read_sheet_rename_rules_can_group_by_workbook_path(tmp_path):
    first_path = _create_file(tmp_path / "一.xlsx")
    second_path = _create_file(tmp_path / "二.xlsx")
    rule_path = create_sheet_rename_rule_workbook(
        workbook_infos=[
            _workbook_info(first_path, ["A"]),
            _workbook_info(second_path, ["A"]),
        ],
    )
    workbook = load_workbook(rule_path)
    sheet = workbook[RULE_SHEET_NAME]
    sheet["D2"] = "新A"
    sheet["D3"] = "新A"
    workbook.save(rule_path)
    workbook.close()

    rules = read_sheet_rename_rules(rule_path)
    grouped_rules = group_sheet_rename_rules_by_workbook_path(rules)

    assert len(rules) == 2
    assert len(grouped_rules) == 2
    assert {rule.workbook_name for rule in rules} == {"一.xlsx", "二.xlsx"}


def test_different_workbooks_same_sheet_names_do_not_conflict(tmp_path):
    first_path = _create_file(tmp_path / "一.xlsx")
    second_path = _create_file(tmp_path / "二.xlsx")
    first_rules = [_rule(first_path, "A", "同名", row_number=2)]
    second_rules = [_rule(second_path, "A", "同名", row_number=3)]

    first_plan = build_sheet_rename_plan(first_rules, ["A"], SheetRenameSettings())
    second_plan = build_sheet_rename_plan(second_rules, ["A"], SheetRenameSettings())

    assert first_plan[0].target_sheet_name == "同名"
    assert second_plan[0].target_sheet_name == "同名"


def test_illegal_sheet_name_chars_can_be_cleaned_or_skipped(tmp_path):
    workbook_path = _create_file(tmp_path / "目标.xlsx")
    rules = [_rule(workbook_path, "Sheet1", "新:名字")]

    plan = build_sheet_rename_plan(rules, ["Sheet1"], SheetRenameSettings(clean_illegal_chars=True))
    assert plan[0].target_sheet_name == "新_名字"
    assert "已清洗非法字符" in plan[0].message

    plan = build_sheet_rename_plan(rules, ["Sheet1"], SheetRenameSettings(clean_illegal_chars=False))
    assert plan[0].status == "跳过"
    assert plan[0].message == "工作表名包含非法字符"


def test_too_long_sheet_name_can_be_truncated_or_skipped(tmp_path):
    workbook_path = _create_file(tmp_path / "目标.xlsx")
    long_name = "超" * 40
    rules = [_rule(workbook_path, "Sheet1", long_name)]

    plan = build_sheet_rename_plan(rules, ["Sheet1"], SheetRenameSettings(too_long_mode="自动截断"))
    assert len(plan[0].target_sheet_name) == 31
    assert "已自动截断" in plan[0].message

    plan = build_sheet_rename_plan(rules, ["Sheet1"], SheetRenameSettings(too_long_mode="跳过"))
    assert plan[0].status == "跳过"
    assert plan[0].message == "工作表名超过 31 字符"


def test_external_target_conflict_auto_numbers_or_skips(tmp_path):
    workbook_path = _create_file(tmp_path / "目标.xlsx")
    rules = [_rule(workbook_path, "A", "C")]

    plan = build_sheet_rename_plan(
        rules,
        ["A", "C", "C_1"],
        SheetRenameSettings(conflict_mode="自动编号"),
    )
    assert plan[0].target_sheet_name == "C_2"
    assert "已自动编号" in plan[0].message

    plan = build_sheet_rename_plan(
        rules,
        ["A", "C"],
        SheetRenameSettings(conflict_mode="跳过"),
    )
    assert plan[0].status == "跳过"
    assert plan[0].message == "目标工作表名冲突"


def test_internal_target_conflict_auto_numbers_or_skips(tmp_path):
    workbook_path = _create_file(tmp_path / "目标.xlsx")
    rules = [
        _rule(workbook_path, "A", "C", row_number=2),
        _rule(workbook_path, "B", "C", row_number=3),
    ]

    plan = build_sheet_rename_plan(rules, ["A", "B"], SheetRenameSettings(conflict_mode="自动编号"))
    assert [action.target_sheet_name for action in plan] == ["C", "C_1"]

    plan = build_sheet_rename_plan(rules, ["A", "B"], SheetRenameSettings(conflict_mode="跳过"))
    assert plan[0].status == "成功"
    assert plan[1].status == "跳过"
    assert plan[1].message == "目标工作表名冲突"


def test_chain_rename_plan_allows_names_that_will_be_vacated(tmp_path):
    workbook_path = _create_file(tmp_path / "目标.xlsx")
    rules = [
        _rule(workbook_path, "Sheet1", "Sheet2", row_number=2),
        _rule(workbook_path, "Sheet2", "Sheet3", row_number=3),
        _rule(workbook_path, "Sheet3", "Sheet4", row_number=4),
    ]

    plan = build_sheet_rename_plan(rules, ["Sheet1", "Sheet2", "Sheet3"], SheetRenameSettings())

    assert [action.target_sheet_name for action in plan] == ["Sheet2", "Sheet3", "Sheet4"]


def test_missing_original_file_is_skipped(tmp_path):
    workbook_path = tmp_path / "missing.xlsx"

    plan = build_sheet_rename_plan(
        [_rule(workbook_path, "Sheet1", "目标")],
        ["Sheet1"],
        SheetRenameSettings(),
    )

    assert plan[0].status == "跳过"
    assert plan[0].message == "原文件不存在"


def test_missing_original_sheet_is_skipped(tmp_path):
    workbook_path = _create_file(tmp_path / "目标.xlsx")

    plan = build_sheet_rename_plan(
        [_rule(workbook_path, "Missing", "目标")],
        ["Sheet1"],
        SheetRenameSettings(),
    )

    assert plan[0].status == "跳过"
    assert plan[0].message == "原工作表不存在"


def test_hidden_sheet_is_skipped_by_default_and_allowed_by_setting(tmp_path):
    workbook_path = _create_file(tmp_path / "目标.xlsx")
    rules = [_rule(workbook_path, "隐藏表", "目标", is_hidden=True)]

    plan = build_sheet_rename_plan(rules, ["隐藏表"], SheetRenameSettings())
    assert plan[0].status == "跳过"
    assert plan[0].message == "隐藏工作表已跳过"

    plan = build_sheet_rename_plan(rules, ["隐藏表"], SheetRenameSettings(rename_hidden_sheets=True))
    assert plan[0].status == "成功"
    assert plan[0].target_sheet_name == "目标"


def test_same_sheet_name_is_skipped(tmp_path):
    workbook_path = _create_file(tmp_path / "目标.xlsx")

    plan = build_sheet_rename_plan(
        [_rule(workbook_path, "Sheet1", "Sheet1")],
        ["Sheet1"],
        SheetRenameSettings(),
    )

    assert plan[0].status == "跳过"
    assert plan[0].message == "新旧工作表名相同"


def test_execute_sheet_rename_plan_uses_fake_workbook_and_keeps_other_actions(tmp_path):
    workbook_path = _create_file(tmp_path / "目标.xlsx")
    workbook = FakeWorkbook(["A", "B"])
    actions = build_sheet_rename_plan(
        [
            _rule(workbook_path, "A", "新A", row_number=2),
            _rule(workbook_path, "Missing", "新Missing", row_number=3),
            _rule(workbook_path, "B", "新:B", row_number=4),
        ],
        ["A", "B"],
        SheetRenameSettings(clean_illegal_chars=False),
    )

    summary = execute_sheet_rename_plan(workbook, actions)

    assert summary == {"success_count": 1, "skipped_count": 2, "failed_count": 0}
    assert [sheet.Name for sheet in workbook.Worksheets] == ["新A", "B"]
    assert actions[0].message == "已重命名"
    assert actions[1].message == "原工作表不存在"
    assert actions[2].message == "工作表名包含非法字符"


def test_execute_sheet_rename_plan_supports_chain_rename_without_numbering(tmp_path):
    workbook_path = _create_file(tmp_path / "目标.xlsx")
    workbook = FakeWorkbook(["Sheet1", "Sheet2", "Sheet3"])
    actions = build_sheet_rename_plan(
        [
            _rule(workbook_path, "Sheet1", "Sheet2", row_number=2),
            _rule(workbook_path, "Sheet2", "Sheet3", row_number=3),
            _rule(workbook_path, "Sheet3", "Sheet4", row_number=4),
        ],
        ["Sheet1", "Sheet2", "Sheet3"],
        SheetRenameSettings(),
    )

    summary = execute_sheet_rename_plan(workbook, actions)

    assert summary == {"success_count": 3, "skipped_count": 0, "failed_count": 0}
    assert [sheet.Name for sheet in workbook.Worksheets] == ["Sheet2", "Sheet3", "Sheet4"]
    assert not any(sheet.Name.startswith("__tmp_rename_") for sheet in workbook.Worksheets)


def test_execute_sheet_rename_plan_supports_swap_rename(tmp_path):
    workbook_path = _create_file(tmp_path / "目标.xlsx")
    workbook = FakeWorkbook(["A", "B"])
    actions = build_sheet_rename_plan(
        [
            _rule(workbook_path, "A", "B", row_number=2),
            _rule(workbook_path, "B", "A", row_number=3),
        ],
        ["A", "B"],
        SheetRenameSettings(),
    )

    summary = execute_sheet_rename_plan(workbook, actions)

    assert summary == {"success_count": 2, "skipped_count": 0, "failed_count": 0}
    assert [sheet.Name for sheet in workbook.Worksheets] == ["B", "A"]
    assert not any(sheet.Name.startswith("__tmp_rename_") for sheet in workbook.Worksheets)


def test_write_sheet_rename_results_updates_status_columns_and_log_sheet(tmp_path):
    workbook_path = _create_file(tmp_path / "目标.xlsx")
    rule_path = create_sheet_rename_rule_workbook(
        workbook_infos=[_workbook_info(workbook_path, ["Sheet1"])],
    )
    actions = build_sheet_rename_plan(
        [_rule(workbook_path, "Sheet1", "新表")],
        ["Sheet1"],
        SheetRenameSettings(),
    )
    execute_sheet_rename_plan(FakeWorkbook(["Sheet1"]), actions)

    write_sheet_rename_results_to_workbook(rule_path, actions)

    workbook = load_workbook(rule_path)
    try:
        rule_sheet = workbook[RULE_SHEET_NAME]
        log_sheet = workbook[LOG_SHEET_NAME]
        assert rule_sheet["E2"].value == "成功"
        assert rule_sheet["F2"].value == "已重命名"
        assert log_sheet.max_row == 2
        assert log_sheet["B2"].value == str(workbook_path)
        assert log_sheet["C2"].value == "目标.xlsx"
        assert log_sheet["D2"].value == "Sheet1"
        assert log_sheet["E2"].value == "新表"
        assert log_sheet["F2"].value == "成功"
        assert log_sheet["G2"].value == "已重命名"
    finally:
        workbook.close()


def _create_file(path: Path, content: str = "data") -> Path:
    path.write_text(content, encoding="utf-8")
    return path.resolve()


def _workbook_info(workbook_path: Path, sheet_names):
    return {
        "workbook_path": str(workbook_path),
        "workbook_name": workbook_path.name,
        "sheet_infos": [
            {"name": sheet_name, "is_hidden": False, "order": index}
            for index, sheet_name in enumerate(sheet_names, start=1)
        ],
    }


def _rule(workbook_path, original_name, new_name, row_number=2, is_hidden=False):
    workbook_path = Path(workbook_path)
    return SheetRenameRule(
        row_number=row_number,
        workbook_path=str(workbook_path),
        workbook_name=workbook_path.name,
        original_sheet_name=original_name,
        new_sheet_name_raw=new_name,
        is_hidden=is_hidden,
        original_order=row_number - 1,
    )

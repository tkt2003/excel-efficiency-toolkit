import os

import pytest
from openpyxl import load_workbook

from src.excel_efficiency_toolkit.template_tb_report_ops import (
    LOG_HEADERS,
    LOG_WORKBOOK_NAME,
    OUTPUT_STEM_SUFFIX,
    TemplateTbReportRecord,
    build_output_filename,
    build_tb_report_plan,
    build_unique_output_path,
    is_office_temp_file,
    is_supported_tb_file,
    is_supported_template_file,
    summarize_tb_report_records,
    write_tb_report_log_workbook,
)


def _make_template(tmp_path, name="template.xlsx"):
    template_path = tmp_path / name
    template_path.write_text("placeholder", encoding="utf-8")
    return str(template_path)


def _make_tb_file(tmp_path, name):
    path = tmp_path / name
    path.write_text("placeholder", encoding="utf-8")
    return str(path)


def test_template_file_detection_accepts_xlsx_xlsm_and_skips_others():
    assert is_supported_template_file("template.xlsx")
    assert is_supported_template_file("TEMPLATE.XLSM")
    assert not is_supported_template_file("template.xls")
    assert not is_supported_template_file("template.csv")
    assert not is_supported_template_file("~$template.xlsx")


def test_tb_file_detection_accepts_xlsx_xlsm_only():
    assert is_supported_tb_file("tb.xlsx")
    assert is_supported_tb_file("TB.XLSM")
    assert not is_supported_tb_file("tb.xls")
    assert not is_supported_tb_file("tb.csv")
    assert not is_supported_tb_file("~$tb.xlsx")


def test_office_temp_file_detection():
    assert is_office_temp_file(r"D:\tmp\~$tb.xlsx")
    assert not is_office_temp_file(r"D:\tmp\tb.xlsx")


def test_build_output_filename_uses_tb_stem_and_template_suffix():
    assert build_output_filename(r"D:\path\子公司1.xlsx", ".xlsx") == f"子公司1{OUTPUT_STEM_SUFFIX}.xlsx"
    assert build_output_filename(r"D:\path\子公司1.xlsx", ".xlsm") == f"子公司1{OUTPUT_STEM_SUFFIX}.xlsm"
    assert build_output_filename("子公司1TB.xlsx", ".xlsx") == f"子公司1TB{OUTPUT_STEM_SUFFIX}.xlsx"


def test_build_unique_output_path_adds_sequence_without_overwrite(tmp_path):
    first_path = tmp_path / f"子公司1{OUTPUT_STEM_SUFFIX}.xlsx"
    second_path = tmp_path / f"子公司1{OUTPUT_STEM_SUFFIX}_2.xlsx"
    first_path.write_text("exists", encoding="utf-8")
    second_path.write_text("exists", encoding="utf-8")

    output_path = build_unique_output_path(str(tmp_path), f"子公司1{OUTPUT_STEM_SUFFIX}.xlsx")

    assert output_path == str(tmp_path / f"子公司1{OUTPUT_STEM_SUFFIX}_3.xlsx")
    assert first_path.read_text(encoding="utf-8") == "exists"
    assert second_path.read_text(encoding="utf-8") == "exists"


def test_build_unique_output_path_respects_reserved_paths(tmp_path):
    target_name = f"子公司1{OUTPUT_STEM_SUFFIX}.xlsx"
    reserved = {str(tmp_path / target_name)}

    output_path = build_unique_output_path(str(tmp_path), target_name, reserved_paths=reserved)

    assert output_path == str(tmp_path / f"子公司1{OUTPUT_STEM_SUFFIX}_2.xlsx")


def test_build_plan_requires_template_path(tmp_path):
    with pytest.raises(ValueError, match="模板"):
        build_tb_report_plan(
            template_path="",
            tb_paths=[_make_tb_file(tmp_path, "tb.xlsx")],
            output_dir=str(tmp_path),
            old_link_path=r"D:\old\old_tb.xlsx",
        )


def test_build_plan_requires_supported_template_extension(tmp_path):
    template_path = tmp_path / "template.xls"
    template_path.write_text("placeholder", encoding="utf-8")

    with pytest.raises(ValueError, match="模板工作簿"):
        build_tb_report_plan(
            template_path=str(template_path),
            tb_paths=[_make_tb_file(tmp_path, "tb.xlsx")],
            output_dir=str(tmp_path),
            old_link_path=r"D:\old\old_tb.xlsx",
        )


def test_build_plan_requires_old_link_path(tmp_path):
    template_path = _make_template(tmp_path)
    with pytest.raises(ValueError, match="旧链接"):
        build_tb_report_plan(
            template_path=template_path,
            tb_paths=[_make_tb_file(tmp_path, "tb.xlsx")],
            output_dir=str(tmp_path),
            old_link_path="",
        )


def test_build_plan_requires_at_least_one_tb_file(tmp_path):
    template_path = _make_template(tmp_path)
    with pytest.raises(ValueError, match="TB"):
        build_tb_report_plan(
            template_path=template_path,
            tb_paths=[],
            output_dir=str(tmp_path),
            old_link_path=r"D:\old\old_tb.xlsx",
        )


def test_build_plan_skips_temp_and_unsupported_files_and_keeps_xlsm_extension(tmp_path):
    template_path = _make_template(tmp_path, name="template.xlsm")
    tb_xlsx = _make_tb_file(tmp_path, "tb_子公司1.xlsx")
    tb_xlsm = _make_tb_file(tmp_path, "tb_子公司2.xlsm")
    tb_temp = _make_tb_file(tmp_path, "~$tb_子公司3.xlsx")
    tb_xls = _make_tb_file(tmp_path, "tb_子公司4.xls")
    tb_csv = _make_tb_file(tmp_path, "tb_子公司5.csv")

    plan = build_tb_report_plan(
        template_path=template_path,
        tb_paths=[tb_xlsx, tb_xlsm, tb_temp, tb_xls, tb_csv],
        output_dir=str(tmp_path),
        old_link_path=r"D:\old\old_tb.xlsx",
    )

    assert plan.template_suffix == ".xlsm"
    assert plan.old_link_path == r"D:\old\old_tb.xlsx"
    assert [record.status for record in plan.records] == [
        "待执行",
        "待执行",
        "跳过",
        "跳过",
        "跳过",
    ]
    assert plan.records[0].output_name == f"tb_子公司1{OUTPUT_STEM_SUFFIX}.xlsm"
    assert plan.records[1].output_name == f"tb_子公司2{OUTPUT_STEM_SUFFIX}.xlsm"
    assert plan.records[2].message == "临时文件已跳过"
    assert "仅支持" in plan.records[3].message
    assert "仅支持" in plan.records[4].message
    assert plan.records[0].new_link_path == os.path.abspath(tb_xlsx)


def test_build_plan_auto_numbers_existing_output_files(tmp_path):
    template_path = _make_template(tmp_path)
    tb_path = _make_tb_file(tmp_path, "tb_子公司1.xlsx")
    existing_output = tmp_path / f"tb_子公司1{OUTPUT_STEM_SUFFIX}.xlsx"
    existing_output.write_text("exists", encoding="utf-8")

    plan = build_tb_report_plan(
        template_path=template_path,
        tb_paths=[tb_path],
        output_dir=str(tmp_path),
        old_link_path=r"D:\old\old_tb.xlsx",
    )

    assert plan.records[0].output_name == f"tb_子公司1{OUTPUT_STEM_SUFFIX}_2.xlsx"
    assert existing_output.read_text(encoding="utf-8") == "exists"


def test_build_plan_handles_duplicate_tb_names_with_unique_outputs(tmp_path):
    template_path = _make_template(tmp_path)
    sub_dir = tmp_path / "sub"
    sub_dir.mkdir()
    tb_first = _make_tb_file(tmp_path, "tb_子公司1.xlsx")
    tb_second = _make_tb_file(sub_dir, "tb_子公司1.xlsx")

    plan = build_tb_report_plan(
        template_path=template_path,
        tb_paths=[tb_first, tb_second],
        output_dir=str(tmp_path),
        old_link_path=r"D:\old\old_tb.xlsx",
    )

    output_names = [record.output_name for record in plan.records]
    assert output_names[0] == f"tb_子公司1{OUTPUT_STEM_SUFFIX}.xlsx"
    assert output_names[1] == f"tb_子公司1{OUTPUT_STEM_SUFFIX}_2.xlsx"


def test_summarize_records_counts_statuses():
    records = [
        TemplateTbReportRecord(index=1, tb_path="a", tb_name="a.xlsx", status="成功"),
        TemplateTbReportRecord(index=2, tb_path="b", tb_name="b.xlsx", status="成功"),
        TemplateTbReportRecord(index=3, tb_path="c", tb_name="c.xlsx", status="跳过"),
        TemplateTbReportRecord(index=4, tb_path="d", tb_name="d.xlsx", status="失败"),
    ]

    assert summarize_tb_report_records(records) == {
        "total_count": 4,
        "success_count": 2,
        "skipped_count": 1,
        "failed_count": 1,
    }


def test_write_log_workbook_creates_unique_path_and_writes_records(tmp_path):
    template_path = _make_template(tmp_path)
    records = [
        TemplateTbReportRecord(
            index=1,
            tb_path=str(tmp_path / "tb_子公司1.xlsx"),
            tb_name="tb_子公司1.xlsx",
            output_name=f"tb_子公司1{OUTPUT_STEM_SUFFIX}.xlsx",
            output_path=str(tmp_path / f"tb_子公司1{OUTPUT_STEM_SUFFIX}.xlsx"),
            old_link_path=r"D:\old\old_tb.xlsx",
            new_link_path=str(tmp_path / "tb_子公司1.xlsx"),
            status="成功",
            message="已生成并替换链接",
        ),
        TemplateTbReportRecord(
            index=2,
            tb_path=str(tmp_path / "~$tb_x.xlsx"),
            tb_name="~$tb_x.xlsx",
            old_link_path=r"D:\old\old_tb.xlsx",
            new_link_path=str(tmp_path / "~$tb_x.xlsx"),
            status="跳过",
            message="临时文件已跳过",
        ),
    ]

    first_log = write_tb_report_log_workbook(
        records=records,
        output_dir=str(tmp_path),
        template_path=template_path,
        old_link_path=r"D:\old\old_tb.xlsx",
    )
    second_log = write_tb_report_log_workbook(
        records=records,
        output_dir=str(tmp_path),
        template_path=template_path,
        old_link_path=r"D:\old\old_tb.xlsx",
    )

    assert os.path.basename(first_log) == LOG_WORKBOOK_NAME
    assert os.path.basename(second_log) != LOG_WORKBOOK_NAME
    assert first_log != second_log

    workbook = load_workbook(first_log)
    try:
        sheet = workbook.active
        header_values = [sheet.cell(row=3, column=column).value for column in range(1, len(LOG_HEADERS) + 1)]
        assert header_values == LOG_HEADERS
        assert sheet.cell(row=4, column=1).value == 1
        assert sheet.cell(row=4, column=2).value == "tb_子公司1.xlsx"
        assert sheet.cell(row=4, column=8).value == "成功"
        assert sheet.cell(row=5, column=8).value == "跳过"
    finally:
        workbook.close()

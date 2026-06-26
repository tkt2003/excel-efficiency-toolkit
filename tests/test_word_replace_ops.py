import os
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from src.excel_efficiency_toolkit.word_replace_ops import (
    COMMON_RULE_HEADERS,
    MODE_COMMON,
    PER_FILE_RULE_HEADERS,
    WordReplaceAction,
    WordReplaceFileLog,
    _build_execution_result,
    _build_preview_result,
    build_common_replace_rules_from_entries,
    build_word_backup_path,
    create_per_file_replace_rule_template,
    load_common_replace_rules,
    load_per_file_replace_rules,
    parse_rule_enabled,
)


def test_load_common_replace_rules_reads_enabled_rows_and_allows_empty_replace(tmp_path):
    rule_path = tmp_path / "common_rules.xlsx"
    _write_rows(
        rule_path,
        [
            COMMON_RULE_HEADERS,
            ["RRR", "替换内容1", "是"],
            ["SSS", "替换内容2", ""],
            ["清空", "", "Y"],
            ["禁用", "不应读取", "否"],
        ],
    )

    rules = load_common_replace_rules(str(rule_path))

    assert [rule.find_text for rule in rules] == ["RRR", "SSS", "清空"]
    assert rules[2].replace_text == ""


def test_load_per_file_replace_rules_groups_by_full_word_path(tmp_path):
    first_path = tmp_path / "B公司附注.docx"
    second_path = tmp_path / "C公司附注.docx"
    rule_path = tmp_path / "per_file_rules.xlsx"
    _write_rows(
        rule_path,
        [
            PER_FILE_RULE_HEADERS,
            [str(first_path), "A公司", "B公司", "是"],
            [str(second_path), "A公司", "C公司", "TRUE"],
            [str(second_path), "禁用", "不应读取", "N"],
        ],
    )

    rules_by_file = load_per_file_replace_rules(str(rule_path))

    first_key = os.path.normcase(str(first_path.resolve()))
    second_key = os.path.normcase(str(second_path.resolve()))
    assert rules_by_file[first_key][0].replace_text == "B公司"
    assert rules_by_file[second_key][0].replace_text == "C公司"
    assert len(rules_by_file[second_key]) == 1


def test_empty_find_text_is_rejected(tmp_path):
    rule_path = tmp_path / "bad_rules.xlsx"
    _write_rows(
        rule_path,
        [
            COMMON_RULE_HEADERS,
            ["", "替换", "是"],
        ],
    )

    with pytest.raises(ValueError, match="查找内容为空"):
        load_common_replace_rules(str(rule_path))


def test_enabled_field_parses_supported_values():
    assert parse_rule_enabled("")
    assert parse_rule_enabled(None)
    assert parse_rule_enabled("是")
    assert parse_rule_enabled("Y")
    assert parse_rule_enabled("TRUE")
    assert not parse_rule_enabled("否")
    assert not parse_rule_enabled("N")
    assert not parse_rule_enabled("FALSE")


def test_build_common_replace_rules_from_entries_ignores_blank_rows_and_rejects_invalid_rows():
    rules = build_common_replace_rules_from_entries(
        [
            ("A公司", "B公司"),
            ("", ""),
            ("2024年", "2025年"),
            ("清空", ""),
        ]
    )

    assert [rule.find_text for rule in rules] == ["A公司", "2024年", "清空"]
    assert rules[2].replace_text == ""

    with pytest.raises(ValueError, match="查找内容为空"):
        build_common_replace_rules_from_entries([("", "B公司")])

    with pytest.raises(ValueError, match="请至少填写一条替换规则"):
        build_common_replace_rules_from_entries([("", ""), (" ", " ")])


def test_create_per_file_replace_rule_template_prefills_word_paths(tmp_path):
    first_path = tmp_path / "B公司附注.docx"
    second_path = tmp_path / "C公司附注.docx"

    template_path = create_per_file_replace_rule_template([str(first_path), str(second_path)])

    workbook = load_workbook(template_path)
    try:
        sheet = workbook.active
        assert [sheet.cell(row=1, column=column).value for column in range(1, 5)] == PER_FILE_RULE_HEADERS
        assert sheet.cell(row=2, column=1).value == str(first_path.resolve())
        assert sheet.cell(row=2, column=2).value is None
        assert sheet.cell(row=2, column=3).value is None
        assert sheet.cell(row=2, column=4).value == "是"
        assert sheet.cell(row=3, column=1).value == str(second_path.resolve())
    finally:
        workbook.close()


def test_backup_path_uses_batch_folder_and_does_not_overwrite(tmp_path):
    target_path = tmp_path / "target.docx"
    target_path.write_text("placeholder", encoding="utf-8")
    backup_dir = tmp_path / "老头表格助手备份" / "word_replace_20260101_010203"
    backup_dir.mkdir(parents=True)
    existing_backup = backup_dir / "target.docx"
    existing_backup.write_text("exists", encoding="utf-8")

    backup_path = build_word_backup_path(str(target_path), timestamp="20260101_010203")

    assert backup_path == str(backup_dir / "target_2.docx")
    assert existing_backup.read_text(encoding="utf-8") == "exists"


def test_preview_and_execution_result_structures_have_expected_fields(tmp_path):
    word_path = tmp_path / "target.docx"
    action = WordReplaceAction(
        file_path=str(word_path),
        find_text="A公司",
        replace_text="B公司",
        rule_row_number=2,
        status="待执行",
        estimated_count=3,
    )
    preview_result = _build_preview_result(MODE_COMMON, [action])

    assert {
        "file_path",
        "find_text",
        "replace_text",
        "estimated_count",
        "status",
        "message",
        "rule_row_number",
    } <= set(preview_result["items"][0])

    file_log = WordReplaceFileLog(
        file_path=str(word_path),
        file_name=word_path.name,
        rule_count=1,
        replace_count=3,
        status="成功",
        message="已保存",
        backup_path=str(tmp_path / "backup.docx"),
        backup_status="已备份",
    )
    execution_result = _build_execution_result([action], [file_log])

    assert {
        "file_path",
        "file_name",
        "rule_count",
        "replace_count",
        "status",
        "message",
        "backup_path",
        "backup_status",
    } <= set(execution_result["file_logs"][0])
    assert {
        "success_file_count",
        "skipped_file_count",
        "failed_file_count",
        "replace_count",
    } <= set(execution_result["summary"])


def _write_rows(path: Path, rows: list[list[object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()

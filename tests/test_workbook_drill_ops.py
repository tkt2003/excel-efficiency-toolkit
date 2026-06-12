from openpyxl import Workbook

from src.excel_efficiency_toolkit.workbook_drill_ops import (
    build_multi_file_result_headers,
    build_single_file_result_headers,
    build_unique_result_sheet_name,
    expand_range_addresses,
    range_value_to_address_map,
    should_skip_history_result_sheet,
    stringify_excel_error_value,
    write_multi_file_result_sheet,
    write_single_file_result_sheet,
    write_single_workbook_drill_result_to_com_sheet,
)


def test_expand_range_addresses_single_cell():
    assert expand_range_addresses("C5") == ["C5"]


def test_expand_range_addresses_area():
    assert expand_range_addresses("B2:C3") == ["B2", "C2", "B3", "C3"]


def test_build_unique_result_sheet_name_adds_sequence():
    assert (
        build_unique_result_sheet_name(["Sheet1", "数据穿透查询结果", "数据穿透查询结果_2"])
        == "数据穿透查询结果_3"
    )


def test_should_skip_history_result_sheet():
    assert should_skip_history_result_sheet("数据穿透查询结果")
    assert should_skip_history_result_sheet("数据穿透查询结果_2")
    assert not should_skip_history_result_sheet("数据穿透结果")


def test_stringify_excel_error_value_keeps_readable_error_text():
    assert stringify_excel_error_value("#VALUE!") == "#VALUE!"
    assert stringify_excel_error_value(2042) == "#N/A"
    assert stringify_excel_error_value("普通文本") == "普通文本"


def test_single_file_headers_and_sheet_write():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据穿透查询结果"

    write_single_file_result_sheet(
        sheet=sheet,
        base_sheet_name="货币资金",
        range_address="B2:C3",
        created_at_text="2026-06-12 10:11:12",
        records=[
            {
                "sheet_name": "Sheet1",
                "visible_status": "可见",
                "values_by_address": {"B2": 1, "C2": 2, "B3": 3, "C3": "#N/A"},
            }
        ],
    )

    assert sheet["A1"].value == "基准工作表：货币资金    基准选区：B2:C3    生成时间：2026-06-12 10:11:12"
    assert [sheet.cell(row=3, column=index).value for index in range(1, 7)] == [
        "工作表名",
        "工作表可见状态",
        "B2",
        "C2",
        "B3",
        "C3",
    ]
    assert [sheet.cell(row=4, column=index).value for index in range(1, 7)] == [
        "Sheet1",
        "可见",
        1,
        2,
        3,
        "#N/A",
    ]


def test_multi_file_headers_and_sheet_write():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据穿透查询结果"

    write_multi_file_result_sheet(
        sheet=sheet,
        records=[
            {
                "source_file_name": "a.xlsx",
                "source_file_path": r"D:\tmp\a.xlsx",
                "target_sheet_name": "货币资金",
                "values_by_address": {"B2": 1, "C2": 2, "B3": 3, "C3": 4},
                "status": "成功",
                "message": "读取成功",
            }
        ],
        range_address="B2:C3",
    )

    assert [cell.value for cell in sheet[1]] == [
        "源文件名",
        "源文件路径",
        "工作表名",
        "B2",
        "C2",
        "B3",
        "C3",
        "状态",
        "说明",
    ]
    assert [sheet.cell(row=2, column=index).value for index in range(1, 10)] == [
        "a.xlsx",
        r"D:\tmp\a.xlsx",
        "货币资金",
        1,
        2,
        3,
        4,
        "成功",
        "读取成功",
    ]


def test_single_workbook_result_can_write_to_com_sheet_like_object():
    sheet = _FakeComSheet()

    write_single_workbook_drill_result_to_com_sheet(
        sheet=sheet,
        base_sheet_name="Sheet1",
        range_address="C3:D4",
        created_at_text="2026-06-12 12:00:00",
        records=[
            {
                "sheet_name": "Sheet1",
                "visible_status": "可见",
                "values_by_address": {"C3": 1, "D3": 2, "C4": 3, "D4": 4},
            }
        ],
    )

    assert sheet.values[(1, 1)] == "基准工作表：Sheet1    基准选区：C3:D4    生成时间：2026-06-12 12:00:00"
    assert [sheet.values[(3, index)] for index in range(1, 7)] == [
        "工作表名",
        "工作表可见状态",
        "C3",
        "D3",
        "C4",
        "D4",
    ]
    assert [sheet.values[(4, index)] for index in range(1, 7)] == [
        "Sheet1",
        "可见",
        1,
        2,
        3,
        4,
    ]


def test_range_value_to_address_map_supports_single_and_area():
    assert range_value_to_address_map("C5", 8) == {"C5": 8}
    assert range_value_to_address_map("B2:C3", ((1, 2), (3, 4))) == {
        "B2": 1,
        "C2": 2,
        "B3": 3,
        "C3": 4,
    }


def test_build_headers_helpers():
    assert build_single_file_result_headers("C5") == ["工作表名", "工作表可见状态", "C5"]
    assert build_multi_file_result_headers("C5") == ["源文件名", "源文件路径", "工作表名", "C5", "状态", "说明"]


class _FakeComCell:
    def __init__(self, values, row_index, column_index):
        self._values = values
        self._key = (row_index, column_index)

    @property
    def Value(self):
        return self._values.get(self._key)

    @Value.setter
    def Value(self, value):
        self._values[self._key] = value


class _FakeComSheet:
    def __init__(self):
        self.values = {}

    def Cells(self, row_index, column_index):
        return _FakeComCell(self.values, row_index, column_index)

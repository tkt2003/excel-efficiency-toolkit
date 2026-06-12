from pathlib import Path

from openpyxl import load_workbook

from src.excel_efficiency_toolkit.rename_file_ops import (
    LOG_SHEET_NAME,
    RULE_HEADERS,
    RULE_SHEET_NAME,
    SETTINGS_SHEET_NAME,
    RenameRule,
    RenameSettings,
    build_rename_plan,
    create_rename_rule_workbook,
    execute_rename_plan,
    read_rename_settings,
    write_rename_results_to_workbook,
)


def _create_file(path: Path, content: str = "data") -> Path:
    path.write_text(content, encoding="utf-8")
    return path


def test_create_rename_rule_workbook_writes_expected_sheets_and_rows(tmp_path):
    source_path = _create_file(tmp_path / "原文件.xlsx")

    rule_path = create_rename_rule_workbook([str(source_path)])

    workbook = load_workbook(rule_path)
    try:
        assert workbook.sheetnames == ["使用说明", "参数设置", "重命名清单", "处理日志"]
        sheet = workbook[RULE_SHEET_NAME]
        assert [sheet.cell(row=1, column=i).value for i in range(1, 9)] == RULE_HEADERS
        assert sheet["A2"].value == str(source_path.resolve())
        assert sheet["B2"].value == "原文件"
        assert sheet["C2"].value is None
        assert sheet["D2"].value == ".xlsx"
        assert sheet["H2"].value == ".xlsx"
    finally:
        workbook.close()


def test_read_rename_settings_returns_defaults_and_falls_back_on_invalid_values(tmp_path):
    source_path = _create_file(tmp_path / "a.xlsx")
    rule_path = create_rename_rule_workbook([str(source_path)])

    settings = read_rename_settings(rule_path)
    assert settings.existing_target_mode == "自动编号"
    assert settings.skip_temp_files is True
    assert settings.clean_illegal_chars is True

    workbook = load_workbook(rule_path)
    sheet = workbook[SETTINGS_SHEET_NAME]
    sheet["B2"] = "覆盖"
    sheet["B3"] = "maybe"
    sheet["B4"] = "unknown"
    workbook.save(rule_path)
    workbook.close()

    settings = read_rename_settings(rule_path)
    assert settings.existing_target_mode == "自动编号"
    assert settings.skip_temp_files is True
    assert settings.clean_illegal_chars is True
    assert len(settings.warnings) == 3


def test_extension_rules_use_d_column_add_dot_and_split_suffix_from_new_name(tmp_path):
    source_path = _create_file(tmp_path / "old.xlsx")

    plan = build_rename_plan(
        [_rule(source_path, new_name="新名字", extension=".xlsx")],
        RenameSettings(),
    )
    assert Path(plan[0].target_path).name == "新名字.xlsx"

    plan = build_rename_plan(
        [_rule(source_path, new_name="新名字", extension="xlsm")],
        RenameSettings(),
    )
    assert Path(plan[0].target_path).suffix == ".xlsm"
    assert Path(plan[0].target_path).name == "新名字.xlsm"

    plan = build_rename_plan(
        [_rule(source_path, new_name="新名字.xlsm", extension=".xlsx")],
        RenameSettings(),
    )
    assert Path(plan[0].target_path).name == "新名字.xlsm"
    assert "已从新文件名中识别后缀" in plan[0].message


def test_suffix_column_rules_cover_expected_cases(tmp_path):
    xlsx_path = _create_file(tmp_path / "1.xlsx")
    xlsm_path = _create_file(tmp_path / "1.xlsm", "m")

    cases = [
        (xlsx_path, "2", ".xlsx", "2.xlsx"),
        (xlsm_path, "2", ".xlsm", "2.xlsm"),
        (xlsx_path, "2", ".xls", "2.xls"),
        (xlsx_path, "2.xlsx", ".xlsx", "2.xlsx"),
        (xlsx_path, "2.xlsm", ".xlsx", "2.xlsm"),
        (xlsx_path, "2", "", "2.xlsx"),
    ]

    for index, (source_path, new_name, extension, expected_name) in enumerate(cases, start=2):
        plan = build_rename_plan(
            [_rule(source_path, new_name=new_name, extension=extension, row_number=index)],
            RenameSettings(),
        )
        assert Path(plan[0].target_path).name == expected_name


def test_illegal_filename_chars_can_be_cleaned_or_skipped(tmp_path):
    source_path = _create_file(tmp_path / "old.xlsx")

    plan = build_rename_plan(
        [_rule(source_path, new_name="新:名字", extension=".xlsx")],
        RenameSettings(clean_illegal_chars=True),
    )
    assert Path(plan[0].target_path).name == "新_名字.xlsx"
    assert "已清洗非法字符" in plan[0].message

    plan = build_rename_plan(
        [_rule(source_path, new_name="新:名字", extension=".xlsx")],
        RenameSettings(clean_illegal_chars=False),
    )
    assert plan[0].status == "跳过"
    assert plan[0].message == "文件名包含非法字符"


def test_existing_target_auto_number_avoids_overwrite_and_skip_mode_skips(tmp_path):
    source_path = _create_file(tmp_path / "old.xlsx")
    _create_file(tmp_path / "目标.xlsx", "existing")

    plan = build_rename_plan(
        [_rule(source_path, new_name="目标", extension=".xlsx")],
        RenameSettings(existing_target_mode="自动编号"),
    )
    assert Path(plan[0].target_path).name == "目标_1.xlsx"

    plan = build_rename_plan(
        [_rule(source_path, new_name="目标", extension=".xlsx")],
        RenameSettings(existing_target_mode="跳过"),
    )
    assert plan[0].status == "跳过"
    assert plan[0].message == "目标文件已存在"


def test_internal_target_conflict_auto_numbers_or_skips(tmp_path):
    first_path = _create_file(tmp_path / "a.xlsx")
    second_path = _create_file(tmp_path / "b.xlsx")
    rules = [
        _rule(first_path, new_name="目标", extension=".xlsx", row_number=2),
        _rule(second_path, new_name="目标", extension=".xlsx", row_number=3),
    ]

    plan = build_rename_plan(rules, RenameSettings(existing_target_mode="自动编号"))
    assert [Path(action.target_path).name for action in plan] == ["目标.xlsx", "目标_1.xlsx"]

    plan = build_rename_plan(rules, RenameSettings(existing_target_mode="跳过"))
    assert plan[0].status == "成功"
    assert plan[1].status == "跳过"
    assert plan[1].message == "目标文件名冲突"


def test_chain_rename_plan_allows_targets_that_will_be_vacated(tmp_path):
    first_path = _create_file(tmp_path / "1.xlsx")
    second_path = _create_file(tmp_path / "2.xlsx")
    third_path = _create_file(tmp_path / "3.xlsx")
    rules = [
        _rule(first_path, new_name="2", extension=".xlsx", row_number=2),
        _rule(second_path, new_name="3", extension=".xlsx", row_number=3),
        _rule(third_path, new_name="4", extension=".xlsx", row_number=4),
    ]

    plan = build_rename_plan(rules, RenameSettings(existing_target_mode="自动编号"))

    assert [Path(action.target_path).name for action in plan] == ["2.xlsx", "3.xlsx", "4.xlsx"]


def test_execute_rename_plan_supports_chain_rename_without_numbering(tmp_path):
    first_path = _create_file(tmp_path / "1.xlsx")
    second_path = _create_file(tmp_path / "2.xlsx")
    third_path = _create_file(tmp_path / "3.xlsx")
    actions = build_rename_plan(
        [
            _rule(first_path, new_name="2", extension=".xlsx", row_number=2),
            _rule(second_path, new_name="3", extension=".xlsx", row_number=3),
            _rule(third_path, new_name="4", extension=".xlsx", row_number=4),
        ],
        RenameSettings(),
    )

    summary = execute_rename_plan(actions)

    assert summary == {"success_count": 3, "skipped_count": 0, "failed_count": 0}
    assert not (tmp_path / "1.xlsx").exists()
    assert (tmp_path / "2.xlsx").exists()
    assert (tmp_path / "3.xlsx").exists()
    assert (tmp_path / "4.xlsx").exists()
    assert not list(tmp_path.glob(".__rename_tmp_*"))


def test_execute_rename_plan_supports_swap_rename(tmp_path):
    first_path = _create_file(tmp_path / "A.xlsx", "a")
    second_path = _create_file(tmp_path / "B.xlsx", "b")
    actions = build_rename_plan(
        [
            _rule(first_path, new_name="B", extension=".xlsx", row_number=2),
            _rule(second_path, new_name="A", extension=".xlsx", row_number=3),
        ],
        RenameSettings(),
    )

    summary = execute_rename_plan(actions)

    assert summary == {"success_count": 2, "skipped_count": 0, "failed_count": 0}
    assert (tmp_path / "A.xlsx").read_text(encoding="utf-8") == "b"
    assert (tmp_path / "B.xlsx").read_text(encoding="utf-8") == "a"
    assert not list(tmp_path.glob(".__rename_tmp_*"))


def test_missing_original_file_is_skipped(tmp_path):
    source_path = tmp_path / "missing.xlsx"

    plan = build_rename_plan(
        [_rule(source_path, new_name="目标", extension=".xlsx")],
        RenameSettings(),
    )

    assert plan[0].status == "跳过"
    assert plan[0].message == "原文件不存在"


def test_execute_rename_plan_renames_file(tmp_path):
    source_path = _create_file(tmp_path / "old.xlsx")
    action = build_rename_plan(
        [_rule(source_path, new_name="new", extension=".xlsx")],
        RenameSettings(),
    )[0]

    summary = execute_rename_plan([action])

    assert summary == {"success_count": 1, "skipped_count": 0, "failed_count": 0}
    assert not source_path.exists()
    assert (tmp_path / "new.xlsx").exists()


def test_single_failure_does_not_block_other_success(monkeypatch, tmp_path):
    ok_path = _create_file(tmp_path / "ok.xlsx")
    fail_path = _create_file(tmp_path / "fail.xlsx")
    actions = build_rename_plan(
        [
            _rule(ok_path, new_name="ok_new", extension=".xlsx", row_number=2),
            _rule(fail_path, new_name="fail_new", extension=".xlsx", row_number=3),
        ],
        RenameSettings(),
    )
    original_rename = Path.rename

    def fake_rename(self, target):
        if self.name == "fail.xlsx":
            raise OSError("locked")
        return original_rename(self, target)

    monkeypatch.setattr(Path, "rename", fake_rename)

    summary = execute_rename_plan(actions)

    assert summary == {"success_count": 1, "skipped_count": 0, "failed_count": 1}
    assert (tmp_path / "ok_new.xlsx").exists()
    assert fail_path.exists()
    assert actions[1].status == "失败"
    assert "locked" in actions[1].message


def test_write_rename_results_updates_status_columns_and_log_sheet(tmp_path):
    source_path = _create_file(tmp_path / "old.xlsx")
    rule_path = create_rename_rule_workbook([str(source_path)])
    action = build_rename_plan(
        [_rule(source_path, new_name="new", extension=".xlsx")],
        RenameSettings(),
    )[0]
    execute_rename_plan([action])

    write_rename_results_to_workbook(rule_path, [action])

    workbook = load_workbook(rule_path)
    try:
        rule_sheet = workbook[RULE_SHEET_NAME]
        log_sheet = workbook[LOG_SHEET_NAME]
        assert rule_sheet["E2"].value == str(tmp_path / "new.xlsx")
        assert rule_sheet["F2"].value == "成功"
        assert rule_sheet["G2"].value == "已重命名"
        assert log_sheet.max_row == 2
        assert log_sheet["B2"].value == str(source_path)
        assert log_sheet["C2"].value == str(tmp_path / "new.xlsx")
        assert log_sheet["D2"].value == "old"
        assert log_sheet["F2"].value == "成功"
    finally:
        workbook.close()


def test_rule_workbook_keeps_original_name_column_as_stem_only(tmp_path):
    source_path = _create_file(tmp_path / "1.xlsx")

    rule_path = create_rename_rule_workbook([str(source_path)])

    workbook = load_workbook(rule_path)
    try:
        sheet = workbook[RULE_SHEET_NAME]
        assert sheet["B2"].value == "1"
        assert sheet["D2"].value == ".xlsx"
    finally:
        workbook.close()


def _rule(
    source_path: Path,
    new_name: str,
    extension: str,
    row_number: int = 2,
) -> RenameRule:
    return RenameRule(
        row_number=row_number,
        original_path=str(source_path),
        original_name=source_path.name,
        new_name_raw=new_name,
        extension_raw=extension,
        original_extension=source_path.suffix,
    )

import pytest
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from src.excel_efficiency_toolkit.color_sum_ops import (
    MissingSheetError,
    _cell_address_from_row_column,
    _get_fast_source_skip_reason,
    _get_cell_fill_color,
    build_limited_log_items,
    build_external_sum_formula,
    build_cell_address,
    build_matching_sheet_plan,
    can_read_with_openpyxl,
    column_number_to_letter,
    get_openpyxl_fill_key,
    get_value_from_used_range_array,
    group_cells_by_contiguous_rows,
    is_excel_file,
    is_visible_sheet_state,
    read_sheet_values_with_openpyxl,
    same_excel_color,
    scan_target_workbook_color_plan_with_openpyxl,
    sheet_exists_with_openpyxl,
    summarize_sheet_results,
    normalize_write_mode,
    to_number_or_none,
)


def _create_openpyxl_test_workbook(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "货币资金"
    sheet.cell(row=4, column=3, value=12.5)
    sheet.cell(row=4, column=4, value="文本")

    chinese_sheet = workbook.create_sheet("中文表")
    chinese_sheet.cell(row=2, column=2, value=88)

    workbook.save(path)
    workbook.close()


def _create_color_target_workbook(path):
    yellow = PatternFill(fill_type="solid", fgColor="FFFF00")
    green = PatternFill(fill_type="solid", fgColor="00FF00")

    workbook = Workbook()
    cash = workbook.active
    cash.title = "货币资金"
    cash["C4"] = 1
    cash["C4"].fill = yellow
    cash["C5"] = 2
    cash["C5"].fill = yellow
    cash["D5"] = 3
    cash["D5"].fill = yellow
    cash["E5"] = 4
    cash["E5"].fill = green

    trading = workbook.create_sheet("交易性金融资产")
    trading["C4"] = 5
    trading["C4"].fill = yellow

    no_color = workbook.create_sheet("无颜色表")
    no_color["C4"] = 6

    hidden = workbook.create_sheet("隐藏表")
    hidden.sheet_state = "hidden"
    hidden["C4"] = 7
    hidden["C4"].fill = yellow

    chinese = workbook.create_sheet("中文表")
    chinese["B2"] = 8
    chinese["B2"].fill = yellow

    workbook.save(path)
    workbook.close()


class _FakeInterior:
    def __init__(self, color, color_index, pattern):
        self.Color = color
        self.ColorIndex = color_index
        self.Pattern = pattern


class _FakeCell:
    def __init__(self, color, color_index, pattern):
        self.Interior = _FakeInterior(color, color_index, pattern)


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        (1, 1.0),
        (1.25, 1.25),
        ("2.5", 2.5),
        (None, None),
        ("", None),
        ("   ", None),
        ("abc", None),
        (True, None),
        (False, None),
    ],
)
def test_to_number_or_none_handles_common_values(value, expected):
    assert to_number_or_none(value) == expected


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        ("1", "formula"),
        ("公式", "formula"),
        ("formula", "formula"),
        ("写入公式", "formula"),
        ("写入求和公式", "formula"),
        ("2", "value"),
        ("数值", "value"),
        ("value", "value"),
        ("只保留数值", "value"),
        ("只写入汇总数值", "value"),
    ],
)
def test_normalize_write_mode_accepts_number_chinese_and_english(value, expected):
    assert normalize_write_mode(value) == expected


def test_normalize_write_mode_rejects_invalid_value():
    with pytest.raises(ValueError):
        normalize_write_mode("其他")


def test_build_external_sum_formula_starts_with_equal_sign():
    formula = build_external_sum_formula([
        ("C:/data/sub1.xlsx", "Sheet1", "C4"),
        ("C:/data/sub2.xlsx", "Sheet1", "C4"),
    ])

    assert formula.startswith("=")
    assert "+'" in formula
    assert "[sub1.xlsx]Sheet1'!C4" in formula
    assert "[sub2.xlsx]Sheet1'!C4" in formula


def test_build_external_sum_formula_handles_chinese_file_and_sheet_names():
    formula = build_external_sum_formula([
        ("C:/审计/子公司1.xlsx", "货币资金", "D5"),
    ])

    assert "[子公司1.xlsx]货币资金'!D5" in formula


def test_build_external_sum_formula_escapes_single_quotes():
    formula = build_external_sum_formula([
        ("C:/审计/子'公司.xlsx", "O'Brien", "A1"),
    ])

    assert "子''公司.xlsx" in formula
    assert "O''Brien" in formula


def test_build_external_sum_formula_rejects_empty_refs():
    with pytest.raises(ValueError):
        build_external_sum_formula([])


@pytest.mark.parametrize(
    ("column", "expected"),
    [
        (1, "A"),
        (3, "C"),
        (26, "Z"),
        (27, "AA"),
        (52, "AZ"),
        (53, "BA"),
    ],
)
def test_column_number_to_letter(column, expected):
    assert column_number_to_letter(column) == expected


def test_column_number_to_letter_rejects_invalid_column():
    with pytest.raises(ValueError):
        column_number_to_letter(0)


def test_build_cell_address_uses_relative_excel_address():
    assert build_cell_address(4, 3) == "C4"
    assert build_cell_address(1, 27) == "AA1"


def test_build_cell_address_rejects_invalid_row():
    with pytest.raises(ValueError):
        build_cell_address(0, 1)


@pytest.mark.parametrize(
    ("row", "column", "expected"),
    [
        (4, 3, "C4"),
        (1, 26, "Z1"),
        (1, 27, "AA1"),
    ],
)
def test_cell_address_from_row_column_returns_relative_excel_address(row, column, expected):
    assert _cell_address_from_row_column(row, column) == expected


def test_group_cells_by_contiguous_rows_groups_same_row_adjacent_columns():
    cells = [(5, 3), (4, 4), (4, 3), (4, 6)]

    assert group_cells_by_contiguous_rows(cells) == [
        [(4, 3), (4, 4)],
        [(4, 6)],
        [(5, 3)],
    ]


def test_group_cells_by_contiguous_rows_returns_empty_for_no_cells():
    assert group_cells_by_contiguous_rows([]) == []


def test_get_value_from_used_range_array_reads_ordinary_2d_tuple():
    values = (
        (1, 2),
        (3, 4),
    )

    assert get_value_from_used_range_array(values, 4, 3, 4, 3) == 1
    assert get_value_from_used_range_array(values, 4, 3, 5, 4) == 4


def test_get_value_from_used_range_array_reads_single_row():
    values = (("A", "B", "C"),)

    assert get_value_from_used_range_array(values, 2, 3, 2, 4) == "B"


def test_get_value_from_used_range_array_reads_flat_single_row():
    values = ("A", "B", "C")

    assert get_value_from_used_range_array(values, 2, 3, 2, 5) == "C"


def test_get_value_from_used_range_array_reads_single_column():
    values = (
        (10,),
        (20,),
        (30,),
    )

    assert get_value_from_used_range_array(values, 2, 3, 4, 3) == 30


def test_get_value_from_used_range_array_reads_flat_single_column():
    values = (10, 20, 30)

    assert get_value_from_used_range_array(values, 2, 3, 4, 3) == 30


def test_get_value_from_used_range_array_reads_single_cell_scalar():
    assert get_value_from_used_range_array(99, 4, 3, 4, 3) == 99


def test_get_value_from_used_range_array_returns_none_for_out_of_range():
    values = ((1, 2),)

    assert get_value_from_used_range_array(values, 4, 3, 3, 3) is None
    assert get_value_from_used_range_array(values, 4, 3, 4, 5) is None
    assert get_value_from_used_range_array(values, 4, 3, 5, 3) is None


def test_get_value_from_used_range_array_keeps_blank_none_value():
    values = ((None,),)

    assert get_value_from_used_range_array(values, 4, 3, 4, 3) is None


def test_same_excel_color_matches_integer_colors():
    assert same_excel_color(65535, 65535) is True
    assert same_excel_color(65535, 255) is False
    assert same_excel_color(None, 65535) is False
    assert same_excel_color(65535, None) is False


def test_get_cell_fill_color_keeps_valid_color_even_if_pattern_reports_none():
    cell = _FakeCell(color=65535, color_index=6, pattern=-4142)

    assert _get_cell_fill_color(cell) == 65535


def test_get_cell_fill_color_rejects_no_fill_color_index():
    cell = _FakeCell(color=16777215, color_index=-4142, pattern=-4142)

    assert _get_cell_fill_color(cell) is None


def test_is_excel_file_accepts_supported_extensions_case_insensitive():
    assert is_excel_file("book.xlsx") is True
    assert is_excel_file("book.XLSM") is True
    assert is_excel_file("book.xls") is True


def test_is_excel_file_rejects_temporary_excel_files():
    assert is_excel_file("~$book.xlsx") is False
    assert is_excel_file("C:/tmp/~$book.xlsm") is False


def test_is_excel_file_rejects_non_excel_files():
    assert is_excel_file("book.csv") is False
    assert is_excel_file("book.txt") is False


@pytest.mark.parametrize(
    ("visible_value", "expected"),
    [
        (-1, True),
        ("-1", True),
        ("visible", True),
        (True, True),
        (0, False),
        (2, False),
        ("hidden", False),
        (False, False),
    ],
)
def test_is_visible_sheet_state_handles_excel_and_openpyxl_values(visible_value, expected):
    assert is_visible_sheet_state(visible_value) is expected


def test_build_matching_sheet_plan_keeps_visible_sheets_with_matched_cells_only():
    plan = build_matching_sheet_plan([
        {"sheet_name": "货币资金", "visible": -1, "cells": [(4, 3), (5, 3)]},
        {"sheet_name": "隐藏表", "visible": 0, "cells": [(4, 3)]},
        {"sheet_name": "无颜色表", "visible": -1, "cells": []},
    ])

    assert plan == [
        {
            "sheet_name": "货币资金",
            "cells": [(4, 3), (5, 3)],
            "matched_cell_count": 2,
        }
    ]


def test_summarize_sheet_results_sums_multi_sheet_counts():
    summary = summarize_sheet_results([
        {
            "sheet_name": "货币资金",
            "matched_cell_count": 3,
            "written_cell_count": 3,
            "matched_source_sheet_count": 2,
            "missing_sheet_file_count": 1,
            "ignored_non_numeric_count": 4,
        },
        {
            "sheet_name": "交易性金融资产",
            "matched_cell_count": 2,
            "written_cell_count": 2,
            "matched_source_sheet_count": 1,
            "missing_sheet_file_count": 2,
            "ignored_non_numeric_count": 0,
        },
    ])

    assert summary == {
        "matched_sheet_count": 2,
        "matched_cell_count": 5,
        "written_cell_count": 5,
        "matched_source_sheet_count": 3,
        "missing_sheet_file_count": 3,
        "ignored_non_numeric_count": 4,
    }


def test_external_formulas_keep_different_sheet_references_independent():
    cash_formula = build_external_sum_formula([
        ("C:/审计/子公司1.xlsx", "货币资金", "C4"),
        ("C:/审计/子公司2.xlsx", "货币资金", "C4"),
    ])
    trading_formula = build_external_sum_formula([
        ("C:/审计/子公司1.xlsx", "交易性金融资产", "C4"),
        ("C:/审计/子公司2.xlsx", "交易性金融资产", "C4"),
    ])

    assert "货币资金'!C4" in cash_formula
    assert "交易性金融资产" not in cash_formula
    assert "交易性金融资产'!C4" in trading_formula
    assert "货币资金" not in trading_formula


def test_build_limited_log_items_limits_long_lists():
    items = [f"item-{index}" for index in range(12)]

    limited = build_limited_log_items(items, limit=10)

    assert limited[:10] == items[:10]
    assert limited[-1] == "其余 2 项已省略（共 12 项）。"


def test_get_openpyxl_fill_key_handles_blank_same_and_different_colors():
    workbook = Workbook()
    sheet = workbook.active
    yellow = PatternFill(fill_type="solid", fgColor="FFFF00")
    green = PatternFill(fill_type="solid", fgColor="00FF00")
    sheet["A1"].fill = yellow
    sheet["B1"].fill = yellow
    sheet["C1"].fill = green

    assert get_openpyxl_fill_key(sheet["D1"]) is None
    assert get_openpyxl_fill_key(sheet["A1"]) == get_openpyxl_fill_key(sheet["B1"])
    assert get_openpyxl_fill_key(sheet["A1"]) != get_openpyxl_fill_key(sheet["C1"])


def test_scan_target_workbook_color_plan_single_sheet_uses_selected_sheet_when_target_blank(tmp_path):
    path = tmp_path / "target.xlsx"
    _create_color_target_workbook(path)

    plan = scan_target_workbook_color_plan_with_openpyxl(
        str(path),
        selected_sheet_name="货币资金",
        selected_cell_address="C5",
        scope="single_sheet",
    )

    assert plan["scope"] == "single_sheet"
    assert plan["matched_sheet_count"] == 1
    assert plan["matched_cell_count"] == 3
    assert plan["sheet_plans"] == [
        {"sheet_name": "货币资金", "cells": ["C4", "C5", "D5"]},
    ]


def test_scan_target_workbook_color_plan_single_sheet_uses_specified_sheet(tmp_path):
    path = tmp_path / "target.xlsx"
    _create_color_target_workbook(path)

    plan = scan_target_workbook_color_plan_with_openpyxl(
        str(path),
        selected_sheet_name="货币资金",
        selected_cell_address="C5",
        scope="single_sheet",
        target_sheet_name="中文表",
    )

    assert plan["matched_sheet_count"] == 1
    assert plan["matched_cell_count"] == 1
    assert plan["sheet_plans"] == [
        {"sheet_name": "中文表", "cells": ["B2"]},
    ]


def test_scan_target_workbook_color_plan_all_matching_sheets_skips_hidden_and_blank_sheets(tmp_path):
    path = tmp_path / "target.xlsx"
    _create_color_target_workbook(path)

    plan = scan_target_workbook_color_plan_with_openpyxl(
        str(path),
        selected_sheet_name="货币资金",
        selected_cell_address="$C$5",
        scope="all_matching_sheets",
    )

    assert plan["matched_sheet_count"] == 3
    assert plan["matched_cell_count"] == 5
    assert plan["sheet_plans"] == [
        {"sheet_name": "货币资金", "cells": ["C4", "C5", "D5"]},
        {"sheet_name": "交易性金融资产", "cells": ["C4"]},
        {"sheet_name": "中文表", "cells": ["B2"]},
    ]


def test_scan_target_workbook_color_plan_rejects_unsupported_target_format(tmp_path):
    path = tmp_path / "target.xls"
    path.write_text("not excel", encoding="utf-8")

    with pytest.raises(ValueError, match="格式不支持"):
        scan_target_workbook_color_plan_with_openpyxl(
            str(path),
            selected_sheet_name="货币资金",
            selected_cell_address="C5",
            scope="single_sheet",
        )


def test_scan_target_workbook_color_plan_rejects_selected_cell_without_fill(tmp_path):
    path = tmp_path / "target.xlsx"
    _create_color_target_workbook(path)

    with pytest.raises(ValueError, match="没有识别到填充色"):
        scan_target_workbook_color_plan_with_openpyxl(
            str(path),
            selected_sheet_name="货币资金",
            selected_cell_address="A1",
            scope="single_sheet",
        )


def test_can_read_with_openpyxl_accepts_xlsx_and_xlsm():
    assert can_read_with_openpyxl("book.xlsx") is True
    assert can_read_with_openpyxl("book.XLSM") is True


def test_can_read_with_openpyxl_rejects_xls_temp_and_non_excel_files():
    assert can_read_with_openpyxl("book.xls") is False
    assert can_read_with_openpyxl("~$book.xlsx") is False
    assert can_read_with_openpyxl("book.csv") is False


def test_fast_source_skip_reason_rejects_xls_and_temporary_files(tmp_path):
    xlsx_path = tmp_path / "book.xlsx"
    xlsx_path.write_bytes(b"placeholder")
    xls_path = tmp_path / "book.xls"
    xls_path.write_bytes(b"placeholder")
    temp_path = tmp_path / "~$book.xlsx"
    temp_path.write_bytes(b"placeholder")

    assert _get_fast_source_skip_reason(str(xlsx_path)) is None
    assert "暂不支持快速汇总" in _get_fast_source_skip_reason(str(xls_path))
    assert "临时文件" in _get_fast_source_skip_reason(str(temp_path))


def test_sheet_exists_with_openpyxl_handles_existing_missing_and_chinese_sheet(tmp_path):
    path = tmp_path / "source.xlsx"
    _create_openpyxl_test_workbook(path)

    assert sheet_exists_with_openpyxl(str(path), "货币资金") is True
    assert sheet_exists_with_openpyxl(str(path), "中文表") is True
    assert sheet_exists_with_openpyxl(str(path), "不存在") is False


def test_read_sheet_values_with_openpyxl_reads_selected_cells_and_blank(tmp_path):
    path = tmp_path / "source.xlsx"
    _create_openpyxl_test_workbook(path)

    values = read_sheet_values_with_openpyxl(
        str(path),
        "货币资金",
        [(4, 3), (4, 4), (6, 3)],
    )

    assert values[(4, 3)] == 12.5
    assert values[(4, 4)] == "文本"
    assert values[(6, 3)] is None


def test_read_sheet_values_with_openpyxl_reads_chinese_sheet_name(tmp_path):
    path = tmp_path / "source.xlsx"
    _create_openpyxl_test_workbook(path)

    values = read_sheet_values_with_openpyxl(str(path), "中文表", [(2, 2)])

    assert values[(2, 2)] == 88


def test_read_sheet_values_with_openpyxl_raises_for_missing_sheet(tmp_path):
    path = tmp_path / "source.xlsx"
    _create_openpyxl_test_workbook(path)

    with pytest.raises(MissingSheetError):
        read_sheet_values_with_openpyxl(str(path), "不存在", [(1, 1)])


def test_read_sheet_values_with_openpyxl_closes_file_handle(tmp_path):
    path = tmp_path / "source.xlsx"
    renamed_path = tmp_path / "renamed.xlsx"
    _create_openpyxl_test_workbook(path)

    read_sheet_values_with_openpyxl(str(path), "货币资金", [(4, 3)])
    path.rename(renamed_path)

    assert renamed_path.exists()

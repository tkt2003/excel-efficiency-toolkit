from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side

from src.excel_efficiency_toolkit.workbook_merge_ops import (
    DIRECTORY_SHEET_NAME,
    LOG_SHEET_NAME,
    build_unique_sheet_name,
    copy_worksheet_basic,
    create_target_backup,
    get_first_visible_sheet_name,
    is_office_temp_file,
    is_supported_openpyxl_workbook,
    merge_workbooks_to_existing_workbook,
    resolve_source_sheet_name,
)


def test_supported_openpyxl_workbook_extensions():
    assert is_supported_openpyxl_workbook("book.xlsx")
    assert is_supported_openpyxl_workbook("book.XLSM")
    assert not is_supported_openpyxl_workbook("book.xls")
    assert not is_supported_openpyxl_workbook("book.csv")


def test_office_temp_file_detection():
    assert is_office_temp_file(r"D:\tmp\~$book.xlsx")
    assert not is_office_temp_file(r"D:\tmp\book.xlsx")


def test_blank_requested_sheet_uses_first_visible_sheet():
    workbook = Workbook()
    workbook.active.title = "隐藏"
    workbook.active.sheet_state = "hidden"
    workbook.create_sheet("第一个可见")

    assert get_first_visible_sheet_name(workbook) == "第一个可见"

    actual_name, available_names, message = resolve_source_sheet_name(workbook, "")
    assert actual_name == "第一个可见"
    assert available_names == ["第一个可见"]
    assert "第一个可见" in message


def test_requested_sheet_name_matches_case_and_outer_spaces():
    workbook = Workbook()
    workbook.active.title = " Data "

    actual_name, available_names, message = resolve_source_sheet_name(workbook, " data ")

    assert actual_name == " Data "
    assert available_names == [" Data "]
    assert message == "匹配指定源 Sheet 名"


def test_missing_requested_sheet_returns_available_sheet_list():
    workbook = Workbook()
    workbook.active.title = "资产负债表"
    workbook.create_sheet("利润表")

    actual_name, available_names, message = resolve_source_sheet_name(workbook, "货币资金")

    assert actual_name is None
    assert available_names == ["资产负债表", "利润表"]
    assert "找不到指定源 Sheet" in message


def test_build_unique_sheet_name_adds_sequence_and_keeps_limit():
    assert build_unique_sheet_name("来源", {"来源", "来源_2"}) == "来源_3"

    long_name = "A" * 40
    assert build_unique_sheet_name(long_name, set()) == "A" * 31

    duplicate_long_name = build_unique_sheet_name(long_name, {"A" * 31})
    assert duplicate_long_name == f"{'A' * 29}_2"
    assert len(duplicate_long_name) == 31


def test_copy_worksheet_basic_copies_values_formulas_and_basic_formatting():
    source_wb = Workbook()
    source_ws = source_wb.active
    source_ws.title = "货币资金"
    source_ws["A1"] = "标题"
    source_ws["A1"].hyperlink = "https://example.com"
    source_ws["B1"] = "=SUM(1,2)"
    source_ws["A2"] = 10
    source_ws["A2"].font = Font(bold=True, color="00FF0000")
    source_ws["A2"].fill = PatternFill(fill_type="solid", fgColor="00FFFF00")
    source_ws["A2"].border = Border(left=Side(style="thin"))
    source_ws["A2"].alignment = Alignment(horizontal="center")
    source_ws["A2"].number_format = "#,##0"
    source_ws["A2"].protection = Protection(locked=False)
    source_ws.row_dimensions[2].height = 28
    source_ws.column_dimensions["A"].width = 18
    source_ws.merge_cells("C1:D1")
    source_ws["C1"] = "合并"
    source_ws.freeze_panes = "A2"
    source_ws.sheet_properties.tabColor = "00FF0000"

    target_wb = Workbook()
    target_ws = copy_worksheet_basic(source_ws, target_wb, "来源")

    assert target_ws["A1"].value == "标题"
    assert target_ws["A1"].hyperlink.target == "https://example.com"
    assert target_ws["B1"].value == "=SUM(1,2)"
    assert target_ws["A2"].font.bold is True
    assert target_ws["A2"].fill.fgColor.rgb == "00FFFF00"
    assert target_ws["A2"].border.left.style == "thin"
    assert target_ws["A2"].alignment.horizontal == "center"
    assert target_ws["A2"].number_format == "#,##0"
    assert target_ws["A2"].protection.locked is False
    assert target_ws.row_dimensions[2].height == 28
    assert target_ws.column_dimensions["A"].width == 18
    assert "C1:D1" in [str(item) for item in target_ws.merged_cells.ranges]
    assert target_ws.freeze_panes == "A2"
    assert target_ws.sheet_properties.tabColor.rgb == "00FF0000"


def test_create_target_backup_generates_non_overwriting_paths(tmp_path):
    target_path = tmp_path / "目标.xlsx"
    _save_workbook(target_path, "目标")

    first_backup_path = create_target_backup(str(target_path))
    second_backup_path = create_target_backup(str(target_path))

    assert Path(first_backup_path).exists()
    assert Path(second_backup_path).exists()
    assert first_backup_path != second_backup_path


def test_merge_workbooks_creates_target_sheets_directory_log_and_skips_unsupported_files(tmp_path):
    source_dir_1 = tmp_path / "one"
    source_dir_2 = tmp_path / "two"
    source_dir_1.mkdir()
    source_dir_2.mkdir()
    source_path_1 = source_dir_1 / "来源.xlsx"
    source_path_2 = source_dir_2 / "来源.xlsx"
    missing_sheet_path = tmp_path / "缺少指定sheet.xlsx"
    old_xls_path = tmp_path / "旧格式.xls"
    temp_path = tmp_path / "~$临时.xlsx"
    target_path = tmp_path / "目标.xlsx"

    _save_workbook(source_path_1, "货币资金", value=1)
    _save_workbook(source_path_2, "货币资金", value=2)
    _save_workbook(missing_sheet_path, "其他", value=3)
    old_xls_path.write_text("old", encoding="utf-8")
    temp_path.write_text("temp", encoding="utf-8")
    _save_workbook(target_path, "已有")

    result = merge_workbooks_to_existing_workbook(
        source_paths=[str(source_path_1), str(source_path_2), str(missing_sheet_path), str(old_xls_path), str(temp_path)],
        target_path=str(target_path),
        requested_sheet_name=" 货币资金 ",
        values_only=False,
    )

    assert result["success_count"] == 2
    assert result["skipped_count"] == 3
    assert result["failed_count"] == 0
    assert Path(result["backup_path"]).exists()

    target_wb = load_workbook(target_path)
    assert "已有" in target_wb.sheetnames
    assert "来源" in target_wb.sheetnames
    assert "来源_2" in target_wb.sheetnames
    assert DIRECTORY_SHEET_NAME in target_wb.sheetnames
    assert LOG_SHEET_NAME in target_wb.sheetnames

    directory_ws = target_wb[DIRECTORY_SHEET_NAME]
    assert directory_ws["B2"].value == "来源"
    assert directory_ws["F2"].hyperlink is not None

    log_ws = target_wb[LOG_SHEET_NAME]
    statuses = [row[6] for row in log_ws.iter_rows(min_row=2, values_only=True)]
    descriptions = [row[7] for row in log_ws.iter_rows(min_row=2, values_only=True)]
    available_lists = [row[8] for row in log_ws.iter_rows(min_row=2, values_only=True)]
    assert statuses.count("成功") == 2
    assert statuses.count("跳过") == 3
    assert any("请另存为 xlsx/xlsm 后再处理" in description for description in descriptions)
    assert any("临时文件已跳过" in description for description in descriptions)
    assert "其他" in available_lists


def test_merge_workbooks_skips_source_file_when_it_is_target_workbook(tmp_path):
    source_path = tmp_path / "来源.xlsx"
    target_path = tmp_path / "目标.xlsx"
    _save_workbook(source_path, "货币资金", value=1)
    _save_workbook(target_path, "货币资金", value=2)

    result = merge_workbooks_to_existing_workbook(
        source_paths=[str(source_path), str(target_path)],
        target_path=str(target_path),
        requested_sheet_name="货币资金",
        values_only=False,
    )

    assert result["success_count"] == 1
    assert result["skipped_count"] == 1
    assert result["failed_count"] == 0

    target_wb = load_workbook(target_path)
    assert "来源" in target_wb.sheetnames
    assert "目标" not in target_wb.sheetnames

    log_ws = target_wb[LOG_SHEET_NAME]
    descriptions = [row[7] for row in log_ws.iter_rows(min_row=2, values_only=True)]
    assert "源文件与目标工作簿相同，已跳过" in descriptions


def test_values_only_keeps_formula_when_formula_cache_is_empty_and_logs_message(tmp_path):
    source_path = tmp_path / "公式来源.xlsx"
    target_path = tmp_path / "目标.xlsx"
    _save_workbook(source_path, "货币资金", value=1, formula="=SUM(A2,1)")
    _save_workbook(target_path, "已有")

    result = merge_workbooks_to_existing_workbook(
        source_paths=[str(source_path)],
        target_path=str(target_path),
        requested_sheet_name="货币资金",
        values_only=True,
    )

    assert result["success_count"] == 1

    target_wb = load_workbook(target_path, data_only=False)
    imported_ws = target_wb["公式来源"]
    assert imported_ws["B2"].value == "=SUM(A2,1)"

    log_ws = target_wb[LOG_SHEET_NAME]
    assert log_ws["H2"].value == "匹配指定源 Sheet 名；公式缓存为空，已保留公式 1 个"
    assert log_ws["L2"].value == "是"
    assert log_ws["M2"].value == "是"


def _save_workbook(path: Path, sheet_name: str, value=1, formula: str | None = None) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet["A1"] = "项目"
    worksheet["A2"] = value
    if formula:
        worksheet["B2"] = formula
    workbook.save(path)

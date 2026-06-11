import os
import time


EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
OPENPYXL_EXTENSIONS = {".xlsx", ".xlsm"}
XL_COLOR_INDEX_NONE = -4142
XL_PATTERN_NONE = -4142


class MissingSheetError(ValueError):
    pass


def column_number_to_letter(col: int) -> str:
    if col < 1:
        raise ValueError("列号必须大于等于 1。")

    letters = ""
    while col > 0:
        col, remainder = divmod(col - 1, 26)
        letters = chr(ord("A") + remainder) + letters
    return letters


def build_cell_address(row: int, col: int) -> str:
    if row < 1:
        raise ValueError("行号必须大于等于 1。")
    return f"{column_number_to_letter(col)}{row}"


def group_cells_by_contiguous_rows(cells: list[tuple[int, int]]) -> list[list[tuple[int, int]]]:
    groups: list[list[tuple[int, int]]] = []
    for row, col in sorted(cells):
        if groups and groups[-1][-1][0] == row and groups[-1][-1][1] + 1 == col:
            groups[-1].append((row, col))
        else:
            groups.append([(row, col)])
    return groups


def get_value_from_used_range_array(
    values,
    used_range_first_row: int,
    used_range_first_col: int,
    row: int,
    col: int,
):
    row_index = row - used_range_first_row
    col_index = col - used_range_first_col
    if row_index < 0 or col_index < 0:
        return None

    if not isinstance(values, (tuple, list)):
        return values if row_index == 0 and col_index == 0 else None

    if not values:
        return None

    first_item = values[0]
    if isinstance(first_item, (tuple, list)):
        if row_index >= len(values):
            return None
        row_values = values[row_index]
        if isinstance(row_values, (tuple, list)):
            if col_index >= len(row_values):
                return None
            return row_values[col_index]
        return row_values if col_index == 0 else None

    if row_index == 0 and col_index < len(values):
        return values[col_index]
    if col_index == 0 and row_index < len(values):
        return values[row_index]
    return None


def is_valid_number(value: object) -> bool:
    return to_number_or_none(value) is not None


def to_number_or_none(value: object) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        try:
            return float(stripped)
        except ValueError:
            return None
    return None


def normalize_write_mode(user_input: str) -> str:
    normalized = str(user_input).strip().lower()
    if normalized in {"1", "公式", "formula", "写入公式", "写入求和公式"}:
        return "formula"
    if normalized in {"2", "数值", "value", "只保留数值", "只写入汇总数值"}:
        return "value"
    raise ValueError("写入方式无效，请输入 1/公式 或 2/数值。")


def build_external_sum_formula(source_refs: list[tuple[str, str, str]]) -> str:
    if not source_refs:
        raise ValueError("公式引用不能为空。")

    refs = [
        _build_external_cell_ref(source_file_path, sheet_name, cell_address)
        for source_file_path, sheet_name, cell_address in source_refs
    ]
    return "=" + "+".join(refs)


def same_excel_color(color1, color2) -> bool:
    if color1 is None or color2 is None:
        return False
    return color1 == color2


def is_excel_file(path: str) -> bool:
    filename = os.path.basename(path)
    if filename.startswith("~$"):
        return False

    _, ext = os.path.splitext(filename)
    return ext.lower() in EXCEL_EXTENSIONS


def is_visible_sheet_state(visible_value: object) -> bool:
    if isinstance(visible_value, bool):
        return visible_value

    try:
        return int(visible_value) == -1
    except (TypeError, ValueError):
        pass

    return str(visible_value).strip().lower() in {"visible", "true"}


def build_matching_sheet_plan(sheet_infos: list[dict]) -> list[dict]:
    plan = []
    for sheet_info in sheet_infos:
        cells = list(sheet_info.get("cells") or [])
        if not is_visible_sheet_state(sheet_info.get("visible")) or not cells:
            continue

        plan.append({
            "sheet_name": str(sheet_info.get("sheet_name", "")),
            "cells": cells,
            "matched_cell_count": len(cells),
        })
    return plan


def summarize_sheet_results(sheet_results: list[dict]) -> dict:
    return {
        "matched_sheet_count": len(sheet_results),
        "matched_cell_count": sum(result.get("matched_cell_count", 0) for result in sheet_results),
        "written_cell_count": sum(result.get("written_cell_count", 0) for result in sheet_results),
        "matched_source_sheet_count": sum(result.get("matched_source_sheet_count", 0) for result in sheet_results),
        "missing_sheet_file_count": sum(result.get("missing_sheet_file_count", 0) for result in sheet_results),
        "ignored_non_numeric_count": sum(result.get("ignored_non_numeric_count", 0) for result in sheet_results),
    }


def build_limited_log_items(items: list[str], limit: int = 10) -> list[str]:
    if len(items) <= limit:
        return list(items)
    return list(items[:limit]) + [f"其余 {len(items) - limit} 项已省略（共 {len(items)} 项）。"]


def get_openpyxl_fill_key(cell) -> tuple | None:
    fill = getattr(cell, "fill", None)
    if fill is None or not fill.fill_type:
        return None

    fg_color = fill.fgColor
    return (
        fill.fill_type,
        fg_color.type,
        fg_color.rgb,
        fg_color.indexed,
        fg_color.theme,
        fg_color.tint,
    )


def scan_target_workbook_color_plan_with_openpyxl(
    workbook_path: str,
    selected_sheet_name: str,
    selected_cell_address: str,
    scope: str,
    target_sheet_name: str | None = None,
) -> dict:
    _validate_target_workbook_path_for_color_scan(workbook_path)

    workbook = None
    try:
        workbook = _load_openpyxl_workbook(workbook_path)
        selected_sheet_actual_name = _find_openpyxl_sheet_name(workbook.sheetnames, selected_sheet_name)
        if selected_sheet_actual_name is None:
            raise ValueError(f"目标工作簿中不存在当前工作表：{selected_sheet_name}。")

        selected_sheet = workbook[selected_sheet_actual_name]
        selected_cell = selected_sheet[_normalize_cell_address(selected_cell_address)]
        selected_color_key = get_openpyxl_fill_key(selected_cell)
        if selected_color_key is None:
            raise ValueError("当前选中单元格在已保存文件中没有识别到填充色；如果刚修改了颜色，请先保存目标工作簿。")

        sheet_plans = []
        if scope == "single_sheet":
            scan_sheet_name = target_sheet_name.strip() if target_sheet_name else selected_sheet_actual_name
            actual_sheet_name = _find_openpyxl_sheet_name(workbook.sheetnames, scan_sheet_name)
            if actual_sheet_name is None:
                raise ValueError(f"目标工作簿中不存在 sheet：{scan_sheet_name}。")
            cells = _scan_openpyxl_sheet_for_fill_key(workbook[actual_sheet_name], selected_color_key)
            if not cells:
                raise ValueError(f"目标 sheet 中没有找到与当前选中单元格同色的单元格：{actual_sheet_name}。")
            sheet_plans.append({"sheet_name": actual_sheet_name, "cells": cells})
        elif scope == "all_matching_sheets":
            for sheet in workbook.worksheets:
                if not is_visible_sheet_state(sheet.sheet_state):
                    continue
                cells = _scan_openpyxl_sheet_for_fill_key(sheet, selected_color_key)
                if cells:
                    sheet_plans.append({"sheet_name": sheet.title, "cells": cells})
            if not sheet_plans:
                raise ValueError("目标工作簿的可见 sheet 中没有找到与当前选中单元格同色的单元格。")
        else:
            raise ValueError(f"汇总范围无效：{scope}")

        return {
            "workbook_path": os.path.abspath(workbook_path),
            "selected_color_key": selected_color_key,
            "scope": scope,
            "sheet_plans": sheet_plans,
            "matched_sheet_count": len(sheet_plans),
            "matched_cell_count": sum(len(plan["cells"]) for plan in sheet_plans),
        }
    except ValueError:
        raise
    except Exception as e:
        raise RuntimeError(f"无法读取目标工作簿颜色信息：{e}。请确认文件已保存且未损坏。") from e
    finally:
        if workbook is not None:
            workbook.close()


def can_read_with_openpyxl(path: str) -> bool:
    if not is_excel_file(path):
        return False

    _, ext = os.path.splitext(os.path.basename(path))
    return ext.lower() in OPENPYXL_EXTENSIONS


def sheet_exists_with_openpyxl(path: str, sheet_name: str) -> bool:
    workbook = _load_openpyxl_workbook(path)
    try:
        return _find_openpyxl_sheet_name(workbook.sheetnames, sheet_name) is not None
    finally:
        workbook.close()


def read_sheet_values_with_openpyxl(
    path: str,
    sheet_name: str,
    cells: list[tuple[int, int]],
) -> dict[tuple[int, int], object]:
    workbook = _load_openpyxl_workbook(path)
    try:
        actual_sheet_name = _find_openpyxl_sheet_name(workbook.sheetnames, sheet_name)
        if actual_sheet_name is None:
            raise MissingSheetError(f"源文件缺少同名 sheet：{sheet_name}")

        sheet = workbook[actual_sheet_name]
        return _read_openpyxl_cells(sheet, cells)
    finally:
        workbook.close()


def sum_current_sheet_by_fill_color(
    source_paths: list[str],
    write_mode: str,
    target_sheet_name: str | None = None,
    logger=None,
) -> dict:
    normalized_write_mode = normalize_write_mode(write_mode)
    source_paths = list(source_paths)
    if not source_paths:
        raise ValueError("请选择至少一个有效的 Excel 文件。")

    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    target_excel = None
    target_excel_state = {}
    try:
        target_excel = _get_active_excel(win32com.client, pythoncom, logger=logger)
        target_workbook, active_sheet, active_cell = _get_active_target_context(target_excel, logger=logger)
        workbook_path = _get_workbook_full_name(target_workbook)
        selected_sheet_name = active_sheet.Name
        selected_cell_address = _get_cell_address_for_log(active_cell)
        _log_fast_scan_context(logger, target_workbook, workbook_path, selected_sheet_name, selected_cell_address)

        scan_start = time.perf_counter()
        _log(logger, "info", "使用 openpyxl 扫描目标工作簿颜色区域。")
        color_plan = scan_target_workbook_color_plan_with_openpyxl(
            workbook_path=workbook_path,
            selected_sheet_name=selected_sheet_name,
            selected_cell_address=selected_cell_address,
            scope="single_sheet",
            target_sheet_name=target_sheet_name,
        )
        scan_seconds = time.perf_counter() - scan_start
        _log_target_scan_summary(logger, color_plan, scan_seconds)

        sheet_plans = _prepare_sheet_plans_for_write(target_workbook, color_plan)
        target_sheet_plan = sheet_plans[0]
        target_sheet_actual_name = target_sheet_plan["sheet_name"]

        _log(logger, "info", f"汇总范围：仅汇总一个 sheet")
        _log(logger, "info", f"写入方式：{_get_write_mode_label(normalized_write_mode)}")
        _log(logger, "info", f"源文件数量：{len(source_paths)}")

        target_excel_state = _set_excel_fast_mode(target_excel)

        aggregation = _aggregate_sources_with_openpyxl(
            source_paths=source_paths,
            sheet_plans=sheet_plans,
            write_mode=normalized_write_mode,
            logger=logger,
        )

        matched_sources = aggregation["matched_sources_by_sheet"][target_sheet_actual_name]
        if not matched_sources:
            raise RuntimeError(f"所有有效源文件都没有找到同名 sheet：{target_sheet_actual_name}。")

        write_groups = group_cells_by_contiguous_rows(target_sheet_plan["cells"])
        _log(
            logger,
            "info",
            f"正在写入目标工作表：{target_sheet_actual_name}；待写入单元格 {len(target_sheet_plan['cells'])} 个，批量区域 {len(write_groups)} 组。",
        )
        written_cell_count = _write_target_cells(
            target_sheet=target_sheet_plan["sheet"],
            cell_groups=write_groups,
            write_mode=normalized_write_mode,
            totals=aggregation["totals_by_sheet"][target_sheet_actual_name],
            matched_sources=matched_sources,
        )

        return {
            "target_workbook_name": target_workbook.Name,
            "target_sheet_name": target_sheet_actual_name,
            "scope": "single_sheet",
            "matched_sheet_count": 1,
            "matched_cell_count": len(target_sheet_plan["cells"]),
            "source_file_count": len(source_paths),
            "valid_source_file_count": aggregation["valid_source_file_count"],
            "skipped_source_file_count": aggregation["skipped_source_file_count"],
            "matched_source_file_count": len(matched_sources),
            "missing_sheet_file_count": aggregation["missing_count_by_sheet"][target_sheet_actual_name],
            "written_cell_count": written_cell_count,
            "ignored_non_numeric_count": aggregation["ignored_count_by_sheet"][target_sheet_actual_name],
            "write_mode": normalized_write_mode,
            "target_scan_seconds": scan_seconds,
        }

    finally:
        try:
            _restore_excel_state(target_excel, target_excel_state)
        except Exception:
            pass
        pythoncom.CoUninitialize()


def sum_matching_sheets_by_fill_color(
    source_paths: list[str],
    write_mode: str,
    logger=None,
) -> dict:
    normalized_write_mode = normalize_write_mode(write_mode)
    source_paths = list(source_paths)
    if not source_paths:
        raise ValueError("请选择至少一个有效的 Excel 文件。")

    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    target_excel = None
    target_excel_state = {}
    try:
        target_excel = _get_active_excel(win32com.client, pythoncom, logger=logger)
        target_workbook, active_sheet, active_cell = _get_active_target_context(target_excel, logger=logger)
        workbook_path = _get_workbook_full_name(target_workbook)
        selected_sheet_name = active_sheet.Name
        selected_cell_address = _get_cell_address_for_log(active_cell)
        _log_fast_scan_context(logger, target_workbook, workbook_path, selected_sheet_name, selected_cell_address)

        scan_start = time.perf_counter()
        _log(logger, "info", "使用 openpyxl 扫描目标工作簿颜色区域。")
        color_plan = scan_target_workbook_color_plan_with_openpyxl(
            workbook_path=workbook_path,
            selected_sheet_name=selected_sheet_name,
            selected_cell_address=selected_cell_address,
            scope="all_matching_sheets",
        )
        scan_seconds = time.perf_counter() - scan_start
        _log_target_scan_summary(logger, color_plan, scan_seconds)

        sheet_plans = _prepare_sheet_plans_for_write(target_workbook, color_plan)
        _log(logger, "info", f"汇总范围：所有匹配 sheet")
        _log(logger, "info", f"写入方式：{_get_write_mode_label(normalized_write_mode)}")
        _log(logger, "info", f"源文件数量：{len(source_paths)}")

        target_excel_state = _set_excel_fast_mode(target_excel)

        aggregation = _aggregate_sources_with_openpyxl(
            source_paths=source_paths,
            sheet_plans=sheet_plans,
            write_mode=normalized_write_mode,
            logger=logger,
        )

        if not any(aggregation["matched_sources_by_sheet"].values()):
            raise RuntimeError("所有有效源文件都没有找到任何匹配的同名 sheet。")

        sheet_results = []
        skipped_write_sheets = []
        for plan in sheet_plans:
            sheet_name = plan["sheet_name"]
            matched_sources = aggregation["matched_sources_by_sheet"][sheet_name]
            if not matched_sources:
                skipped_write_sheets.append(f"{sheet_name}：所有源文件都缺少同名工作表")
                continue

            write_groups = group_cells_by_contiguous_rows(plan["cells"])
            _log(
                logger,
                "info",
                f"正在写入目标工作表：{sheet_name}；待写入单元格 {len(plan['cells'])} 个，批量区域 {len(write_groups)} 组。",
            )
            written_cell_count = _write_target_cells(
                target_sheet=plan["sheet"],
                cell_groups=write_groups,
                write_mode=normalized_write_mode,
                totals=aggregation["totals_by_sheet"][sheet_name],
                matched_sources=matched_sources,
            )
            sheet_results.append({
                "sheet_name": sheet_name,
                "matched_cell_count": len(plan["cells"]),
                "written_cell_count": written_cell_count,
                "matched_source_sheet_count": len(matched_sources),
                "missing_sheet_file_count": aggregation["missing_count_by_sheet"][sheet_name],
                "ignored_non_numeric_count": aggregation["ignored_count_by_sheet"][sheet_name],
            })

        _log_limited_items(logger, "跳过写入目标工作表", skipped_write_sheets)

        summary = summarize_sheet_results(sheet_results)
        _log(logger, "info", f"写入 sheet 数：{summary['matched_sheet_count']}")
        _log(logger, "info", f"写入单元格数：{summary['written_cell_count']}")
        return {
            "target_workbook_name": target_workbook.Name,
            "target_sheet_name": "",
            "scope": "all_matching_sheets",
            "matched_sheet_count": color_plan["matched_sheet_count"],
            "written_sheet_count": summary["matched_sheet_count"],
            "matched_cell_count": color_plan["matched_cell_count"],
            "source_file_count": len(source_paths),
            "valid_source_file_count": aggregation["valid_source_file_count"],
            "skipped_source_file_count": aggregation["skipped_source_file_count"],
            "matched_source_file_count": len(aggregation["matched_source_file_paths"]),
            "matched_source_sheet_count": summary["matched_source_sheet_count"],
            "missing_sheet_file_count": summary["missing_sheet_file_count"],
            "written_cell_count": summary["written_cell_count"],
            "ignored_non_numeric_count": summary["ignored_non_numeric_count"],
            "write_mode": normalized_write_mode,
            "target_scan_seconds": scan_seconds,
            "sheet_results": sheet_results,
        }

    finally:
        try:
            _restore_excel_state(target_excel, target_excel_state)
        except Exception:
            pass
        pythoncom.CoUninitialize()


def _prepare_sheet_plans_for_write(target_workbook, color_plan: dict) -> list[dict]:
    sheet_plans = []
    for sheet_plan in color_plan["sheet_plans"]:
        sheet_name = sheet_plan["sheet_name"]
        target_sheet = _find_worksheet_by_name(target_workbook, sheet_name)
        if target_sheet is None:
            raise ValueError(f"当前打开的目标工作簿中不存在 sheet：{sheet_name}。")

        sheet_plans.append({
            "sheet_name": sheet_name,
            "sheet": target_sheet,
            "cells": [_cell_address_to_row_col(address) for address in sheet_plan["cells"]],
            "cell_addresses": list(sheet_plan["cells"]),
        })
    return sheet_plans


def _aggregate_sources_with_openpyxl(
    source_paths: list[str],
    sheet_plans: list[dict],
    write_mode: str,
    logger=None,
) -> dict:
    totals_by_sheet = {
        plan["sheet_name"]: {cell: 0.0 for cell in plan["cells"]}
        for plan in sheet_plans
    }
    matched_sources_by_sheet = {plan["sheet_name"]: [] for plan in sheet_plans}
    missing_count_by_sheet = {plan["sheet_name"]: 0 for plan in sheet_plans}
    ignored_count_by_sheet = {plan["sheet_name"]: 0 for plan in sheet_plans}
    matched_source_file_paths: set[str] = set()
    valid_source_file_paths: set[str] = set()
    skipped_source_items = []
    skipped_sheet_items = []

    source_file_count = len(source_paths)
    for index, source_path in enumerate(source_paths, start=1):
        source_path = os.path.abspath(source_path)
        source_filename = os.path.basename(source_path)
        skip_reason = _get_fast_source_skip_reason(source_path)
        if skip_reason:
            skipped_source_items.append(skip_reason)
            continue

        _log(logger, "info", f"正在读取源文件 {index}/{source_file_count}：{source_filename}")
        try:
            _log(logger, "info", f"使用 openpyxl 只读读取源文件：{source_filename}")
            source_result = _read_multi_sheet_source_with_openpyxl(
                source_path=source_path,
                sheet_plans=sheet_plans,
                write_mode=write_mode,
                logger=logger,
            )
        except Exception as e:
            skipped_source_items.append(f"源文件 {source_filename} 读取失败，已跳过：{e}")
            continue

        valid_source_file_paths.add(source_path)
        source_has_matched_sheet = False
        for plan in sheet_plans:
            sheet_name = plan["sheet_name"]
            sheet_result = source_result.get(sheet_name, {"status": "missing_sheet"})
            if sheet_result["status"] == "missing_sheet":
                missing_count_by_sheet[sheet_name] += 1
                skipped_sheet_items.append(f"{source_filename} 缺少工作表 {sheet_name}")
                continue
            if sheet_result["status"] == "failed":
                skipped_sheet_items.append(f"{source_filename} 工作表 {sheet_name} 读取失败：{sheet_result['error']}")
                continue

            source_has_matched_sheet = True
            matched_sources_by_sheet[sheet_name].append((source_path, sheet_result["sheet_name"]))
            ignored_count_by_sheet[sheet_name] += sheet_result["ignored_non_numeric_count"]
            for cell in plan["cells"]:
                totals_by_sheet[sheet_name][cell] += sheet_result["totals"][cell]

        if source_has_matched_sheet:
            matched_source_file_paths.add(source_path)

    _log(logger, "info", f"有效源文件数量：{len(valid_source_file_paths)}")
    _log(logger, "info", f"跳过源文件数量：{len(skipped_source_items)}")
    _log_limited_items(logger, "跳过源文件", skipped_source_items)
    _log_limited_items(logger, "缺少或跳过的源工作表", skipped_sheet_items)

    if not valid_source_file_paths:
        raise ValueError("没有可用于汇总的有效源文件。")

    return {
        "totals_by_sheet": totals_by_sheet,
        "matched_sources_by_sheet": matched_sources_by_sheet,
        "missing_count_by_sheet": missing_count_by_sheet,
        "ignored_count_by_sheet": ignored_count_by_sheet,
        "valid_source_file_count": len(valid_source_file_paths),
        "skipped_source_file_count": len(skipped_source_items),
        "matched_source_file_paths": matched_source_file_paths,
    }


def _escape_excel_quote(value: object) -> str:
    return str(value).replace("'", "''")


def _build_external_cell_ref(source_file_path: str, sheet_name: str, cell_address: str) -> str:
    abs_path = os.path.abspath(source_file_path)
    directory, filename = os.path.split(abs_path)
    workbook_ref = os.path.join(directory, f"[{filename}]") if directory else f"[{filename}]"
    quoted_sheet_ref = _escape_excel_quote(f"{workbook_ref}{sheet_name}")
    return f"'{quoted_sheet_ref}'!{str(cell_address).strip()}"


def _log(logger, level: str, message: str) -> None:
    if logger:
        getattr(logger, level)(message)


def _log_limited_items(logger, title: str, items: list[str], limit: int = 10) -> None:
    if not items:
        return

    _log(logger, "info", f"{title}：共 {len(items)} 项。")
    for item in build_limited_log_items(items, limit=limit):
        _log(logger, "info", f"{title}：{item}")


def _get_workbook_full_name(workbook) -> str:
    full_name = _get_com_property(workbook, "FullName")
    return "" if full_name is None else str(full_name)


def _log_fast_scan_context(
    logger,
    target_workbook,
    workbook_path: str,
    selected_sheet_name: str,
    selected_cell_address: str,
) -> None:
    _log(logger, "info", "当前 Excel 连接成功。")
    _log(logger, "info", f"目标工作簿：{_get_com_object_name(target_workbook)}")
    _log(logger, "info", f"当前目标工作簿路径：{workbook_path or '未保存'}")
    _log(logger, "info", f"当前工作表：{selected_sheet_name}")
    _log(logger, "info", f"当前选中单元格地址：{selected_cell_address}")


def _log_target_scan_summary(logger, color_plan: dict, scan_seconds: float) -> None:
    _log(logger, "info", f"目标扫描耗时：{scan_seconds:.2f} 秒。")
    _log(logger, "info", f"匹配 sheet 数：{color_plan['matched_sheet_count']}")
    _log(logger, "info", f"同色单元格总数：{color_plan['matched_cell_count']}")


def _cell_address_to_row_col(cell_address: str) -> tuple[int, int]:
    from openpyxl.utils.cell import coordinate_to_tuple

    return coordinate_to_tuple(_normalize_cell_address(cell_address))


def _get_fast_source_skip_reason(source_path: str) -> str | None:
    filename = os.path.basename(source_path)
    if filename.startswith("~$"):
        return f"源文件 {filename} 为临时文件，已跳过。"

    if not os.path.exists(source_path):
        return f"源文件 {filename} 不存在，已跳过。"

    _, ext = os.path.splitext(filename)
    ext = ext.lower()
    if ext == ".xls":
        return f"源文件 {filename} 为 .xls，暂不支持快速汇总，请另存为 xlsx/xlsm。"
    if ext not in OPENPYXL_EXTENSIONS:
        return f"源文件 {filename} 不是支持的 xlsx/xlsm 文件，已跳过。"

    return None


def _load_openpyxl_workbook(path: str):
    from openpyxl import load_workbook

    return load_workbook(path, read_only=True, data_only=True)


def _validate_target_workbook_path_for_color_scan(workbook_path: str) -> None:
    if not workbook_path or not os.path.exists(workbook_path):
        raise ValueError("快速颜色扫描需要读取当前工作簿文件，请先保存目标工作簿后重试。")

    _, ext = os.path.splitext(workbook_path)
    if ext.lower() not in OPENPYXL_EXTENSIONS:
        raise ValueError("当前目标工作簿格式不支持快速颜色扫描，请另存为 xlsx/xlsm 后重试。")


def _normalize_cell_address(cell_address: str) -> str:
    return str(cell_address).replace("$", "").strip()


def _scan_openpyxl_sheet_for_fill_key(sheet, selected_color_key: tuple) -> list[str]:
    cells = []
    for row in sheet.iter_rows():
        for cell in row:
            if get_openpyxl_fill_key(cell) == selected_color_key:
                cells.append(cell.coordinate)
    return cells


def _find_openpyxl_sheet_name(sheet_names: list[str], sheet_name: str) -> str | None:
    expected_name = _normalize_sheet_name(sheet_name)
    for actual_name in sheet_names:
        if _normalize_sheet_name(actual_name) == expected_name:
            return actual_name
    return None


def _get_openpyxl_sheet_name(path: str, sheet_name: str) -> str | None:
    workbook = _load_openpyxl_workbook(path)
    try:
        return _find_openpyxl_sheet_name(workbook.sheetnames, sheet_name)
    finally:
        workbook.close()


def _read_openpyxl_cells(sheet, cells: list[tuple[int, int]]) -> dict[tuple[int, int], object]:
    result = {cell: None for cell in cells}
    if not cells:
        return result

    min_row = min(row for row, _ in cells)
    max_row = max(row for row, _ in cells)
    min_col = min(col for _, col in cells)
    max_col = max(col for _, col in cells)
    target_cols_by_row: dict[int, set[int]] = {}
    for row, col in cells:
        target_cols_by_row.setdefault(row, set()).add(col)

    for row_offset, row_values in enumerate(
        sheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        )
    ):
        actual_row = min_row + row_offset
        target_cols = target_cols_by_row.get(actual_row)
        if not target_cols:
            continue
        for col in target_cols:
            value_index = col - min_col
            if value_index < len(row_values):
                result[(actual_row, col)] = row_values[value_index]

    return result


def _set_excel_fast_mode(excel) -> dict:
    if excel is None:
        return {}

    state = {}
    for attr, value in {
        "ScreenUpdating": False,
        "EnableEvents": False,
        "DisplayAlerts": False,
    }.items():
        try:
            state[attr] = getattr(excel, attr)
            setattr(excel, attr, value)
        except Exception:
            pass
    return state


def _restore_excel_state(excel, state: dict) -> None:
    if excel is None:
        return
    for attr, value in state.items():
        try:
            setattr(excel, attr, value)
        except Exception:
            pass


def _get_write_mode_label(write_mode: str) -> str:
    if write_mode == "formula":
        return "公式"
    return "数值"


def _get_active_excel(win32com_client, pythoncom_module, logger=None):
    attempts: list[str] = []

    connectors = [
        (
            "win32com.client.GetActiveObject",
            lambda: [(win32com_client.GetActiveObject("Excel.Application"), None)],
        ),
        (
            "pythoncom.GetActiveObject",
            lambda: [(win32com_client.Dispatch(pythoncom_module.GetActiveObject("Excel.Application")), None)],
        ),
        (
            "运行对象表",
            lambda: _get_excel_candidates_from_rot(win32com_client, pythoncom_module),
        ),
        (
            "win32com.client.Dispatch",
            lambda: [(win32com_client.Dispatch("Excel.Application"), None)],
        ),
    ]

    for method_name, connect in connectors:
        try:
            candidates = connect()
        except Exception as e:
            detail = f"{method_name} 失败：{_format_exception(e)}"
            attempts.append(detail)
            _log(logger, "info", detail)
            continue

        if not candidates:
            detail = f"{method_name} 未找到可用的 Excel.Application"
            attempts.append(detail)
            _log(logger, "info", detail)
            continue

        for excel, source_name in candidates:
            has_workbook, workbook_detail = _excel_has_active_workbook(excel)
            source_detail = f"，来源：{source_name}" if source_name else ""
            detail = f"{method_name} 成功{source_detail}，{workbook_detail}"
            attempts.append(detail)
            _log(logger, "info", detail)
            if has_workbook:
                _log(logger, "info", f"当前连接到的 Excel 应用：成功，方式：{method_name}")
                return excel

            if method_name == "win32com.client.Dispatch":
                _quit_empty_hidden_excel(excel)

    raise RuntimeError(
        "无法连接到当前 Excel 实例或无法读取活动工作簿。"
        "请确认目标 Excel 已打开，未处于单元格编辑、弹窗或受保护视图状态，"
        "并确认工具和 Excel 以相同权限运行。"
        f"连接诊断：{'；'.join(attempts)}"
    )


def _get_excel_candidates_from_rot(win32com_client, pythoncom_module) -> list[tuple[object, str]]:
    candidates: list[tuple[object, str]] = []
    rot = pythoncom_module.GetRunningObjectTable()
    bind_context = pythoncom_module.CreateBindCtx(0)
    enum_moniker = rot.EnumRunning()

    while True:
        monikers = enum_moniker.Next(1)
        if not monikers:
            break

        moniker = monikers[0]
        try:
            display_name = moniker.GetDisplayName(bind_context, None)
        except Exception:
            display_name = ""

        if not _looks_like_excel_rot_entry(display_name):
            continue

        try:
            running_object = rot.GetObject(moniker)
        except Exception:
            try:
                running_object = moniker.BindToObject(bind_context, None, pythoncom_module.IID_IDispatch)
            except Exception:
                continue

        try:
            dispatch = win32com_client.Dispatch(running_object)
        except Exception:
            continue

        excel = _get_com_property(dispatch, "Application") or dispatch
        candidates.append((excel, display_name or "ROT"))

    return candidates


def _looks_like_excel_rot_entry(display_name: str) -> bool:
    lower_name = str(display_name).lower()
    return (
        "excel" in lower_name
        or lower_name.endswith(".xlsx")
        or lower_name.endswith(".xlsm")
        or lower_name.endswith(".xls")
    )


def _excel_has_active_workbook(excel) -> tuple[bool, str]:
    try:
        workbook = excel.ActiveWorkbook
    except Exception as e:
        return False, f"ActiveWorkbook 读取失败：{_format_exception(e)}"

    if workbook is None:
        return False, "ActiveWorkbook 为空"

    workbook_name = _get_com_object_name(workbook)
    return True, f"ActiveWorkbook.Name：{workbook_name}"


def _quit_empty_hidden_excel(excel) -> None:
    try:
        workbooks = excel.Workbooks
        workbook_count = int(workbooks.Count)
        visible = bool(excel.Visible)
    except Exception:
        return

    if workbook_count == 0 and not visible:
        try:
            excel.Quit()
        except Exception:
            pass


def _format_exception(error: Exception) -> str:
    return f"{type(error).__name__}: {error}"


def _get_active_target_context(excel, logger=None):
    target_workbook = _get_com_property(excel, "ActiveWorkbook")
    active_sheet = _get_com_property(excel, "ActiveSheet")
    active_cell = _get_selected_top_left_cell(excel)

    if active_sheet is None and active_cell is not None:
        active_sheet = _get_com_property(active_cell, "Worksheet")
    if target_workbook is None and active_sheet is not None:
        target_workbook = _get_com_property(active_sheet, "Parent")

    _log(logger, "info", f"ActiveWorkbook.Name：{_get_com_object_name(target_workbook)}")
    _log(logger, "info", f"ActiveSheet.Name：{_get_com_object_name(active_sheet)}")
    _log(logger, "info", f"ActiveCell.Address：{_get_cell_address_for_log(active_cell)}")

    if target_workbook is None or active_sheet is None or active_cell is None:
        raise RuntimeError("请先打开目标合并工作簿，并选中一个带颜色的单元格。")

    return target_workbook, active_sheet, active_cell


def _get_com_property(obj, attr: str):
    if obj is None:
        return None
    try:
        return getattr(obj, attr)
    except Exception:
        return None


def _get_com_object_name(obj) -> str:
    name = _get_com_property(obj, "Name")
    if name is None:
        return "无法获取"
    return str(name)


def _get_selected_top_left_cell(excel):
    selection = _get_com_property(excel, "Selection")
    cell = _get_first_cell_from_range(selection)
    if cell is not None:
        return cell

    return _get_com_property(excel, "ActiveCell")


def _get_first_cell_from_range(range_obj):
    if range_obj is None:
        return None

    cells = _get_com_property(range_obj, "Cells")
    if cells is None:
        return None

    try:
        return cells(1, 1)
    except Exception:
        pass

    item = _get_com_property(cells, "Item")
    if item is None:
        return None
    try:
        return item(1, 1)
    except Exception:
        return None


def _get_cell_address_for_log(cell) -> str:
    if cell is None:
        return "无法获取"

    try:
        return build_cell_address(int(cell.Row), int(cell.Column))
    except Exception:
        address = _get_com_property(cell, "Address")
        if address is None:
            return "无法获取"
        return str(address)


def _normalize_sheet_name(sheet_name: object) -> str:
    if sheet_name is None:
        return ""
    return str(sheet_name).strip().lower()


def _find_worksheet_by_name(workbook, sheet_name: str):
    expected_name = _normalize_sheet_name(sheet_name)
    for sheet in workbook.Worksheets:
        if _normalize_sheet_name(sheet.Name) == expected_name:
            return sheet
    return None


def _resolve_target_sheet(workbook, active_sheet, target_sheet_name: str | None):
    requested_name = "" if target_sheet_name is None else str(target_sheet_name).strip()
    if not requested_name:
        return active_sheet

    target_sheet = _find_worksheet_by_name(workbook, requested_name)
    if target_sheet is None:
        raise ValueError(f"目标工作簿中不存在 sheet：{requested_name}。")
    return target_sheet


def _get_cell_fill_color(cell):
    fill_info = _get_cell_fill_color_info(cell)
    return fill_info["color"] if fill_info["has_fill"] else None


def _get_cell_fill_color_info(cell) -> dict:
    fill_info = {
        "color": None,
        "color_index": None,
        "pattern": None,
        "has_fill": False,
    }
    interior = _get_com_property(cell, "Interior")
    if interior is None:
        return fill_info

    fill_info["color"] = _get_com_property(interior, "Color")
    fill_info["color_index"] = _get_com_property(interior, "ColorIndex")
    fill_info["pattern"] = _get_com_property(interior, "Pattern")
    fill_info["has_fill"] = _has_effective_fill_color(
        fill_info["color"],
        fill_info["color_index"],
        fill_info["pattern"],
    )
    return fill_info


def _has_effective_fill_color(color, color_index, pattern) -> bool:
    if _same_excel_constant(color_index, XL_COLOR_INDEX_NONE):
        return False
    if color is not None:
        return True
    if _same_excel_constant(pattern, XL_PATTERN_NONE):
        return False
    return False


def _same_excel_constant(value, constant: int) -> bool:
    try:
        return int(value) == constant
    except (TypeError, ValueError):
        return False


def _log_active_cell_diagnostics(logger, target_workbook, active_sheet, active_cell, fill_info: dict) -> None:
    _log(logger, "info", f"当前活动工作簿：{_get_com_object_name(target_workbook)}")
    _log(logger, "info", f"当前活动工作表：{_get_com_object_name(active_sheet)}")
    _log(logger, "info", f"当前选中单元格：{_get_cell_address_for_log(active_cell)}")
    _log(logger, "info", f"当前选中单元格 Interior.Color：{fill_info['color']}")
    _log(logger, "info", f"当前选中单元格 Interior.ColorIndex：{fill_info['color_index']}")
    _log(logger, "info", f"当前选中单元格 Interior.Pattern：{fill_info['pattern']}")
    fill_status = "有填充色" if fill_info["has_fill"] else "无填充色"
    _log(logger, "info", f"当前选中单元格填充色判定：{fill_status}")


def _cell_address_from_row_column(row: int, column: int) -> str:
    return build_cell_address(row, column)


def _column_number_to_letters(column: int) -> str:
    return column_number_to_letter(column)


def _is_blank_value(value: object) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _accumulate_values_from_mapping(
    source_values: dict[tuple[int, int], object],
    totals: dict[tuple[int, int], float],
    source_path: str,
    sheet_name: str,
    logger=None,
) -> int:
    ignored_non_numeric_count = 0
    for row, col in totals:
        value = source_values.get((row, col))
        number = to_number_or_none(value)
        if number is None:
            if not _is_blank_value(value):
                ignored_non_numeric_count += 1
                address = build_cell_address(row, col)
                _log(logger, "info", f"非数字已忽略：{os.path.basename(source_path)} {sheet_name}!{address} = {value}")
            continue
        totals[(row, col)] += number
    return ignored_non_numeric_count


def _read_multi_sheet_source_with_openpyxl(
    source_path: str,
    sheet_plans: list[dict],
    write_mode: str,
    logger=None,
) -> dict:
    source_filename = os.path.basename(source_path)
    workbook = _load_openpyxl_workbook(source_path)
    try:
        result = {}
        for plan in sheet_plans:
            sheet_name = plan["sheet_name"]
            totals = {cell: 0.0 for cell in plan["cells"]}
            actual_sheet_name = _find_openpyxl_sheet_name(workbook.sheetnames, sheet_name)
            if actual_sheet_name is None:
                result[sheet_name] = {
                    "status": "missing_sheet",
                    "sheet_name": None,
                    "totals": totals,
                    "ignored_non_numeric_count": 0,
                    "error": None,
                }
                continue

            if len(sheet_plans) == 1:
                _log(logger, "info", f"找到同名工作表：{source_filename} -> {actual_sheet_name}")
            ignored_non_numeric_count = 0
            if write_mode == "value":
                source_values = _read_openpyxl_cells(workbook[actual_sheet_name], plan["cells"])
                ignored_non_numeric_count = _accumulate_values_from_mapping(
                    source_values=source_values,
                    totals=totals,
                    source_path=source_path,
                    sheet_name=actual_sheet_name,
                    logger=logger,
                )

            result[sheet_name] = {
                "status": "ok",
                "sheet_name": actual_sheet_name,
                "totals": totals,
                "ignored_non_numeric_count": ignored_non_numeric_count,
                "error": None,
            }
        return result
    finally:
        workbook.close()


def _write_target_cells(
    target_sheet,
    cell_groups: list[list[tuple[int, int]]],
    write_mode: str,
    totals: dict[tuple[int, int], float],
    matched_sources: list[tuple[str, str]],
) -> int:
    written_cell_count = 0
    for group in cell_groups:
        row = group[0][0]
        start_col = group[0][1]
        end_col = group[-1][1]
        row_values = [_build_write_value(row, col, write_mode, totals, matched_sources) for row, col in group]

        if len(group) == 1:
            target_cell = target_sheet.Cells(row, start_col)
            _write_single_cell(target_cell, write_mode, row_values[0])
        else:
            target_range = target_sheet.Range(
                target_sheet.Cells(row, start_col),
                target_sheet.Cells(row, end_col),
            )
            _write_range_row(target_range, write_mode, row_values)
        written_cell_count += len(group)

    return written_cell_count


def _build_write_value(
    row: int,
    col: int,
    write_mode: str,
    totals: dict[tuple[int, int], float],
    matched_sources: list[tuple[str, str]],
):
    if write_mode == "formula":
        address = build_cell_address(row, col)
        source_refs = [
            (source_path, source_sheet_name, address)
            for source_path, source_sheet_name in matched_sources
        ]
        return build_external_sum_formula(source_refs)
    return totals[(row, col)]


def _write_single_cell(target_cell, write_mode: str, value) -> None:
    if write_mode == "formula":
        target_cell.Formula = value
    else:
        try:
            target_cell.Value2 = value
        except Exception:
            target_cell.Value = value


def _write_range_row(target_range, write_mode: str, values: list[object]) -> None:
    row_values = (tuple(values),)
    if write_mode == "formula":
        target_range.Formula = row_values
    else:
        try:
            target_range.Value2 = row_values
        except Exception:
            target_range.Value = row_values

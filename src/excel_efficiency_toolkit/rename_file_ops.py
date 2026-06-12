import os
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


INSTRUCTION_SHEET_NAME = "使用说明"
SETTINGS_SHEET_NAME = "参数设置"
RULE_SHEET_NAME = "重命名清单"
LOG_SHEET_NAME = "处理日志"

RULE_HEADERS = [
    "原文件路径",
    "原文件名（不含后缀）",
    "新文件名（不含后缀）",
    "后缀名",
    "目标文件路径预览",
    "状态",
    "说明",
    "原后缀名（系统列）",
]
LOG_HEADERS = ["处理时间", "原文件路径", "目标文件路径", "原文件名", "新文件名", "状态", "说明"]
ILLEGAL_FILENAME_PATTERN = re.compile(r'[<>:"/\\|?*]')


@dataclass
class RenameRule:
    row_number: int
    original_path: str
    original_name: str
    new_name_raw: str
    extension_raw: str
    original_extension: str


@dataclass
class RenameSettings:
    existing_target_mode: str = "自动编号"
    skip_temp_files: bool = True
    clean_illegal_chars: bool = True
    warnings: list[str] | None = None


@dataclass
class RenameAction:
    row_number: int
    original_path: str
    target_path: str
    status: str
    message: str
    final_name: str


def create_rename_rule_workbook(file_paths: list[str], output_dir: str | None = None) -> str:
    valid_paths = [str(Path(path).resolve()) for path in file_paths if str(path).strip()]
    if not valid_paths:
        raise ValueError("请选择至少一个需要重命名的文件。")

    target_dir = Path(output_dir).resolve() if output_dir else _infer_output_dir(valid_paths)
    target_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = _get_unique_path(target_dir / f"批量重命名规则_{timestamp}.xlsx")

    workbook = Workbook()
    instruction_sheet = workbook.active
    instruction_sheet.title = INSTRUCTION_SHEET_NAME
    settings_sheet = workbook.create_sheet(SETTINGS_SHEET_NAME)
    rule_sheet = workbook.create_sheet(RULE_SHEET_NAME)
    log_sheet = workbook.create_sheet(LOG_SHEET_NAME)

    _write_instruction_sheet(instruction_sheet)
    _write_settings_sheet(settings_sheet)
    _write_rule_sheet(rule_sheet, valid_paths)
    _write_log_sheet(log_sheet)

    workbook.active = workbook.sheetnames.index(RULE_SHEET_NAME)
    workbook.save(output_path)
    workbook.close()
    return str(output_path)


def read_rename_settings(rule_workbook_path: str) -> RenameSettings:
    workbook = load_workbook(rule_workbook_path, data_only=True)
    warnings: list[str] = []
    try:
        sheet = workbook[SETTINGS_SHEET_NAME]
        values = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            key = _cell_to_text(row[0] if len(row) > 0 else None)
            value = _cell_to_text(row[1] if len(row) > 1 else None)
            if key:
                values[key] = value

        existing_target_mode = values.get("目标文件已存在处理方式", "").strip()
        if existing_target_mode not in {"自动编号", "跳过"}:
            if existing_target_mode:
                warnings.append("参数“目标文件已存在处理方式”无效，已使用默认值“自动编号”。")
            existing_target_mode = "自动编号"

        skip_temp_files = _read_yes_no_setting(
            values.get("是否跳过临时文件", ""),
            default=True,
            setting_name="是否跳过临时文件",
            warnings=warnings,
        )
        clean_illegal_chars = _read_yes_no_setting(
            values.get("是否清洗非法字符", ""),
            default=True,
            setting_name="是否清洗非法字符",
            warnings=warnings,
        )

        return RenameSettings(
            existing_target_mode=existing_target_mode,
            skip_temp_files=skip_temp_files,
            clean_illegal_chars=clean_illegal_chars,
            warnings=warnings,
        )
    finally:
        workbook.close()


def read_rename_rules(rule_workbook_path: str) -> list[RenameRule]:
    workbook = load_workbook(rule_workbook_path, data_only=True)
    try:
        sheet = workbook[RULE_SHEET_NAME]
        rules: list[RenameRule] = []
        for row_number in range(2, sheet.max_row + 1):
            original_path = _cell_to_text(sheet.cell(row=row_number, column=1).value)
            original_name = _cell_to_text(sheet.cell(row=row_number, column=2).value)
            new_name_raw = _cell_to_text(sheet.cell(row=row_number, column=3).value)
            extension_raw = _cell_to_text(sheet.cell(row=row_number, column=4).value)
            original_extension = _cell_to_text(sheet.cell(row=row_number, column=8).value)
            rules.append(
                RenameRule(
                    row_number=row_number,
                    original_path=original_path,
                    original_name=original_name,
                    new_name_raw=new_name_raw,
                    extension_raw=extension_raw,
                    original_extension=original_extension,
                )
            )
        return rules
    finally:
        workbook.close()


def build_rename_plan(rules: list[RenameRule], settings: RenameSettings) -> list[RenameAction]:
    base_actions = [_build_base_action(rule, settings) for rule in rules]
    original_paths_to_be_renamed = _success_original_paths(base_actions)

    actions = base_actions
    for _ in range(len(base_actions) + 1):
        actions = _resolve_target_conflicts(base_actions, settings, original_paths_to_be_renamed)
        next_original_paths = _success_original_paths(actions)
        if next_original_paths == original_paths_to_be_renamed:
            return actions
        original_paths_to_be_renamed = next_original_paths

    return actions


def execute_rename_plan(actions: list[RenameAction]) -> dict:
    success_count = 0
    skipped_count = len([action for action in actions if action.status == "跳过"])
    failed_count = 0
    executable_actions = [action for action in actions if action.status == "成功"]
    moved_actions: list[tuple[RenameAction, Path, Path]] = []
    reserved_temp_paths: set[str] = set()

    for index, action in enumerate(executable_actions, start=1):
        original_path = Path(action.original_path)
        if not original_path.exists():
            action.status = "跳过"
            action.message = _join_messages(action.message, "原文件不存在")
            skipped_count += 1
            continue

        temp_path = _build_temp_path(original_path, index, reserved_temp_paths)
        reserved_temp_paths.add(_normalize_path_for_compare(str(temp_path)))
        try:
            original_path.rename(temp_path)
            moved_actions.append((action, original_path, temp_path))
        except Exception as e:
            action.status = "失败"
            action.message = _join_messages(action.message, f"第一阶段临时重命名失败：{e}")
            failed_count += 1

    for action, original_path, temp_path in moved_actions:
        target_path = Path(action.target_path)
        try:
            if target_path.exists():
                raise FileExistsError("目标文件已存在，未覆盖")
            temp_path.rename(target_path)
            action.message = "已重命名" if action.message == "已重命名" else _join_messages(action.message, "已重命名")
            success_count += 1
        except Exception as e:
            rollback_message = _rollback_temp_file(temp_path, original_path)
            action.status = "失败"
            action.message = _join_messages(action.message, f"第二阶段重命名失败：{e}", rollback_message)
            failed_count += 1

    return {
        "success_count": success_count,
        "skipped_count": skipped_count,
        "failed_count": failed_count,
    }


def _rollback_temp_file(temp_path: Path, original_path: Path) -> str:
    if not temp_path.exists():
        return ""
    if original_path.exists():
        return f"临时文件保留在：{temp_path}"
    try:
        temp_path.rename(original_path)
        return "已回滚到原文件名"
    except Exception as rollback_error:
        return f"回滚失败，临时文件保留在：{temp_path}；回滚错误：{rollback_error}"


def _build_temp_path(original_path: Path, index: int, reserved_temp_paths: set[str]) -> Path:
    parent = original_path.parent
    suffix = original_path.suffix
    while True:
        candidate = parent / f".__rename_tmp_{uuid4().hex}_{index}{suffix}"
        normalized_candidate = _normalize_path_for_compare(str(candidate))
        if not candidate.exists() and normalized_candidate not in reserved_temp_paths:
            return candidate
        index += 1


def _success_original_paths(actions: list[RenameAction]) -> set[str]:
    return {
        _normalize_path_for_compare(action.original_path)
        for action in actions
        if action.status == "成功"
    }


def _resolve_target_conflicts(
    base_actions: list[RenameAction],
    settings: RenameSettings,
    original_paths_to_be_renamed: set[str],
) -> list[RenameAction]:
    actions: list[RenameAction] = []
    assigned_targets: set[str] = set()

    for base_action in base_actions:
        action = _copy_action(base_action)
        if action.status != "成功":
            actions.append(action)
            continue

        target_path = Path(action.target_path)
        target_compare = _normalize_path_for_compare(str(target_path))
        target_exists_external = target_path.exists() and target_compare not in original_paths_to_be_renamed
        target_assigned = target_compare in assigned_targets
        if target_exists_external or target_assigned:
            if settings.existing_target_mode == "跳过":
                action.status = "跳过"
                action.message = "目标文件已存在" if target_exists_external else "目标文件名冲突"
                actions.append(action)
                continue

            numbered_target_path = _build_numbered_target_path(
                target_path,
                assigned_targets,
                original_paths_to_be_renamed,
            )
            action.target_path = str(numbered_target_path)
            action.final_name = numbered_target_path.name
            action.message = _join_messages(
                action.message if action.message != "已重命名" else "",
                "目标文件名冲突，已自动编号" if target_assigned else "目标文件已存在，已自动编号",
            ) or "已重命名"
            target_compare = _normalize_path_for_compare(action.target_path)

        assigned_targets.add(target_compare)
        actions.append(action)

    return actions


def write_rename_results_to_workbook(rule_workbook_path: str, actions: list[RenameAction]) -> None:
    workbook = load_workbook(rule_workbook_path)
    try:
        rule_sheet = workbook[RULE_SHEET_NAME]
        log_sheet = workbook[LOG_SHEET_NAME]
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for action in actions:
            rule_sheet.cell(row=action.row_number, column=5, value=action.target_path)
            rule_sheet.cell(row=action.row_number, column=6, value=action.status)
            rule_sheet.cell(row=action.row_number, column=7, value=action.message)
            log_sheet.append(
                [
                    now,
                    action.original_path,
                    action.target_path,
                    Path(action.original_path).stem if action.original_path else "",
                    action.final_name,
                    action.status,
                    action.message,
                ]
            )

        workbook.save(rule_workbook_path)
    finally:
        workbook.close()


def _infer_output_dir(file_paths: list[str]) -> Path:
    parent_dirs = [str(Path(path).parent) for path in file_paths]
    try:
        common_dir = os.path.commonpath(parent_dirs)
    except ValueError:
        common_dir = parent_dirs[0]
    if common_dir and Path(common_dir).exists():
        return Path(common_dir)
    return Path(file_paths[0]).parent


def _write_instruction_sheet(sheet) -> None:
    rows = [
        ["批量重命名文件使用说明"],
        ["1. 在【重命名清单】C列填写新文件名；不带后缀时，程序会拼接 D 列当前后缀。"],
        ["2. 如果 C 列已填写完整后缀，则以 C 列完整内容为准，不再拼接 D 列。"],
        ["3. D列为后缀名，默认使用原后缀，可修改；D列为空时，会回退使用原文件后缀。"],
        ["4. 修改后缀名只改变文件名，不转换文件格式。"],
        ["5. 保存规则表后，回到工具台点击“执行批量重命名”。"],
        ["6. 程序不会覆盖已有文件。"],
        ["7. 目标文件名重复时，根据【参数设置】自动编号或跳过。"],
        ["8. 执行后会在 F/G 列写入状态和说明，并在【处理日志】记录明细。"],
    ]
    for row in rows:
        sheet.append(row)
    sheet["A1"].font = Font(bold=True)
    sheet.column_dimensions["A"].width = 92


def _write_settings_sheet(sheet) -> None:
    rows = [
        ["参数项", "参数值", "说明"],
        ["目标文件已存在处理方式", "自动编号", "自动编号会在文件名后追加 _1、_2；跳过则不重命名。"],
        ["是否跳过临时文件", "是", "跳过以 ~$ 开头的 Office 临时文件。"],
        ["是否清洗非法字符", "是", "将 Windows 文件名非法字符替换为下划线；如果为否，遇到非法字符则跳过。"],
    ]
    _append_rows(sheet, rows)
    _style_header(sheet, 1, len(rows[0]))
    widths = {"A": 24, "B": 18, "C": 72}
    _set_widths(sheet, widths)


def _write_rule_sheet(sheet, file_paths: list[str]) -> None:
    _append_rows(sheet, [RULE_HEADERS])
    _style_header(sheet, 1, len(RULE_HEADERS))
    fill = PatternFill(fill_type="solid", fgColor="FFF2CC")

    for file_path in file_paths:
        path = Path(file_path)
        extension = path.suffix
        sheet.append([str(path), path.stem, "", extension, "", "", "", extension])

    for row in range(2, sheet.max_row + 1):
        sheet.cell(row=row, column=3).fill = fill
        sheet.cell(row=row, column=4).fill = fill

    sheet.freeze_panes = "A2"
    _set_widths(
        sheet,
        {
            "A": 58,
            "B": 30,
            "C": 28,
            "D": 14,
            "E": 58,
            "F": 12,
            "G": 42,
            "H": 18,
        },
    )


def _write_log_sheet(sheet) -> None:
    _append_rows(sheet, [LOG_HEADERS])
    _style_header(sheet, 1, len(LOG_HEADERS))
    sheet.freeze_panes = "A2"
    _set_widths(
        sheet,
        {
            "A": 20,
            "B": 58,
            "C": 58,
            "D": 30,
            "E": 30,
            "F": 12,
            "G": 42,
        },
    )


def _build_base_action(rule: RenameRule, settings: RenameSettings) -> RenameAction:
    original_path_text = rule.original_path.strip()
    if not original_path_text:
        return _skip_action(rule, "", "原文件路径为空")

    original_path = Path(original_path_text)
    original_name = rule.original_name.strip() or original_path.name
    if settings.skip_temp_files and original_name.startswith("~$"):
        return _skip_action(rule, "", "临时文件已跳过")
    if not original_path.exists():
        return _skip_action(rule, "", "原文件不存在")

    new_name_text = rule.new_name_raw.strip()
    if not new_name_text:
        return _skip_action(rule, "", "未填写新文件名")

    message_parts: list[str] = []
    new_stem, extension_from_name = _split_name_and_extension(new_name_text)
    extension = _normalize_extension(rule.extension_raw.strip())
    if extension_from_name:
        new_stem = new_stem.strip()
        extension = extension_from_name
        message_parts.append("已从新文件名中识别后缀")
    if not extension:
        extension = _normalize_extension(rule.original_extension.strip()) or original_path.suffix

    if not new_stem.strip():
        return _skip_action(rule, "", "新文件名为空")

    if ILLEGAL_FILENAME_PATTERN.search(new_stem) or ILLEGAL_FILENAME_PATTERN.search(extension):
        if settings.clean_illegal_chars:
            new_stem = ILLEGAL_FILENAME_PATTERN.sub("_", new_stem)
            extension = ILLEGAL_FILENAME_PATTERN.sub("_", extension)
            message_parts.append("已清洗非法字符")
        else:
            return _skip_action(rule, "", "文件名包含非法字符")

    final_name = f"{new_stem}{extension}"
    if not final_name.strip():
        return _skip_action(rule, "", "新文件名为空")

    target_path = original_path.with_name(final_name)
    if _normalize_path_for_compare(str(target_path)) == _normalize_path_for_compare(str(original_path)):
        return _skip_action(rule, str(target_path), "新旧文件名相同", final_name)

    return RenameAction(
        row_number=rule.row_number,
        original_path=str(original_path),
        target_path=str(target_path),
        status="成功",
        message=_join_messages(*message_parts) or "已重命名",
        final_name=final_name,
    )


def _build_numbered_target_path(
    target_path: Path,
    assigned_targets: set[str],
    original_paths_to_be_renamed: set[str],
) -> Path:
    stem = target_path.stem
    suffix = target_path.suffix
    parent = target_path.parent
    counter = 1
    while True:
        candidate = parent / f"{stem}_{counter}{suffix}"
        candidate_compare = _normalize_path_for_compare(str(candidate))
        candidate_exists_external = candidate.exists() and candidate_compare not in original_paths_to_be_renamed
        if not candidate_exists_external and candidate_compare not in assigned_targets:
            return candidate
        counter += 1


def _copy_action(action: RenameAction) -> RenameAction:
    return RenameAction(
        row_number=action.row_number,
        original_path=action.original_path,
        target_path=action.target_path,
        status=action.status,
        message=action.message,
        final_name=action.final_name,
    )


def _split_name_and_extension(name: str) -> tuple[str, str]:
    path = Path(name)
    if path.suffix:
        return path.stem, path.suffix
    return name, ""


def _normalize_extension(extension: str) -> str:
    value = extension.strip()
    if not value:
        return ""
    if value == ".":
        return ""
    return value if value.startswith(".") else f".{value}"


def _skip_action(rule: RenameRule, target_path: str, message: str, final_name: str = "") -> RenameAction:
    return RenameAction(
        row_number=rule.row_number,
        original_path=rule.original_path,
        target_path=target_path,
        status="跳过",
        message=message,
        final_name=final_name,
    )


def _read_yes_no_setting(value: str, default: bool, setting_name: str, warnings: list[str]) -> bool:
    normalized = value.strip().lower()
    if normalized in {"是", "yes", "y", "true", "1"}:
        return True
    if normalized in {"否", "no", "n", "false", "0"}:
        return False
    if normalized:
        default_text = "是" if default else "否"
        warnings.append(f"参数“{setting_name}”无效，已使用默认值“{default_text}”。")
    return default


def _append_rows(sheet, rows: list[list[object]]) -> None:
    for row in rows:
        sheet.append(row)


def _style_header(sheet, row_number: int, column_count: int) -> None:
    for column in range(1, column_count + 1):
        sheet.cell(row=row_number, column=column).font = Font(bold=True)


def _set_widths(sheet, widths: dict[str, int]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _cell_to_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _get_unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    counter = 2
    while True:
        candidate = path.with_name(f"{path.stem}_{counter}{path.suffix}")
        if not candidate.exists():
            return candidate
        counter += 1


def _normalize_path_for_compare(path: str) -> str:
    return os.path.normcase(os.path.abspath(path))


def _join_messages(*messages: str) -> str:
    return "；".join(message for message in messages if message)

import os

from openpyxl import Workbook, load_workbook

from .workbook_drill_ops import (
    RESULT_SHEET_BASE_NAME,
    is_excel_error_text,
    normalize_range_address,
    range_value_to_address_map,
    write_multi_file_result_sheet,
)


SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm"}
RESULT_SHEET_NAME = RESULT_SHEET_BASE_NAME
def is_supported_openpyxl_workbook(path: str) -> bool:
    return os.path.splitext(str(path))[1].lower() in SUPPORTED_EXTENSIONS


def is_office_temp_file(path: str) -> bool:
    return os.path.basename(str(path)).startswith("~$")


def build_unique_output_path(directory: str, base_name: str = "数据穿透结果", extension: str = ".xlsx") -> str:
    output_dir = os.path.abspath(str(directory))
    normalized_extension = extension if extension.startswith(".") else f".{extension}"
    candidate = os.path.join(output_dir, f"{base_name}{normalized_extension}")
    if not os.path.exists(candidate):
        return candidate

    counter = 2
    while True:
        candidate = os.path.join(output_dir, f"{base_name}_{counter}{normalized_extension}")
        if not os.path.exists(candidate):
            return candidate
        counter += 1


def normalize_sheet_lookup_name(name: str) -> str:
    return str(name or "").strip().casefold()


def resolve_sheet_name(workbook, target_sheet_name: str) -> tuple[str | None, list[str], str]:
    available_sheet_names = list(workbook.sheetnames)
    normalized_target_name = normalize_sheet_lookup_name(target_sheet_name)
    if not normalized_target_name:
        return None, available_sheet_names, "目标 Sheet 名为空"

    for sheet_name in available_sheet_names:
        if normalize_sheet_lookup_name(sheet_name) == normalized_target_name:
            return sheet_name, available_sheet_names, "匹配同名 Sheet"

    if available_sheet_names:
        available_text = "、".join(available_sheet_names)
        return None, available_sheet_names, f"缺少同名 Sheet：{target_sheet_name}；可用 Sheet：{available_text}"
    return None, available_sheet_names, f"缺少同名 Sheet：{target_sheet_name}；源文件没有工作表"


def is_excel_error_value(value) -> bool:
    return is_excel_error_text(value)


def read_drill_cell_from_workbook(source_path: str, sheet_name: str, cell_address: str) -> dict:
    return read_drill_range_from_workbook(source_path, sheet_name, cell_address)


def read_drill_range_from_workbook(source_path: str, sheet_name: str, range_address: str) -> dict:
    source_path = os.path.abspath(str(source_path))
    record = _new_record(source_path, sheet_name, range_address)

    skip_message = _get_source_skip_message(source_path)
    if skip_message is not None:
        status = "失败" if skip_message == "文件不存在" else "跳过"
        return _finish_record(record, None, status, skip_message)

    value_workbook = None
    formula_workbook = None
    try:
        value_workbook = load_workbook(source_path, read_only=True, data_only=True)
        formula_workbook = load_workbook(source_path, read_only=True, data_only=False)

        actual_sheet_name, available_sheet_names, message = resolve_sheet_name(value_workbook, sheet_name)
        if actual_sheet_name is None:
            return _finish_record(record, None, "跳过", message)

        value_sheet = value_workbook[actual_sheet_name]
        formula_sheet = formula_workbook[actual_sheet_name]
        normalized_range_address = normalize_range_address(range_address)
        value = _read_range_value(value_sheet, normalized_range_address)
        formula_value = _read_range_value(formula_sheet, normalized_range_address)
        values_by_address = range_value_to_address_map(normalized_range_address, value)
        formula_values_by_address = range_value_to_address_map(normalized_range_address, formula_value)
        cell_address = normalize_range_address(range_address).split(":")[0]
        record["values_by_address"] = values_by_address
        record["value"] = values_by_address.get(cell_address)

        if _has_formula_without_cached_value(values_by_address, formula_values_by_address):
            return _finish_record(
                record,
                record["value"],
                "成功",
                "公式缓存为空，取值可能需要先打开源文件计算并保存",
            )

        return _finish_record(record, record["value"], "成功", "读取成功")
    except Exception as e:
        return _finish_record(record, None, "失败", f"读取异常：{e}")
    finally:
        if value_workbook is not None:
            value_workbook.close()
        if formula_workbook is not None:
            formula_workbook.close()


def build_data_drill_records(source_paths: list[str], sheet_name: str, cell_address: str, logger=None) -> list[dict]:
    return build_data_drill_range_records(source_paths, sheet_name, cell_address, logger=logger)


def build_data_drill_range_records(source_paths: list[str], sheet_name: str, range_address: str, logger=None) -> list[dict]:
    records = []
    for index, source_path in enumerate(source_paths, start=1):
        _log(logger, "info", f"正在读取源文件 {index}/{len(source_paths)}：{source_path}")
        record = read_drill_range_from_workbook(source_path, sheet_name, range_address)
        records.append(record)
        _log(logger, "info", f"{record['source_file_name']}：{record['status']}，{record['message']}")
    return records


def summarize_data_drill_records(records: list[dict]) -> dict:
    summary = {"success_count": 0, "skipped_count": 0, "failed_count": 0, "total_count": len(records)}
    for record in records:
        status = record.get("status")
        if status == "成功":
            summary["success_count"] += 1
        elif status == "跳过":
            summary["skipped_count"] += 1
        elif status == "失败":
            summary["failed_count"] += 1
    return summary


def write_data_drill_result_workbook(
    records: list[dict],
    output_path: str,
    source_sheet_name: str,
    cell_address: str,
) -> str:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = RESULT_SHEET_NAME
    write_multi_file_result_sheet(sheet, records, cell_address)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    _set_result_column_widths(sheet)
    workbook.save(output_path)
    return output_path


def _get_source_skip_message(source_path: str) -> str | None:
    filename = os.path.basename(source_path)
    if is_office_temp_file(source_path):
        return "临时文件已跳过"
    if not os.path.exists(source_path):
        return "文件不存在"

    extension = os.path.splitext(filename)[1].lower()
    if extension == ".xls":
        return "请另存为 xlsx/xlsm 后再处理"
    if extension not in SUPPORTED_EXTENSIONS:
        return "不支持的文件格式"
    return None


def _new_record(source_path: str, sheet_name: str, cell_address: str) -> dict:
    normalized_address = normalize_range_address(cell_address)
    return {
        "source_file_name": os.path.basename(source_path),
        "source_file_path": source_path,
        "target_sheet_name": sheet_name,
        "cell_address": normalized_address.split(":")[0],
        "range_address": normalized_address,
        "value": None,
        "values_by_address": {},
        "status": "",
        "message": "",
    }


def _finish_record(record: dict, value, status: str, message: str) -> dict:
    record["value"] = value
    record["is_empty"] = _is_empty_value(value)
    record["is_error_value"] = is_excel_error_value(value)
    record["status"] = status
    record["message"] = message
    return record


def _is_formula_value(value) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _read_range_value(sheet, range_address: str):
    cells = sheet[range_address]
    if isinstance(cells, tuple):
        if cells and isinstance(cells[0], tuple):
            return tuple(tuple(cell.value for cell in row) for row in cells)
        return cells[0].value if cells else None
    return cells.value


def _has_formula_without_cached_value(values_by_address: dict[str, object], formula_values_by_address: dict[str, object]) -> bool:
    for address, formula_value in formula_values_by_address.items():
        if _is_formula_value(formula_value) and values_by_address.get(address) is None:
            return True
    return False


def _is_empty_value(value) -> bool:
    return value is None or value == ""


def _set_result_column_widths(sheet) -> None:
    widths = [24, 52, 18]
    for column_index, width in enumerate(widths, start=1):
        column_letter = chr(64 + column_index)
        sheet.column_dimensions[column_letter].width = width

    if sheet.max_column >= 5:
        for column_index in range(4, sheet.max_column - 1):
            column_letter = _column_index_to_letter(column_index)
            sheet.column_dimensions[column_letter].width = 14

    sheet.column_dimensions[_column_index_to_letter(sheet.max_column - 1)].width = 10
    sheet.column_dimensions[_column_index_to_letter(sheet.max_column)].width = 48


def _log(logger, level: str, message: str) -> None:
    if logger is None:
        return
    log_func = getattr(logger, level, None)
    if log_func is None:
        return
    log_func(message)


def _column_index_to_letter(column_index: int) -> str:
    result = ""
    current = column_index
    while current > 0:
        current, remainder = divmod(current - 1, 26)
        result = chr(65 + remainder) + result
    return result

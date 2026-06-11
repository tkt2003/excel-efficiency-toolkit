import os
import re
import shutil
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from .link_replace_ops import XL_EXCEL_LINKS, normalize_com_link_sources


DEFAULT_CHECKLIST_SHEET_NAME = "tb文件名"
EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
INVALID_FILENAME_PATTERN = re.compile(r'[<>:"/\\|?*]')


@dataclass
class ReportChecklistRecord:
    row_number: int
    tb_link_path: str
    output_name_raw: str
    company_name: str
    report_type: str


def read_report_checklist(
    template_path: str,
    sheet_name: str = DEFAULT_CHECKLIST_SHEET_NAME,
) -> list[ReportChecklistRecord]:
    workbook = load_workbook(template_path, read_only=True, data_only=True, keep_vba=_is_xlsm(template_path))
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"模板工作簿缺少清单 sheet：{sheet_name}")

        sheet = workbook[sheet_name]
        records: list[ReportChecklistRecord] = []
        for row_number in range(2, sheet.max_row + 1):
            tb_link_path = _cell_to_text(sheet.cell(row=row_number, column=1).value)
            if not tb_link_path:
                continue
            output_name_raw = _cell_to_text(sheet.cell(row=row_number, column=2).value)
            company_name = _cell_to_text(sheet.cell(row=row_number, column=3).value)
            report_type = _cell_to_text(sheet.cell(row=row_number, column=4).value)
            records.append(
                ReportChecklistRecord(
                    row_number=row_number,
                    tb_link_path=tb_link_path,
                    output_name_raw=output_name_raw,
                    company_name=company_name,
                    report_type=report_type,
                )
            )
        return records
    finally:
        workbook.close()


def sanitize_output_stem(name: str, row_number: int | None = None) -> str:
    text = str(name or "").strip()
    lower_text = text.lower()
    for extension in EXCEL_EXTENSIONS:
        if lower_text.endswith(extension):
            text = text[: -len(extension)]
            break
    text = INVALID_FILENAME_PATTERN.sub("_", text).strip(" .")
    if text:
        return text
    return f"报表_第{row_number or 1}行"


def build_unique_output_path(output_dir: str, output_stem: str, extension: str) -> str:
    directory = Path(output_dir).resolve()
    normalized_extension = extension if extension.startswith(".") else f".{extension}"
    candidate = directory / f"{output_stem}{normalized_extension}"
    if not candidate.exists():
        return str(candidate)

    counter = 2
    while True:
        candidate = directory / f"{output_stem}_{counter}{normalized_extension}"
        if not candidate.exists():
            return str(candidate)
        counter += 1


def summarize_report_records(records: list[dict]) -> dict:
    summary = {"success_count": 0, "skipped_count": 0, "failed_count": 0, "total_count": len(records)}
    for record in records:
        status = record.get("status")
        if status == "成功":
            summary["success_count"] += 1
        elif status == "跳过":
            summary["skipped_count"] += 1
        elif status == "失败":
            summary["failed_count"] += 1
    return summary


def read_workbook_external_links_with_com(workbook_path: str) -> list[str]:
    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False
        workbook = excel.Workbooks.Open(
            os.path.abspath(workbook_path),
            ReadOnly=True,
            UpdateLinks=0,
        )
        return normalize_com_link_sources(workbook.LinkSources(XL_EXCEL_LINKS))
    finally:
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()


def generate_reports_from_checklist(
    template_path: str,
    output_dir: str,
    sheet_name: str = DEFAULT_CHECKLIST_SHEET_NAME,
    logger=None,
) -> dict:
    abs_template_path = os.path.abspath(template_path)
    abs_output_dir = os.path.abspath(output_dir)
    if not os.path.exists(abs_template_path):
        raise FileNotFoundError(f"模板工作簿不存在：{abs_template_path}")
    if Path(abs_template_path).suffix.lower() not in EXCEL_EXTENSIONS:
        raise ValueError("模板工作簿仅支持 .xlsx / .xlsm / .xls。")
    os.makedirs(abs_output_dir, exist_ok=True)

    checklist_records = read_report_checklist(abs_template_path, sheet_name)
    template_links = read_workbook_external_links_with_com(abs_template_path)
    output_records = [_new_output_record(record) for record in checklist_records]

    if not checklist_records:
        return {
            "template_path": abs_template_path,
            "output_dir": abs_output_dir,
            "template_link_count": len(template_links),
            "records": [],
            **summarize_report_records([]),
        }

    if not template_links:
        for output_record in output_records:
            output_record["status"] = "跳过"
            output_record["message"] = "模板没有外部链接，已跳过"
        summary = summarize_report_records(output_records)
        return {
            "template_path": abs_template_path,
            "output_dir": abs_output_dir,
            "template_link_count": 0,
            "records": output_records,
            **summary,
        }

    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False

        for checklist_record, output_record in zip(checklist_records, output_records):
            workbook = None
            try:
                output_stem = sanitize_output_stem(checklist_record.output_name_raw, checklist_record.row_number)
                output_path = build_unique_output_path(
                    abs_output_dir,
                    output_stem,
                    Path(abs_template_path).suffix,
                )
                shutil.copy2(abs_template_path, output_path)
                output_record["output_path"] = output_path
                _log(logger, "info", f"正在生成报表：{output_path}")

                workbook = excel.Workbooks.Open(
                    os.path.abspath(output_path),
                    ReadOnly=False,
                    UpdateLinks=0,
                )
                for old_link_path in template_links:
                    workbook.ChangeLink(old_link_path, checklist_record.tb_link_path, XL_EXCEL_LINKS)
                workbook.Save()

                output_record["status"] = "成功"
                output_record["message"] = _build_success_message(len(template_links))
            except Exception as e:
                output_record["status"] = "失败"
                output_record["message"] = f"生成失败：{e}"
                _log(logger, "error", f"生成失败：第 {checklist_record.row_number} 行。详细信息：{e}")
            finally:
                if workbook is not None:
                    try:
                        workbook.Close(SaveChanges=False)
                    except Exception:
                        pass
                    workbook = None
    finally:
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()

    summary = summarize_report_records(output_records)
    return {
        "template_path": abs_template_path,
        "output_dir": abs_output_dir,
        "template_link_count": len(template_links),
        "records": output_records,
        **summary,
    }


def _new_output_record(record: ReportChecklistRecord) -> dict:
    return {
        "row_number": record.row_number,
        "tb_link_path": record.tb_link_path,
        "output_name_raw": record.output_name_raw,
        "company_name": record.company_name,
        "report_type": record.report_type,
        "output_path": "",
        "status": "",
        "message": "",
    }


def _build_success_message(template_link_count: int) -> str:
    if template_link_count > 1:
        return "已生成；模板存在多个外部链接，已全部替换为该行链接"
    return "已生成并替换链接"


def _is_xlsm(path: str) -> bool:
    return Path(str(path)).suffix.lower() == ".xlsm"


def _cell_to_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _log(logger, level: str, message: str) -> None:
    if logger is None:
        return
    log_func = getattr(logger, level, None)
    if log_func:
        log_func(message)

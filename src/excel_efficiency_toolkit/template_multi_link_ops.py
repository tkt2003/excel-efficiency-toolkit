import os
import shutil
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .link_replace_ops import XL_EXCEL_LINKS, normalize_com_link_sources, should_execute_rule
from .template_tb_report_ops import (
    DEFAULT_OUTPUT_NAME_SUFFIX,
    build_output_filename,
    build_unique_output_path,
    is_office_temp_file,
    is_supported_tb_file,
    is_supported_template_file,
    normalize_output_name_suffix,
)


INSTRUCTION_SHEET_NAME = "使用说明"
SETTINGS_SHEET_NAME = "参数设置"
RULE_SHEET_NAME = "生成清单"
LOG_SHEET_NAME = "处理日志"

RULE_BASE_HEADERS = ["是否执行", "输出文件名"]
RULE_TRAILING_HEADERS = ["状态", "说明", "输出文件路径", "主源文件路径"]
LOG_HEADERS = ["处理时间", "规则行号", "输出文件路径", "旧链接", "新链接", "状态", "说明"]
RULE_WORKBOOK_PREFIX = "按模板多链接生成Excel_规则表"
EDITABLE_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
REFERENCE_FILL = PatternFill(fill_type="solid", fgColor="F3F4F6")


@dataclass
class TemplateMultiLinkSettings:
    template_path: str
    output_dir: str
    output_name_suffix: str
    main_old_link: str
    generated_at: str = ""


@dataclass
class TemplateLinkReplacement:
    old_link_path: str
    new_link_path: str
    status: str = ""
    message: str = ""


@dataclass
class TemplateMultiLinkRule:
    row_number: int
    execute_raw: str
    output_name: str
    output_path: str
    main_source_path: str
    replacements: list[TemplateLinkReplacement] = field(default_factory=list)


@dataclass
class TemplateMultiLinkAction:
    row_number: int
    output_name: str
    output_path: str
    main_source_path: str
    replacements: list[TemplateLinkReplacement]
    status: str
    message: str


def create_template_multi_link_rule_workbook(
    template_path: str,
    output_dir: str,
    old_links: list[str],
    main_old_link: str,
    main_source_paths: list[str],
    output_name_suffix: str | None = None,
) -> str:
    template_text = str(template_path or "").strip()
    if not template_text:
        raise ValueError("请先选择模板工作簿。")
    if not is_supported_template_file(template_text):
        raise ValueError("模板工作簿仅支持 .xlsx / .xlsm。")

    output_dir_text = str(output_dir or "").strip()
    if not output_dir_text:
        raise ValueError("请先选择输出目录。")

    normalized_links = _normalize_old_links(old_links)
    main_link_text = str(main_old_link or "").strip()
    if not main_link_text:
        raise ValueError("请先选择主旧链接。")
    if not any(_same_link_path(link, main_link_text) for link in normalized_links):
        raise ValueError("主旧链接不在模板外部链接清单中。")

    valid_main_sources = _filter_valid_main_source_paths(main_source_paths)
    if not valid_main_sources:
        raise ValueError("请至少选择一个支持的主源文件（.xlsx / .xlsm）。")

    abs_template_path = os.path.abspath(template_text)
    abs_output_dir = os.path.abspath(output_dir_text)
    os.makedirs(abs_output_dir, exist_ok=True)
    name_suffix = normalize_output_name_suffix(output_name_suffix)
    template_suffix = Path(abs_template_path).suffix
    rule_path = _build_unique_rule_workbook_path(abs_output_dir)
    settings = TemplateMultiLinkSettings(
        template_path=abs_template_path,
        output_dir=abs_output_dir,
        output_name_suffix=name_suffix,
        main_old_link=main_link_text,
        generated_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    )

    workbook = Workbook()
    instruction_sheet = workbook.active
    instruction_sheet.title = INSTRUCTION_SHEET_NAME
    settings_sheet = workbook.create_sheet(SETTINGS_SHEET_NAME)
    rule_sheet = workbook.create_sheet(RULE_SHEET_NAME)
    log_sheet = workbook.create_sheet(LOG_SHEET_NAME)

    _write_instruction_sheet(instruction_sheet)
    _write_settings_sheet(settings_sheet, settings)
    _write_rule_sheet(
        rule_sheet,
        settings,
        normalized_links,
        valid_main_sources,
        template_suffix,
    )
    _write_log_sheet(log_sheet)

    workbook.active = workbook.sheetnames.index(RULE_SHEET_NAME)
    workbook.save(rule_path)
    workbook.close()
    return str(rule_path)


def read_template_multi_link_settings(rule_workbook_path: str) -> TemplateMultiLinkSettings:
    workbook = load_workbook(rule_workbook_path, data_only=True)
    try:
        sheet = _require_sheet(workbook, SETTINGS_SHEET_NAME)
        values: dict[str, str] = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            key = _cell_to_text(row[0] if len(row) > 0 else None)
            value = _cell_to_text(row[1] if len(row) > 1 else None)
            if key:
                values[key] = value

        settings = TemplateMultiLinkSettings(
            template_path=values.get("模板路径", ""),
            output_dir=values.get("输出目录", ""),
            output_name_suffix=values.get("输出后缀", ""),
            main_old_link=values.get("主旧链接", ""),
            generated_at=values.get("生成时间", ""),
        )
        _validate_settings(settings)
        return settings
    finally:
        workbook.close()


def read_template_multi_link_rules(rule_workbook_path: str) -> list[TemplateMultiLinkRule]:
    workbook = load_workbook(rule_workbook_path, data_only=True)
    try:
        sheet = _require_sheet(workbook, RULE_SHEET_NAME)
        header_map = _build_header_index_map(sheet)
        link_column_pairs = _read_link_column_pairs(sheet)
        rules: list[TemplateMultiLinkRule] = []
        for row_number in range(2, sheet.max_row + 1):
            execute_raw = _get_text_by_header(sheet, row_number, header_map, "是否执行")
            output_name = _get_text_by_header(sheet, row_number, header_map, "输出文件名")
            output_path = _get_text_by_header(sheet, row_number, header_map, "输出文件路径")
            main_source_path = _get_text_by_header(sheet, row_number, header_map, "主源文件路径")
            replacements: list[TemplateLinkReplacement] = []
            for old_col, new_col in link_column_pairs:
                old_link_path = _cell_to_text(sheet.cell(row=row_number, column=old_col).value)
                new_link_path = _cell_to_text(sheet.cell(row=row_number, column=new_col).value)
                if old_link_path and new_link_path:
                    replacements.append(TemplateLinkReplacement(old_link_path, new_link_path))
            if not any([execute_raw, output_name, output_path, main_source_path, replacements]):
                continue
            rules.append(
                TemplateMultiLinkRule(
                    row_number=row_number,
                    execute_raw=execute_raw,
                    output_name=output_name,
                    output_path=output_path,
                    main_source_path=main_source_path,
                    replacements=replacements,
                )
            )
        return rules
    finally:
        workbook.close()


def build_template_multi_link_actions(
    rules: list[TemplateMultiLinkRule],
    settings: TemplateMultiLinkSettings,
) -> list[TemplateMultiLinkAction]:
    actions: list[TemplateMultiLinkAction] = []
    for rule in rules:
        output_path = _resolve_output_path(rule, settings)
        output_name = os.path.basename(output_path) if output_path else rule.output_name
        if not should_execute_rule(rule.execute_raw):
            actions.append(_skip_action(rule, output_name, output_path, "是否执行为否，已跳过"))
            continue
        if not output_path:
            actions.append(_skip_action(rule, output_name, output_path, "输出文件路径为空"))
            continue
        if not rule.main_source_path:
            actions.append(_skip_action(rule, output_name, output_path, "主源文件路径为空"))
            continue
        if is_office_temp_file(rule.main_source_path):
            actions.append(_skip_action(rule, output_name, output_path, "临时文件已跳过"))
            continue
        if not is_supported_tb_file(rule.main_source_path):
            actions.append(_skip_action(rule, output_name, output_path, "主源文件仅支持 .xlsx / .xlsm"))
            continue
        replacements = _filter_supported_replacements(rule.replacements)
        if not replacements:
            actions.append(_skip_action(rule, output_name, output_path, "未填写任何新链接路径"))
            continue
        actions.append(
            TemplateMultiLinkAction(
                row_number=rule.row_number,
                output_name=output_name,
                output_path=output_path,
                main_source_path=os.path.abspath(rule.main_source_path),
                replacements=replacements,
                status="待执行",
                message="待生成",
            )
        )
    return actions


def summarize_template_multi_link_actions(actions: list[TemplateMultiLinkAction]) -> dict:
    return {
        "total_count": len(actions),
        "success_count": len([action for action in actions if action.status == "成功"]),
        "skipped_count": len([action for action in actions if action.status == "跳过"]),
        "failed_count": len([action for action in actions if action.status == "失败"]),
    }


def execute_template_multi_link_generation_from_rule_workbook(rule_workbook_path: str, logger=None) -> dict:
    settings = read_template_multi_link_settings(rule_workbook_path)
    rules = read_template_multi_link_rules(rule_workbook_path)
    actions = build_template_multi_link_actions(rules, settings)
    execute_template_multi_link_actions_with_com(actions, settings, logger=logger)
    write_template_multi_link_results_to_workbook(rule_workbook_path, actions)
    summary = summarize_template_multi_link_actions(actions)
    summary["rule_count"] = len(rules)
    return summary


def execute_template_multi_link_actions_with_com(
    actions: list[TemplateMultiLinkAction],
    settings: TemplateMultiLinkSettings,
    logger=None,
) -> None:
    executable_actions = [action for action in actions if action.status == "待执行"]
    if not executable_actions:
        return

    abs_template_path = os.path.abspath(settings.template_path)
    if not os.path.exists(abs_template_path):
        _mark_actions(executable_actions, "失败", f"模板工作簿不存在：{abs_template_path}")
        return

    os.makedirs(settings.output_dir, exist_ok=True)

    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    reserved_output_paths: set[str] = set()
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.EnableEvents = False
        excel.AskToUpdateLinks = False

        for action in executable_actions:
            workbook = None
            try:
                missing_link = _find_missing_new_link(action.replacements)
                if missing_link:
                    action.status = "失败"
                    action.message = f"新链接文件不存在：{missing_link}"
                    continue

                action.output_path = build_unique_output_path(
                    str(Path(action.output_path).parent),
                    os.path.basename(action.output_path),
                    reserved_output_paths,
                )
                reserved_output_paths.add(action.output_path)
                action.output_name = os.path.basename(action.output_path)

                _log(logger, "info", f"正在生成多链接 Excel：{action.output_path}")
                shutil.copy2(abs_template_path, action.output_path)
                workbook = excel.Workbooks.Open(
                    action.output_path,
                    ReadOnly=False,
                    UpdateLinks=0,
                )
                current_links = normalize_com_link_sources(workbook.LinkSources(XL_EXCEL_LINKS))
                success_count = 0
                failed_count = 0

                for replacement in action.replacements:
                    try:
                        if not _link_exists(replacement.old_link_path, current_links):
                            replacement.status = "失败"
                            replacement.message = "复制后的工作簿中未找到旧链接"
                            failed_count += 1
                            continue
                        workbook.ChangeLink(replacement.old_link_path, replacement.new_link_path, XL_EXCEL_LINKS)
                        replacement.status = "成功"
                        replacement.message = "已替换"
                        success_count += 1
                        current_links = [
                            replacement.new_link_path if _same_link_path(link, replacement.old_link_path) else link
                            for link in current_links
                        ]
                    except Exception as e:
                        replacement.status = "失败"
                        replacement.message = f"ChangeLink 失败：{e}"
                        failed_count += 1

                if success_count > 0:
                    workbook.Save()
                if failed_count:
                    action.status = "失败"
                    action.message = f"已生成，成功替换 {success_count} 个链接，失败 {failed_count} 个链接"
                else:
                    action.status = "成功"
                    action.message = f"已生成并替换 {success_count} 个链接"
            except Exception as e:
                action.status = "失败"
                action.message = f"生成失败：{e}"
                _mark_pending_replacements_failed(action, action.message)
                _log(logger, "error", f"多链接生成失败：第 {action.row_number} 行。详细信息：{e}")
            finally:
                if workbook is not None:
                    try:
                        workbook.Close(SaveChanges=False)
                    except Exception:
                        pass
                    workbook = None
    finally:
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()


def write_template_multi_link_results_to_workbook(
    rule_workbook_path: str,
    actions: list[TemplateMultiLinkAction],
) -> None:
    workbook = load_workbook(rule_workbook_path)
    try:
        rule_sheet = _require_sheet(workbook, RULE_SHEET_NAME)
        log_sheet = _require_sheet(workbook, LOG_SHEET_NAME)
        header_map = _build_header_index_map(rule_sheet)
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for action in actions:
            _write_value_by_header(rule_sheet, action.row_number, header_map, "输出文件名", action.output_name)
            _write_value_by_header(rule_sheet, action.row_number, header_map, "输出文件路径", action.output_path)
            _write_value_by_header(rule_sheet, action.row_number, header_map, "状态", action.status)
            _write_value_by_header(rule_sheet, action.row_number, header_map, "说明", action.message)
            if action.replacements:
                for replacement in action.replacements:
                    log_sheet.append(
                        [
                            now,
                            action.row_number,
                            action.output_path,
                            replacement.old_link_path,
                            replacement.new_link_path,
                            replacement.status or action.status,
                            replacement.message or action.message,
                        ]
                    )
            else:
                log_sheet.append([now, action.row_number, action.output_path, "", "", action.status, action.message])

        workbook.save(rule_workbook_path)
    finally:
        workbook.close()


def _normalize_old_links(old_links: list[str]) -> list[str]:
    normalized: list[str] = []
    seen: set[str] = set()
    for raw_link in old_links or []:
        link = str(raw_link or "").strip()
        key = os.path.normcase(link)
        if link and key not in seen:
            normalized.append(link)
            seen.add(key)
    if not normalized:
        raise ValueError("模板没有外部链接，无法生成多链接规则表。")
    return normalized


def _filter_valid_main_source_paths(main_source_paths: list[str]) -> list[str]:
    valid_paths: list[str] = []
    for raw_path in main_source_paths or []:
        path = str(raw_path or "").strip()
        if not path or is_office_temp_file(path) or not is_supported_tb_file(path):
            continue
        valid_paths.append(os.path.abspath(path))
    return valid_paths


def _filter_supported_replacements(replacements: list[TemplateLinkReplacement]) -> list[TemplateLinkReplacement]:
    supported: list[TemplateLinkReplacement] = []
    for replacement in replacements:
        old_link = str(replacement.old_link_path or "").strip()
        new_link = str(replacement.new_link_path or "").strip()
        if not old_link or not new_link:
            continue
        if is_office_temp_file(new_link) or not is_supported_tb_file(new_link):
            continue
        supported.append(TemplateLinkReplacement(old_link, os.path.abspath(new_link)))
    return supported


def _write_instruction_sheet(sheet) -> None:
    rows = [
        ["按模板多链接生成 Excel 使用说明"],
        ["1. 本功能适用于模板中存在多个外部链接，并且每个输出文件可能替换不同链接的场景。"],
        ["2. 主链接决定生成清单行数；【生成清单】每一行代表一份输出 Excel。"],
        ["3. 主链接的新链接列已自动填写主源文件路径；其他新链接列默认留空。"],
        ["4. 新链接为空表示不替换；多个旧链接可在同一行分别填写各自的新链接。"],
        ["5. 执行后会回写【生成清单】中的状态和说明，并在【处理日志】记录明细。"],
        ["6. 保存并关闭规则表后，回到工具台点击“我已检查规则，开始生成”。"],
        ["7. 原模板不会被修改；输出文件已存在时自动编号，不覆盖。"],
    ]
    _append_rows(sheet, rows)
    sheet["A1"].font = Font(bold=True)
    sheet.column_dimensions["A"].width = 108


def _write_settings_sheet(sheet, settings: TemplateMultiLinkSettings) -> None:
    rows = [
        ["参数项", "参数值", "说明"],
        ["模板路径", settings.template_path, "只读复制，不修改原模板。"],
        ["输出目录", settings.output_dir, "生成的 Excel 文件保存目录。"],
        ["输出后缀", settings.output_name_suffix, "可留空；默认 _批量生成。"],
        ["主旧链接", settings.main_old_link, "主链接决定生成清单行数。"],
        ["生成时间", settings.generated_at, "规则表生成时间。"],
    ]
    _append_rows(sheet, rows)
    _style_header(sheet, 1, len(rows[0]))
    _set_widths(sheet, {"A": 18, "B": 72, "C": 58})


def _write_rule_sheet(
    sheet,
    settings: TemplateMultiLinkSettings,
    old_links: list[str],
    main_source_paths: list[str],
    template_suffix: str,
) -> None:
    headers = RULE_BASE_HEADERS[:]
    for index in range(1, len(old_links) + 1):
        headers.extend([f"旧链接 {index}", f"新链接 {index}"])
    headers.extend(RULE_TRAILING_HEADERS)
    sheet.append(headers)
    _style_header(sheet, 1, len(headers))

    trailing_start = len(headers) - len(RULE_TRAILING_HEADERS) + 1
    used_output_paths: set[str] = set()
    for main_source_path in main_source_paths:
        output_name = build_output_filename(main_source_path, template_suffix, settings.output_name_suffix)
        output_path = build_unique_output_path(settings.output_dir, output_name, used_output_paths)
        used_output_paths.add(output_path)
        row = ["是", os.path.basename(output_path)]
        for old_link in old_links:
            new_link = main_source_path if _same_link_path(old_link, settings.main_old_link) else ""
            row.extend([old_link, new_link])
        row.extend(["", "", output_path, main_source_path])
        sheet.append(row)

    for row_number in range(2, sheet.max_row + 1):
        sheet.cell(row=row_number, column=2).fill = REFERENCE_FILL
        for index in range(len(old_links)):
            old_col = 3 + index * 2
            new_col = old_col + 1
            sheet.cell(row=row_number, column=old_col).fill = REFERENCE_FILL
            sheet.cell(row=row_number, column=new_col).fill = EDITABLE_FILL
        for column in range(trailing_start, len(headers) + 1):
            if headers[column - 1] in {"输出文件路径", "主源文件路径"}:
                sheet.cell(row=row_number, column=column).fill = REFERENCE_FILL

    _auto_fit_columns(sheet)
    sheet.freeze_panes = "C2"


def _write_log_sheet(sheet) -> None:
    sheet.append(LOG_HEADERS)
    _style_header(sheet, 1, len(LOG_HEADERS))
    _set_widths(sheet, {"A": 20, "B": 10, "C": 64, "D": 58, "E": 58, "F": 12, "G": 42})
    sheet.freeze_panes = "A2"


def _read_link_column_pairs(sheet) -> list[tuple[int, int]]:
    pairs: list[tuple[int, int]] = []
    for column in range(1, sheet.max_column + 1):
        old_header = _cell_to_text(sheet.cell(row=1, column=column).value)
        if not old_header.startswith("旧链接"):
            continue
        new_header = _cell_to_text(sheet.cell(row=1, column=column + 1).value)
        if new_header.startswith("新链接"):
            pairs.append((column, column + 1))
    return pairs


def _build_header_index_map(sheet) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for column in range(1, sheet.max_column + 1):
        header = _cell_to_text(sheet.cell(row=1, column=column).value)
        if header:
            header_map[header] = column
    return header_map


def _get_text_by_header(sheet, row_number: int, header_map: dict[str, int], header: str) -> str:
    column = header_map.get(header)
    if not column:
        return ""
    return _cell_to_text(sheet.cell(row=row_number, column=column).value)


def _write_value_by_header(sheet, row_number: int, header_map: dict[str, int], header: str, value: object) -> None:
    column = header_map.get(header)
    if not column:
        raise ValueError(f"规则表缺少“{header}”列。")
    sheet.cell(row=row_number, column=column, value=value)


def _resolve_output_path(rule: TemplateMultiLinkRule, settings: TemplateMultiLinkSettings) -> str:
    if rule.output_path:
        return os.path.abspath(rule.output_path)
    if rule.output_name:
        return os.path.abspath(os.path.join(settings.output_dir, rule.output_name))
    return ""


def _skip_action(
    rule: TemplateMultiLinkRule,
    output_name: str,
    output_path: str,
    message: str,
) -> TemplateMultiLinkAction:
    return TemplateMultiLinkAction(
        row_number=rule.row_number,
        output_name=output_name,
        output_path=output_path,
        main_source_path=rule.main_source_path,
        replacements=[],
        status="跳过",
        message=message,
    )


def _validate_settings(settings: TemplateMultiLinkSettings) -> None:
    if not settings.template_path:
        raise ValueError("规则表参数缺少模板路径。")
    if not settings.output_dir:
        raise ValueError("规则表参数缺少输出目录。")
    if not settings.main_old_link:
        raise ValueError("规则表参数缺少主旧链接。")


def _require_sheet(workbook, sheet_name: str):
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"规则表缺少“{sheet_name}”sheet。")
    return workbook[sheet_name]


def _find_missing_new_link(replacements: list[TemplateLinkReplacement]) -> str:
    for replacement in replacements:
        if not os.path.exists(replacement.new_link_path):
            return replacement.new_link_path
    return ""


def _mark_actions(actions: list[TemplateMultiLinkAction], status: str, message: str) -> None:
    for action in actions:
        action.status = status
        action.message = message
        _mark_pending_replacements_failed(action, message)


def _mark_pending_replacements_failed(action: TemplateMultiLinkAction, message: str) -> None:
    for replacement in action.replacements:
        if not replacement.status:
            replacement.status = "失败"
            replacement.message = message


def _link_exists(target_link: str, current_links: list[str]) -> bool:
    return any(_same_link_path(link, target_link) for link in current_links)


def _same_link_path(left: str, right: str) -> bool:
    return os.path.normcase(str(left).strip()) == os.path.normcase(str(right).strip())


def _build_unique_rule_workbook_path(output_dir: str) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = Path(output_dir) / f"{RULE_WORKBOOK_PREFIX}_{timestamp}.xlsx"
    if not base.exists():
        return base
    counter = 2
    while True:
        candidate = base.with_name(f"{base.stem}_{counter}{base.suffix}")
        if not candidate.exists():
            return candidate
        counter += 1


def _append_rows(sheet, rows: list[list[object]]) -> None:
    for row in rows:
        sheet.append(row)


def _style_header(sheet, row_number: int, column_count: int) -> None:
    for column in range(1, column_count + 1):
        sheet.cell(row=row_number, column=column).font = Font(bold=True)


def _set_widths(sheet, widths: dict[str, int]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _auto_fit_columns(sheet) -> None:
    for column_number in range(1, sheet.max_column + 1):
        max_length = 0
        for row_number in range(1, sheet.max_row + 1):
            text = _cell_to_text(sheet.cell(row=row_number, column=column_number).value)
            if len(text) > max_length:
                max_length = len(text)
        width = min(max(max_length + 2, 10), 64)
        sheet.column_dimensions[get_column_letter(column_number)].width = width


def _cell_to_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _log(logger, level: str, message: str) -> None:
    if logger is None:
        return
    log_method = getattr(logger, level, None)
    if log_method is None:
        return
    log_method(message)

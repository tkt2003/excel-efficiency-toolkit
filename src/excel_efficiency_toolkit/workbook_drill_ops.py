from __future__ import annotations

from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter, range_boundaries


RESULT_SHEET_BASE_NAME = "数据穿透查询结果"
EXCEL_ERROR_TEXTS = {
    "#VALUE!",
    "#DIV/0!",
    "#REF!",
    "#NAME?",
    "#N/A",
    "#NULL!",
    "#NUM!",
}
EXCEL_ERROR_CODE_MAP = {
    2000: "#NULL!",
    2007: "#DIV/0!",
    2015: "#VALUE!",
    2023: "#REF!",
    2029: "#NAME?",
    2036: "#NUM!",
    2042: "#N/A",
}


def normalize_range_address(range_address: str) -> str:
    return str(range_address or "").replace("$", "").strip().upper()


def is_multi_area_range(range_address: str) -> bool:
    normalized = normalize_range_address(range_address)
    return "," in normalized


def expand_range_addresses(range_address: str) -> list[str]:
    normalized = normalize_range_address(range_address)
    if not normalized:
        raise ValueError("选区地址为空")
    if is_multi_area_range(normalized):
        raise ValueError("不支持多区域选区")

    min_col, min_row, max_col, max_row = range_boundaries(normalized)
    addresses = []
    for row_index in range(min_row, max_row + 1):
        for column_index in range(min_col, max_col + 1):
            addresses.append(f"{get_column_letter(column_index)}{row_index}")
    return addresses


def build_unique_result_sheet_name(
    existing_sheet_names: list[str],
    base_name: str = RESULT_SHEET_BASE_NAME,
) -> str:
    existing_names = {str(name or "").strip() for name in existing_sheet_names}
    if base_name not in existing_names:
        return base_name

    counter = 2
    while True:
        candidate = f"{base_name}_{counter}"
        if candidate not in existing_names:
            return candidate
        counter += 1


def should_skip_history_result_sheet(
    sheet_name: str,
    base_name: str = RESULT_SHEET_BASE_NAME,
) -> bool:
    return str(sheet_name or "").strip().startswith(base_name)


def stringify_excel_error_value(value):
    if isinstance(value, str):
        text = value.strip()
        if text in EXCEL_ERROR_TEXTS:
            return text
        return value

    if isinstance(value, (int, float)):
        error_text = EXCEL_ERROR_CODE_MAP.get(int(value))
        if error_text is not None:
            return error_text
    return value


def build_single_file_result_headers(range_address: str) -> list[str]:
    return ["工作表名", "工作表可见状态", *expand_range_addresses(range_address)]


def build_multi_file_result_headers(range_address: str) -> list[str]:
    return ["源文件名", "源文件路径", "工作表名", *expand_range_addresses(range_address), "状态", "说明"]


def build_single_file_result_row(
    sheet_name: str,
    visible_status: str,
    values_by_address: dict[str, object],
    range_address: str,
) -> list[object]:
    addresses = expand_range_addresses(range_address)
    return [
        sheet_name,
        visible_status,
        *[stringify_excel_error_value(values_by_address.get(address)) for address in addresses],
    ]


def build_multi_file_result_row(record: dict, range_address: str) -> list[object]:
    addresses = expand_range_addresses(range_address)
    values_by_address = record.get("values_by_address", {})
    return [
        record.get("source_file_name", ""),
        record.get("source_file_path", ""),
        record.get("target_sheet_name", ""),
        *[stringify_excel_error_value(values_by_address.get(address)) for address in addresses],
        record.get("status", ""),
        record.get("message", ""),
    ]


def write_single_file_result_sheet(
    sheet,
    base_sheet_name: str,
    range_address: str,
    created_at_text: str,
    records: list[dict],
) -> None:
    sheet.cell(row=1, column=1).value = (
        f"基准工作表：{base_sheet_name}    基准选区：{normalize_range_address(range_address)}    生成时间：{created_at_text}"
    )

    headers = build_single_file_result_headers(range_address)
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(row=3, column=column_index).value = header

    for row_offset, record in enumerate(records, start=4):
        row_values = build_single_file_result_row(
            sheet_name=record.get("sheet_name", ""),
            visible_status=record.get("visible_status", ""),
            values_by_address=record.get("values_by_address", {}),
            range_address=range_address,
        )
        for column_index, value in enumerate(row_values, start=1):
            sheet.cell(row=row_offset, column=column_index).value = value


def write_single_workbook_drill_result_to_com_sheet(
    sheet,
    base_sheet_name: str,
    range_address: str,
    created_at_text: str,
    records: list[dict],
) -> None:
    sheet.Cells(1, 1).Value = (
        f"基准工作表：{base_sheet_name}    基准选区：{normalize_range_address(range_address)}    生成时间：{created_at_text}"
    )

    headers = build_single_file_result_headers(range_address)
    for column_index, header in enumerate(headers, start=1):
        sheet.Cells(3, column_index).Value = header

    for row_offset, record in enumerate(records, start=4):
        row_values = build_single_file_result_row(
            sheet_name=record.get("sheet_name", ""),
            visible_status=record.get("visible_status", ""),
            values_by_address=record.get("values_by_address", {}),
            range_address=range_address,
        )
        for column_index, value in enumerate(row_values, start=1):
            sheet.Cells(row_offset, column_index).Value = value


def write_multi_file_result_sheet(
    sheet,
    records: list[dict],
    range_address: str,
) -> None:
    headers = build_multi_file_result_headers(range_address)
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=column_index).value = header

    for row_offset, record in enumerate(records, start=2):
        row_values = build_multi_file_result_row(record, range_address)
        for column_index, value in enumerate(row_values, start=1):
            sheet.cell(row=row_offset, column=column_index).value = value


def range_value_to_address_map(range_address: str, raw_value) -> dict[str, object]:
    addresses = expand_range_addresses(range_address)
    if len(addresses) == 1:
        return {addresses[0]: stringify_excel_error_value(raw_value)}

    rows = _normalize_range_values_matrix(raw_value, range_address)
    values_by_address = {}
    address_index = 0
    for row_values in rows:
        for value in row_values:
            values_by_address[addresses[address_index]] = stringify_excel_error_value(value)
            address_index += 1
    return values_by_address


def is_excel_error_text(value) -> bool:
    return isinstance(value, str) and value.strip() in EXCEL_ERROR_TEXTS


def _normalize_range_values_matrix(raw_value, range_address: str) -> list[list[object]]:
    normalized = normalize_range_address(range_address)
    min_col, min_row, max_col, max_row = range_boundaries(normalized)
    width = max_col - min_col + 1
    height = max_row - min_row + 1

    if raw_value is None:
        return [[None for _ in range(width)] for _ in range(height)]

    if isinstance(raw_value, tuple):
        rows = []
        for row_values in raw_value:
            if isinstance(row_values, tuple):
                rows.append(list(row_values))
            else:
                rows.append([row_values])
        return _pad_matrix(rows, width, height)

    return _pad_matrix([[raw_value]], width, height)


def _pad_matrix(rows: list[list[object]], width: int, height: int) -> list[list[object]]:
    normalized_rows = []
    for row_index in range(height):
        source_row = rows[row_index] if row_index < len(rows) else []
        normalized_row = list(source_row[:width])
        if len(normalized_row) < width:
            normalized_row.extend([None] * (width - len(normalized_row)))
        normalized_rows.append(normalized_row)
    return normalized_rows

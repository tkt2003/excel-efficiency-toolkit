import os
import re
import tempfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


INSTRUCTION_SHEET_NAME = "使用说明"
SETTINGS_SHEET_NAME = "参数设置"
RULE_SHEET_NAME = "重命名清单"
LOG_SHEET_NAME = "处理日志"

EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
RULE_HEADERS = ["原文件路径", "原文件名", "原工作表名", "新工作表名", "状态", "说明", "是否隐藏", "原顺序"]
LOG_HEADERS = ["处理时间", "文件路径", "原文件名", "原工作表名", "目标工作表名", "状态", "说明"]
ILLEGAL_SHEET_NAME_PATTERN = re.compile(r"[:\\/\?\*\[\]]")
MAX_SHEET_NAME_LENGTH = 31


@dataclass
class SheetRenameRule:
    row_number: int
    workbook_path: str
    workbook_name: str
    original_sheet_name: str
    new_sheet_name_raw: str
    is_hidden: bool
    original_order: int | None = None
    existing_status: str = ""
    existing_message: str = ""


@dataclass
class SheetRenameSettings:
    conflict_mode: str = "自动编号"
    rename_hidden_sheets: bool = False
    clean_illegal_chars: bool = True
    too_long_mode: str = "自动截断"
    skip_temp_files: bool = True
    warnings: list[str] | None = None


@dataclass
class SheetRenameAction:
    row_number: int
    workbook_path: str
    original_sheet_name: str
    target_sheet_name: str
    status: str
    message: str


def is_excel_workbook_file(path: str) -> bool:
    filename = os.path.basename(path)
    if filename.startswith("~$"):
        return False
    return Path(filename).suffix.lower() in EXCEL_EXTENSIONS


def is_office_temp_file(path: str) -> bool:
    return os.path.basename(path).startswith("~$")


def create_sheet_rename_rule_workbook(
    workbook_infos: list[dict],
    output_dir: str | None = None,
) -> str:
    if not workbook_infos:
        raise ValueError("请选择至少一个需要重命名工作表的 Excel 文件。")

    target_dir = Path(output_dir).resolve() if output_dir else _infer_output_dir(workbook_infos)
    target_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = _get_unique_path(target_dir / f"工作表重命名规则_{timestamp}.xlsx")

    workbook = Workbook()
    instruction_sheet = workbook.active
    instruction_sheet.title = INSTRUCTION_SHEET_NAME
    settings_sheet = workbook.create_sheet(SETTINGS_SHEET_NAME)
    rule_sheet = workbook.create_sheet(RULE_SHEET_NAME)
    log_sheet = workbook.create_sheet(LOG_SHEET_NAME)

    _write_instruction_sheet(instruction_sheet)
    _write_settings_sheet(settings_sheet)
    _write_rule_sheet(rule_sheet, workbook_infos)
    _write_log_sheet(log_sheet)

    workbook.active = workbook.sheetnames.index(RULE_SHEET_NAME)
    workbook.save(output_path)
    workbook.close()
    return str(output_path)


def read_sheet_rename_settings(rule_workbook_path: str) -> SheetRenameSettings:
    workbook = load_workbook(rule_workbook_path, data_only=True)
    warnings: list[str] = []
    try:
        sheet = workbook[SETTINGS_SHEET_NAME]
        values = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            key = _cell_to_text(row[0] if len(row) > 0 else None)
            value = _cell_to_text(row[1] if len(row) > 1 else None)
            if key:
                values[key] = value

        conflict_mode = _read_choice_setting(
            values.get("目标工作表名冲突处理方式", ""),
            choices={"自动编号", "跳过"},
            default="自动编号",
            setting_name="目标工作表名冲突处理方式",
            warnings=warnings,
        )
        rename_hidden_sheets = _read_yes_no_setting(
            values.get("是否重命名隐藏工作表", ""),
            default=False,
            setting_name="是否重命名隐藏工作表",
            warnings=warnings,
        )
        clean_illegal_chars = _read_yes_no_setting(
            values.get("是否清洗非法字符", ""),
            default=True,
            setting_name="是否清洗非法字符",
            warnings=warnings,
        )
        too_long_mode = _read_choice_setting(
            values.get("名称超过 31 字符时", ""),
            choices={"自动截断", "跳过"},
            default="自动截断",
            setting_name="名称超过 31 字符时",
            warnings=warnings,
        )
        skip_temp_files = _read_yes_no_setting(
            values.get("是否跳过临时文件", ""),
            default=True,
            setting_name="是否跳过临时文件",
            warnings=warnings,
        )

        return SheetRenameSettings(
            conflict_mode=conflict_mode,
            rename_hidden_sheets=rename_hidden_sheets,
            clean_illegal_chars=clean_illegal_chars,
            too_long_mode=too_long_mode,
            skip_temp_files=skip_temp_files,
            warnings=warnings,
        )
    finally:
        workbook.close()


def read_sheet_rename_rules(rule_workbook_path: str) -> list[SheetRenameRule]:
    workbook = load_workbook(rule_workbook_path, data_only=True)
    try:
        sheet = workbook[RULE_SHEET_NAME]
        rules: list[SheetRenameRule] = []
        for row_number in range(2, sheet.max_row + 1):
            workbook_path = _cell_to_text(sheet.cell(row=row_number, column=1).value)
            workbook_name = _cell_to_text(sheet.cell(row=row_number, column=2).value)
            original_sheet_name = _cell_to_text(sheet.cell(row=row_number, column=3).value)
            new_sheet_name_raw = _cell_to_text(sheet.cell(row=row_number, column=4).value)
            existing_status = _cell_to_text(sheet.cell(row=row_number, column=5).value)
            existing_message = _cell_to_text(sheet.cell(row=row_number, column=6).value)
            is_hidden = _cell_to_text(sheet.cell(row=row_number, column=7).value) == "是"
            original_order = _cell_to_int(sheet.cell(row=row_number, column=8).value)
            rules.append(
                SheetRenameRule(
                    row_number=row_number,
                    workbook_path=workbook_path,
                    workbook_name=workbook_name,
                    original_sheet_name=original_sheet_name,
                    new_sheet_name_raw=new_sheet_name_raw,
                    is_hidden=is_hidden,
                    original_order=original_order,
                    existing_status=existing_status,
                    existing_message=existing_message,
                )
            )
        return rules
    finally:
        workbook.close()


def group_sheet_rename_rules_by_workbook_path(rules: list[SheetRenameRule]) -> dict[str, list[SheetRenameRule]]:
    grouped_rules: dict[str, list[SheetRenameRule]] = {}
    for rule in rules:
        key = _normalize_path_for_compare(rule.workbook_path)
        grouped_rules.setdefault(key, []).append(rule)
    return grouped_rules


def build_sheet_rename_plan(
    rules: list[SheetRenameRule],
    existing_sheet_names: list[str],
    settings: SheetRenameSettings,
) -> list[SheetRenameAction]:
    existing_names = {_normalize_sheet_name(name): name for name in existing_sheet_names}
    base_actions = [_build_base_sheet_action(rule, existing_names, settings) for rule in rules]
    original_names_to_be_renamed = _success_original_sheet_names(base_actions)

    actions = base_actions
    for _ in range(len(base_actions) + 1):
        actions = _resolve_sheet_target_conflicts(
            base_actions,
            existing_names,
            original_names_to_be_renamed,
            settings,
        )
        next_original_names = _success_original_sheet_names(actions)
        if next_original_names == original_names_to_be_renamed:
            return actions
        original_names_to_be_renamed = next_original_names

    return actions


def build_skipped_sheet_rename_actions(
    rules: list[SheetRenameRule],
    message: str,
) -> list[SheetRenameAction]:
    return [
        SheetRenameAction(
            row_number=rule.row_number,
            workbook_path=rule.workbook_path,
            original_sheet_name=rule.original_sheet_name,
            target_sheet_name="",
            status="跳过",
            message=message,
        )
        for rule in rules
    ]


def execute_sheet_rename_plan(workbook, actions: list[SheetRenameAction]) -> dict:
    success_count = 0
    skipped_count = len([action for action in actions if action.status == "跳过"])
    failed_count = 0
    executable_actions = [action for action in actions if action.status == "成功"]
    final_target_names = {_normalize_sheet_name(action.target_sheet_name) for action in executable_actions}
    temp_records = []
    reserved_temp_names: set[str] = set()

    for index, action in enumerate(executable_actions, start=1):
        try:
            sheet = _find_worksheet_by_name(workbook, action.original_sheet_name)
            if sheet is None:
                action.status = "跳过"
                action.message = _join_messages(action.message, "原工作表不存在")
                skipped_count += 1
                continue

            temp_name = _build_temp_sheet_name(workbook, final_target_names, reserved_temp_names, index)
            sheet.Name = temp_name
            reserved_temp_names.add(_normalize_sheet_name(temp_name))
            temp_records.append((action, sheet, temp_name))
        except Exception as e:
            action.status = "失败"
            action.message = _join_messages(action.message, f"第一阶段临时重命名失败：{e}")
            failed_count += 1

    for action, sheet, temp_name in temp_records:
        try:
            existing_target_sheet = _find_worksheet_by_name(workbook, action.target_sheet_name)
            if existing_target_sheet is not None and existing_target_sheet is not sheet:
                raise ValueError("目标工作表名仍被占用")

            sheet.Name = action.target_sheet_name
            action.message = "已重命名" if action.message == "待重命名" else _join_messages(action.message, "已重命名")
            success_count += 1
        except Exception as e:
            rollback_message = _rollback_temp_sheet(workbook, sheet, temp_name, action.original_sheet_name)
            action.status = "失败"
            action.message = _join_messages(action.message, f"第二阶段重命名失败：{e}", rollback_message)
            failed_count += 1

    return {
        "success_count": success_count,
        "skipped_count": skipped_count,
        "failed_count": failed_count,
    }


def summarize_sheet_rename_actions(actions: list[SheetRenameAction]) -> dict:
    return {
        "success_count": len([action for action in actions if action.status == "成功"]),
        "skipped_count": len([action for action in actions if action.status == "跳过"]),
        "failed_count": len([action for action in actions if action.status == "失败"]),
    }


def write_sheet_rename_results_to_workbook(rule_workbook_path: str, actions: list[SheetRenameAction]) -> None:
    workbook = load_workbook(rule_workbook_path)
    try:
        rule_sheet = workbook[RULE_SHEET_NAME]
        log_sheet = workbook[LOG_SHEET_NAME]
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for action in actions:
            rule_sheet.cell(row=action.row_number, column=5, value=action.status)
            rule_sheet.cell(row=action.row_number, column=6, value=action.message)
            log_sheet.append(
                [
                    now,
                    action.workbook_path,
                    Path(action.workbook_path).name if action.workbook_path else "",
                    action.original_sheet_name,
                    action.target_sheet_name,
                    action.status,
                    action.message,
                ]
            )

        workbook.save(rule_workbook_path)
    finally:
        workbook.close()


def is_sheet_hidden_by_visible_value(visible_value) -> bool:
    return visible_value not in (-1, True, "visible")


def _build_base_sheet_action(
    rule: SheetRenameRule,
    existing_names: dict[str, str],
    settings: SheetRenameSettings,
) -> SheetRenameAction:
    if rule.existing_status == "跳过" and rule.existing_message:
        return _skip_action(rule, "", rule.existing_message)

    workbook_path = rule.workbook_path.strip()
    if not workbook_path:
        return _skip_action(rule, "", "原文件路径为空")
    if settings.skip_temp_files and is_office_temp_file(workbook_path):
        return _skip_action(rule, "", "临时文件已跳过")
    if not Path(workbook_path).exists():
        return _skip_action(rule, "", "原文件不存在")

    original_name = rule.original_sheet_name.strip()
    if not original_name:
        return _skip_action(rule, "", "原工作表名为空")

    target_name = rule.new_sheet_name_raw.strip()
    if not target_name:
        return _skip_action(rule, "", "未填写新工作表名")

    original_key = _normalize_sheet_name(original_name)
    if original_key not in existing_names:
        return _skip_action(rule, "", "原工作表不存在")

    if rule.is_hidden and not settings.rename_hidden_sheets:
        return _skip_action(rule, "", "隐藏工作表已跳过")

    messages: list[str] = []
    if ILLEGAL_SHEET_NAME_PATTERN.search(target_name):
        if settings.clean_illegal_chars:
            target_name = ILLEGAL_SHEET_NAME_PATTERN.sub("_", target_name)
            messages.append("已清洗非法字符")
        else:
            return _skip_action(rule, "", "工作表名包含非法字符")

    if not target_name.strip():
        return _skip_action(rule, "", "新工作表名为空")

    if len(target_name) > MAX_SHEET_NAME_LENGTH:
        if settings.too_long_mode == "跳过":
            return _skip_action(rule, "", "工作表名超过 31 字符")
        target_name = target_name[:MAX_SHEET_NAME_LENGTH]
        messages.append("工作表名超过 31 字符，已自动截断")

    target_key = _normalize_sheet_name(target_name)
    if target_key == original_key:
        return _skip_action(rule, target_name, "新旧工作表名相同")

    return SheetRenameAction(
        row_number=rule.row_number,
        workbook_path=workbook_path,
        original_sheet_name=original_name,
        target_sheet_name=target_name,
        status="成功",
        message=_join_messages(*messages) or "待重命名",
    )


def _resolve_sheet_target_conflicts(
    base_actions: list[SheetRenameAction],
    existing_names: dict[str, str],
    original_names_to_be_renamed: set[str],
    settings: SheetRenameSettings,
) -> list[SheetRenameAction]:
    actions: list[SheetRenameAction] = []
    assigned_targets: set[str] = set()

    for base_action in base_actions:
        action = _copy_action(base_action)
        if action.status != "成功":
            actions.append(action)
            continue

        original_key = _normalize_sheet_name(action.original_sheet_name)
        target_key = _normalize_sheet_name(action.target_sheet_name)
        conflicts_with_workbook = (
            target_key in existing_names
            and target_key != original_key
            and target_key not in original_names_to_be_renamed
        )
        conflicts_with_plan = target_key in assigned_targets
        if conflicts_with_workbook or conflicts_with_plan:
            if settings.conflict_mode == "跳过":
                action.status = "跳过"
                action.message = "目标工作表名冲突"
                actions.append(action)
                continue

            unavailable_names = assigned_targets | (set(existing_names.keys()) - original_names_to_be_renamed)
            action.target_sheet_name = _build_numbered_sheet_name(action.target_sheet_name, unavailable_names)
            action.message = _join_messages(
                action.message if action.message != "待重命名" else "",
                "目标工作表名冲突，已自动编号",
            ) or "待重命名"
            target_key = _normalize_sheet_name(action.target_sheet_name)

        assigned_targets.add(target_key)
        actions.append(action)

    return actions


def _success_original_sheet_names(actions: list[SheetRenameAction]) -> set[str]:
    return {
        _normalize_sheet_name(action.original_sheet_name)
        for action in actions
        if action.status == "成功"
    }


def _rollback_temp_sheet(workbook, sheet, temp_name: str, original_sheet_name: str) -> str:
    if sheet.Name != temp_name:
        return ""
    if _find_worksheet_by_name(workbook, original_sheet_name) is not None:
        return f"临时工作表名保留为：{temp_name}"
    try:
        sheet.Name = original_sheet_name
        return "已回滚到原工作表名"
    except Exception as rollback_error:
        return f"回滚失败，临时工作表名保留为：{temp_name}；回滚错误：{rollback_error}"


def _build_temp_sheet_name(
    workbook,
    final_target_names: set[str],
    reserved_temp_names: set[str],
    start_index: int,
) -> str:
    counter = start_index
    while True:
        candidate = f"__tmp_rename_{counter:03d}"
        candidate_key = _normalize_sheet_name(candidate)
        if (
            len(candidate) <= MAX_SHEET_NAME_LENGTH
            and candidate_key not in final_target_names
            and candidate_key not in reserved_temp_names
            and _find_worksheet_by_name(workbook, candidate) is None
        ):
            return candidate
        counter += 1


def _find_worksheet_by_name(workbook, sheet_name: str):
    for sheet in workbook.Worksheets:
        if sheet.Name == sheet_name:
            return sheet
    return None


def _build_numbered_sheet_name(base_name: str, unavailable_names: set[str]) -> str:
    counter = 1
    while True:
        suffix = f"_{counter}"
        trimmed_base = base_name[: MAX_SHEET_NAME_LENGTH - len(suffix)]
        candidate = f"{trimmed_base}{suffix}"
        if _normalize_sheet_name(candidate) not in unavailable_names:
            return candidate
        counter += 1


def _write_instruction_sheet(sheet) -> None:
    rows = [
        ["批量重命名工作表使用说明"],
        ["1. 本工具用于批量重命名用户选择的 Excel 工作簿中的工作表。"],
        ["2. 在【重命名清单】D列填写新工作表名。"],
        ["3. A/C列为原文件路径和原工作表名，请不要修改。"],
        ["4. 保存并关闭/释放规则表后，回到工具台点击“执行批量重命名工作表”。"],
        ["5. 工作表名不能超过 31 个字符。"],
        ["6. 工作表名不能包含 : \\ / ? * [ ]。"],
        ["7. 工作表名不能为空。"],
        ["8. 同一目标工作簿内工作表名不能重复，不同工作簿之间互不冲突。"],
        ["9. 默认遇到重名时自动编号，默认跳过隐藏工作表。"],
        ["10. 执行后会在 E/F 列写入状态和说明，并在【处理日志】记录明细。"],
        ["11. 执行后会保存被处理的 Excel 文件，建议操作前先备份。"],
    ]
    for row in rows:
        sheet.append(row)
    sheet["A1"].font = Font(bold=True)
    sheet.column_dimensions["A"].width = 100


def _write_settings_sheet(sheet) -> None:
    rows = [
        ["参数项", "参数值", "说明"],
        ["目标工作表名冲突处理方式", "自动编号", "自动编号会在名称后追加 _1、_2；跳过则不重命名。"],
        ["是否重命名隐藏工作表", "否", "默认为否，只处理可见工作表。"],
        ["是否清洗非法字符", "是", "将 Excel 工作表名非法字符替换为下划线；如果为否，遇到非法字符则跳过。"],
        ["名称超过 31 字符时", "自动截断", "Excel 工作表名最长 31 个字符。自动截断时追加编号后仍不超过 31 字符。"],
        ["是否跳过临时文件", "是", "跳过以 ~$ 开头的 Office 临时文件。"],
    ]
    _append_rows(sheet, rows)
    _style_header(sheet, 1, len(rows[0]))
    _set_widths(sheet, {"A": 28, "B": 18, "C": 82})


def _write_rule_sheet(sheet, workbook_infos: list[dict]) -> None:
    _append_rows(sheet, [RULE_HEADERS])
    _style_header(sheet, 1, len(RULE_HEADERS))
    fill = PatternFill(fill_type="solid", fgColor="FFF2CC")

    for workbook_info in workbook_infos:
        workbook_path = _cell_to_text(workbook_info.get("workbook_path"))
        workbook_name = _cell_to_text(workbook_info.get("workbook_name")) or Path(workbook_path).name
        error_message = _cell_to_text(workbook_info.get("error"))
        if error_message:
            sheet.append([workbook_path, workbook_name, "", "", "跳过", error_message, "", ""])
            continue

        for index, sheet_info in enumerate(workbook_info.get("sheet_infos", []), start=1):
            name = _cell_to_text(sheet_info.get("name"))
            is_hidden = bool(sheet_info.get("is_hidden", False))
            order = sheet_info.get("order", index)
            sheet.append([workbook_path, workbook_name, name, "", "", "", "是" if is_hidden else "否", order])

    for row in range(2, sheet.max_row + 1):
        sheet.cell(row=row, column=4).fill = fill

    sheet.freeze_panes = "A2"
    _set_widths(sheet, {"A": 58, "B": 30, "C": 28, "D": 28, "E": 12, "F": 42, "G": 12, "H": 12})


def _write_log_sheet(sheet) -> None:
    _append_rows(sheet, [LOG_HEADERS])
    _style_header(sheet, 1, len(LOG_HEADERS))
    sheet.freeze_panes = "A2"
    _set_widths(sheet, {"A": 20, "B": 58, "C": 30, "D": 28, "E": 28, "F": 12, "G": 42})


def _read_choice_setting(
    value: str,
    choices: set[str],
    default: str,
    setting_name: str,
    warnings: list[str],
) -> str:
    normalized = value.strip()
    if normalized in choices:
        return normalized
    if normalized:
        warnings.append(f"参数“{setting_name}”无效，已使用默认值“{default}”。")
    return default


def _read_yes_no_setting(value: str, default: bool, setting_name: str, warnings: list[str]) -> bool:
    normalized = value.strip()
    if normalized == "是":
        return True
    if normalized == "否":
        return False
    if normalized:
        default_text = "是" if default else "否"
        warnings.append(f"参数“{setting_name}”无效，已使用默认值“{default_text}”。")
    return default


def _skip_action(rule: SheetRenameRule, target_name: str, message: str) -> SheetRenameAction:
    return SheetRenameAction(
        row_number=rule.row_number,
        workbook_path=rule.workbook_path,
        original_sheet_name=rule.original_sheet_name,
        target_sheet_name=target_name,
        status="跳过",
        message=message,
    )


def _copy_action(action: SheetRenameAction) -> SheetRenameAction:
    return SheetRenameAction(
        row_number=action.row_number,
        workbook_path=action.workbook_path,
        original_sheet_name=action.original_sheet_name,
        target_sheet_name=action.target_sheet_name,
        status=action.status,
        message=action.message,
    )


def _append_rows(sheet, rows: list[list[object]]) -> None:
    for row in rows:
        sheet.append(row)


def _style_header(sheet, row_number: int, column_count: int) -> None:
    for column in range(1, column_count + 1):
        sheet.cell(row=row_number, column=column).font = Font(bold=True)


def _set_widths(sheet, widths: dict[str, int]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _infer_output_dir(workbook_infos: list[dict]) -> Path:
    parent_dirs = []
    for workbook_info in workbook_infos:
        workbook_path = _cell_to_text(workbook_info.get("workbook_path"))
        if workbook_path:
            parent_dirs.append(str(Path(workbook_path).parent))

    if not parent_dirs:
        return Path(tempfile.gettempdir())
    try:
        common_dir = os.path.commonpath(parent_dirs)
    except ValueError:
        common_dir = parent_dirs[0]
    if common_dir and Path(common_dir).exists():
        return Path(common_dir)
    return Path(parent_dirs[0])


def _get_unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    counter = 2
    while True:
        candidate = path.with_name(f"{path.stem}_{counter}{path.suffix}")
        if not candidate.exists():
            return candidate
        counter += 1


def _normalize_sheet_name(name: str) -> str:
    return name.casefold()


def _normalize_path_for_compare(path: str) -> str:
    return os.path.normcase(os.path.abspath(path))


def _cell_to_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _cell_to_int(value: object) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _join_messages(*messages: str) -> str:
    return "；".join(message for message in messages if message)

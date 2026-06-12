import os
import shutil
import time
import zipfile
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell

from .color_sum_ops import (
    _cell_address_to_row_col,
    _find_openpyxl_sheet_name,
    _find_worksheet_by_name,
    _get_active_excel,
    _get_active_target_context,
    _get_cell_address_for_log,
    _get_cell_fill_color_info,
    _get_com_object_name,
    _get_com_property,
    _get_workbook_full_name,
    _log,
    _log_active_cell_diagnostics,
    build_cell_address,
    get_openpyxl_fill_key,
    group_cells_by_contiguous_rows,
    is_visible_sheet_state,
)


OPENPYXL_COLOR_WORKBOOK_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
KEEP_VBA_EXTENSIONS = {".xlsm", ".xltm"}
LOG_WORKBOOK_PREFIX = "按颜色清空内容_处理日志"
BACKUP_SUFFIX = "按颜色清空内容备份"
BATCH_BACKUP_DIR_PREFIX = "按颜色清空内容_备份"


def is_supported_openpyxl_color_workbook(path: str) -> bool:
    filename = os.path.basename(str(path))
    if not filename or filename.startswith("~$"):
        return False
    _, ext = os.path.splitext(filename)
    return ext.lower() in OPENPYXL_COLOR_WORKBOOK_EXTENSIONS


def should_keep_vba_for_workbook(path: str) -> bool:
    _, ext = os.path.splitext(str(path))
    return ext.lower() in KEEP_VBA_EXTENSIONS


def get_load_workbook_options_for_color_clear(path: str) -> dict:
    return {"keep_vba": should_keep_vba_for_workbook(path)}


def build_clear_by_color_backup_path(workbook_path: str, timestamp: str | None = None) -> str:
    abs_path = os.path.abspath(workbook_path)
    directory = os.path.dirname(abs_path)
    stem, ext = os.path.splitext(os.path.basename(abs_path))
    stamp = timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_base_name = f"{stem}_{BACKUP_SUFFIX}_{stamp}"

    counter = 1
    while True:
        suffix = "" if counter == 1 else f"_{counter}"
        backup_path = os.path.join(directory, f"{backup_base_name}{suffix}{ext}")
        if not os.path.exists(backup_path):
            return backup_path
        counter += 1


def create_clear_by_color_backup(workbook_path: str, timestamp: str | None = None) -> str:
    abs_path = os.path.abspath(workbook_path)
    if not os.path.exists(abs_path):
        raise FileNotFoundError(f"目标工作簿不存在：{abs_path}")

    backup_path = build_clear_by_color_backup_path(abs_path, timestamp=timestamp)
    shutil.copy2(abs_path, backup_path)
    return backup_path


def build_clear_by_color_batch_backup_dir(base_dir: str, timestamp: str | None = None) -> str:
    abs_base_dir = os.path.abspath(base_dir or ".")
    stamp = timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(abs_base_dir, f"{BATCH_BACKUP_DIR_PREFIX}_{stamp}")


def build_unique_backup_file_path(batch_backup_dir: str, file_name: str) -> str:
    stem, ext = os.path.splitext(file_name)
    counter = 1
    while True:
        suffix = "" if counter == 1 else f"_{counter}"
        backup_path = os.path.join(batch_backup_dir, f"{stem}{suffix}{ext}")
        if not os.path.exists(backup_path):
            return backup_path
        counter += 1


def build_clear_by_color_log_path(output_dir: str, timestamp: str | None = None) -> str:
    abs_output_dir = os.path.abspath(output_dir or ".")
    stamp = timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = f"{LOG_WORKBOOK_PREFIX}_{stamp}"

    counter = 1
    while True:
        suffix = "" if counter == 1 else f"_{counter}"
        log_path = os.path.join(abs_output_dir, f"{base_name}{suffix}.xlsx")
        if not os.path.exists(log_path):
            return log_path
        counter += 1


def build_clear_summary_log_record(
    file_path: str,
    matched_sheet_count: int,
    matched_cell_count: int,
    cleared_cell_count: int,
    save_mode: str,
    backup_mode: str,
    backup_path: str,
    status: str,
    note: str,
) -> dict:
    return {
        "文件路径": os.path.abspath(file_path),
        "匹配工作表数量": matched_sheet_count,
        "匹配单元格数量": matched_cell_count,
        "清空单元格数量": cleared_cell_count,
        "保存方式": save_mode,
        "备份方式": backup_mode,
        "备份路径": backup_path,
        "状态": status,
        "说明": note,
    }


def build_clear_detail_log_record(
    file_path: str,
    sheet_name: str,
    cell_address: str,
    status: str,
    note: str,
) -> dict:
    return {
        "文件路径": os.path.abspath(file_path),
        "工作表名": sheet_name,
        "单元格地址": cell_address,
        "状态": status,
        "说明": note,
    }


def write_clear_by_color_log_workbook(
    output_dir: str,
    summary_records: list[dict],
    detail_records: list[dict],
    timestamp: str | None = None,
) -> str:
    log_path = build_clear_by_color_log_path(output_dir, timestamp=timestamp)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "处理汇总"
    summary_headers = ["文件路径", "匹配工作表数量", "匹配单元格数量", "清空单元格数量", "保存方式", "备份方式", "备份路径", "状态", "说明"]
    summary_sheet.append(summary_headers)
    for record in summary_records:
        summary_sheet.append([record.get(header, "") for header in summary_headers])

    detail_sheet = workbook.create_sheet("处理明细")
    detail_headers = ["文件路径", "工作表名", "单元格地址", "状态", "说明"]
    detail_sheet.append(detail_headers)
    for record in detail_records:
        detail_sheet.append([record.get(header, "") for header in detail_headers])

    workbook.save(log_path)
    workbook.close()
    return log_path


def validate_openpyxl_color_scan_path(workbook_path: str, context_name: str = "工作簿") -> None:
    if not workbook_path:
        raise ValueError(f"{context_name}未保存，请先保存为 xlsx/xlsm/xltx/xltm 后重试。")

    abs_path = os.path.abspath(workbook_path)
    if not os.path.exists(abs_path):
        raise ValueError(f"{context_name}磁盘文件不可读，请确认文件已保存且存在：{abs_path}")

    if not is_supported_openpyxl_color_workbook(abs_path):
        raise ValueError(f"{context_name}仅支持 xlsx/xlsm/xltx/xltm；.xls 暂不支持。")


def read_selected_fill_key_from_workbook(
    workbook_path: str,
    selected_sheet_name: str,
    selected_cell_address: str,
):
    workbook = None
    try:
        workbook = _load_workbook_for_scan(workbook_path)
        actual_sheet_name = _find_openpyxl_sheet_name(workbook.sheetnames, selected_sheet_name)
        if actual_sheet_name is None:
            raise ValueError(f"当前活动工作簿中不存在工作表：{selected_sheet_name}。")
        return get_openpyxl_fill_key(workbook[actual_sheet_name][_normalize_cell_address(selected_cell_address)])
    finally:
        if workbook is not None:
            workbook.close()


def prepare_active_fill_color_context(logger=None) -> dict:
    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    try:
        excel = _get_active_excel(win32com.client, pythoncom, logger=logger)
        target_workbook, active_sheet, active_cell = _get_active_target_context(excel, logger=logger)
        workbook_path = _get_workbook_full_name(target_workbook)
        selected_sheet_name = active_sheet.Name
        selected_cell_address = _get_cell_address_for_log(active_cell)
        fill_info = _get_cell_fill_color_info(active_cell)
        _log_active_cell_diagnostics(logger, target_workbook, active_sheet, active_cell, fill_info)

        if not fill_info["has_fill"]:
            raise ValueError("请先选中一个带填充色的单元格。")

        validate_openpyxl_color_scan_path(workbook_path, context_name="当前活动工作簿")
        selected_color_key = read_selected_fill_key_from_workbook(
            workbook_path=workbook_path,
            selected_sheet_name=selected_sheet_name,
            selected_cell_address=selected_cell_address,
        )
        if selected_color_key is None:
            raise ValueError("当前选中单元格在已保存文件中没有识别到填充色；如果刚修改了颜色，请先保存当前工作簿后重试。")

        saved_state = _get_workbook_saved_state(target_workbook)
        current_color_text = f"Interior.Color={fill_info['color']}"
        _log(logger, "info", f"当前颜色：{current_color_text}")
        _log(logger, "info", f"当前颜色 key：{selected_color_key}")
        _log(logger, "info", f"当前工作簿已保存状态：{'已保存' if saved_state else '有未保存修改'}")

        return {
            "workbook_name": _get_com_object_name(target_workbook),
            "workbook_path": os.path.abspath(workbook_path),
            "selected_sheet_name": selected_sheet_name,
            "selected_cell_address": selected_cell_address,
            "selected_color_key": selected_color_key,
            "current_color_text": current_color_text,
            "fill_info": fill_info,
            "saved_state": saved_state,
        }
    finally:
        pythoncom.CoUninitialize()


def scan_workbook_color_matches(
    workbook_path: str,
    selected_color_key: tuple,
) -> dict:
    validate_openpyxl_color_scan_path(workbook_path)
    workbook = None
    try:
        workbook = _load_workbook_for_scan(workbook_path)
        return scan_loaded_workbook_color_matches(workbook, selected_color_key)
    finally:
        if workbook is not None:
            workbook.close()


def scan_loaded_workbook_color_matches(workbook, selected_color_key: tuple) -> dict:
    if selected_color_key is None:
        raise ValueError("颜色 key 不能为空。")

    sheet_plans = []
    for sheet in workbook.worksheets:
        if not is_visible_sheet_state(sheet.sheet_state):
            continue

        matched_cells = []
        for row in sheet.iter_rows():
            for cell in row:
                if get_openpyxl_fill_key(cell) == selected_color_key:
                    matched_cells.append(cell.coordinate)

        if matched_cells:
            sheet_plans.append({
                "sheet_name": sheet.title,
                "cells": matched_cells,
                "matched_cell_count": len(matched_cells),
            })

    return {
        "sheet_plans": sheet_plans,
        "matched_sheet_count": len(sheet_plans),
        "matched_cell_count": sum(plan["matched_cell_count"] for plan in sheet_plans),
    }


def build_clear_ranges_from_cells(cells: list[tuple[int, int]]) -> list[str]:
    range_addresses = []
    for group in group_cells_by_contiguous_rows(cells):
        start_row, start_col = group[0]
        end_row, end_col = group[-1]
        start_address = build_cell_address(start_row, start_col)
        end_address = build_cell_address(end_row, end_col)
        range_addresses.append(start_address if start_address == end_address else f"{start_address}:{end_address}")
    return range_addresses


def build_sheet_clear_plans(sheet_plans: list[dict]) -> list[dict]:
    result = []
    for sheet_plan in sheet_plans:
        cells = [_cell_address_to_row_col(address) for address in sheet_plan["cells"]]
        result.append({
            "sheet_name": sheet_plan["sheet_name"],
            "cells": list(sheet_plan["cells"]),
            "matched_cell_count": len(sheet_plan["cells"]),
            "range_addresses": build_clear_ranges_from_cells(cells),
            "range_count": len(build_clear_ranges_from_cells(cells)),
        })
    return result


def workbook_has_external_links(workbook_path: str) -> bool:
    abs_path = os.path.abspath(workbook_path)
    try:
        with zipfile.ZipFile(abs_path, "r") as archive:
            return any(name.startswith("xl/externalLinks/") for name in archive.namelist())
    except (FileNotFoundError, zipfile.BadZipFile, OSError):
        return False


def resolve_multi_workbook_save_mode(workbook_path: str) -> str:
    return "com" if workbook_has_external_links(workbook_path) else "openpyxl"


def analyze_multi_workbook_clear_target(workbook_path: str, selected_color_key: tuple) -> dict:
    abs_path = os.path.abspath(workbook_path)
    record = {
        "file_path": abs_path,
        "file_name": os.path.basename(abs_path),
        "is_supported": False,
        "has_external_links": False,
        "matched_sheet_count": 0,
        "matched_cell_count": 0,
        "save_mode": "跳过",
        "needs_modify": False,
        "backup_mode": "跳过",
        "backup_path": "",
        "status": "跳过",
        "note": "",
        "sheet_plans": [],
        "sheet_clear_plans": [],
    }

    if os.path.basename(abs_path).startswith("~$"):
        record["note"] = "临时文件已跳过。"
        return record

    if not os.path.exists(abs_path):
        record["status"] = "失败"
        record["note"] = "目标文件不存在。"
        return record

    _, ext = os.path.splitext(abs_path)
    if ext.lower() == ".xls":
        record["note"] = ".xls 暂不支持，请另存为 xlsx/xlsm/xltx/xltm。"
        return record

    if not is_supported_openpyxl_color_workbook(abs_path):
        record["note"] = "不支持的文件类型。"
        return record

    record["is_supported"] = True
    record["has_external_links"] = workbook_has_external_links(abs_path)

    color_plan = scan_workbook_color_matches(abs_path, selected_color_key)
    record["matched_sheet_count"] = color_plan["matched_sheet_count"]
    record["matched_cell_count"] = color_plan["matched_cell_count"]
    record["sheet_plans"] = color_plan["sheet_plans"]
    record["sheet_clear_plans"] = build_sheet_clear_plans(color_plan["sheet_plans"])

    if color_plan["matched_cell_count"] == 0:
        record["backup_mode"] = "无需备份"
        record["note"] = "未找到同色单元格。"
        return record

    record["save_mode"] = resolve_multi_workbook_save_mode(abs_path)
    record["needs_modify"] = True
    record["status"] = "待处理"
    record["backup_mode"] = ""
    return record


def collect_multi_workbook_clear_records(target_paths: list[str], selected_color_key: tuple, logger=None) -> list[dict]:
    records = []
    total = len(target_paths)
    for index, target_path in enumerate(target_paths, start=1):
        abs_path = os.path.abspath(target_path)
        filename = os.path.basename(abs_path)
        _log(logger, "info", f"正在扫描目标文件 {index}/{total}：{filename}")
        try:
            scan_start = time.perf_counter()
            record = analyze_multi_workbook_clear_target(abs_path, selected_color_key)
            scan_seconds = time.perf_counter() - scan_start
            record["scan_seconds"] = scan_seconds
            _log(logger, "info", f"扫描耗时：{scan_seconds:.2f} 秒。")
            _log(logger, "info", f"匹配 sheet 数：{record['matched_sheet_count']}")
            _log(logger, "info", f"匹配单元格数：{record['matched_cell_count']}")
            if record["needs_modify"]:
                _log(logger, "info", f"保存方式：{record['save_mode']}")
            records.append(record)
        except Exception as e:
            records.append({
                "file_path": abs_path,
                "file_name": filename,
                "is_supported": False,
                "has_external_links": False,
                "matched_sheet_count": 0,
                "matched_cell_count": 0,
                "save_mode": "失败",
                "needs_modify": False,
                "backup_mode": "跳过",
                "backup_path": "",
                "status": "失败",
                "note": str(e),
                "sheet_plans": [],
                "sheet_clear_plans": [],
                "scan_seconds": 0.0,
            })
    return records


def get_multi_workbook_backup_candidates(records: list[dict]) -> list[dict]:
    return [record for record in records if record.get("needs_modify")]


def apply_multi_workbook_backup_plan(
    records: list[dict],
    base_dir: str,
    skip_backup: bool,
    timestamp: str | None = None,
) -> str:
    candidates = get_multi_workbook_backup_candidates(records)
    if not candidates:
        for record in records:
            if record.get("backup_mode") == "":
                record["backup_mode"] = "无需备份"
        return ""

    if skip_backup:
        for record in candidates:
            record["backup_mode"] = "用户选择不备份"
            record["backup_path"] = "用户选择不备份"
        return ""

    batch_backup_dir = build_clear_by_color_batch_backup_dir(base_dir, timestamp=timestamp)
    os.makedirs(batch_backup_dir, exist_ok=True)
    for record in candidates:
        backup_path = build_unique_backup_file_path(batch_backup_dir, record["file_name"])
        shutil.copy2(record["file_path"], backup_path)
        record["backup_mode"] = "批次备份"
        record["backup_path"] = backup_path
    return batch_backup_dir


def plan_clear_active_workbook_by_color(logger=None) -> dict:
    context = prepare_active_fill_color_context(logger=logger)
    if not context["saved_state"]:
        raise ValueError("当前活动工作簿存在未保存修改，请先保存后再执行按颜色清空内容。")

    _log(logger, "info", "目标模式：当前活动工作簿")
    _log(logger, "info", f"当前工作簿：{context['workbook_name']}")
    _log(logger, "info", f"当前工作表：{context['selected_sheet_name']}")
    _log(logger, "info", f"当前单元格：{context['selected_cell_address']}")
    _log(logger, "info", f"当前颜色：{context['current_color_text']}")

    scan_start = time.perf_counter()
    _log(logger, "info", "使用 openpyxl 扫描当前工作簿颜色区域。")
    color_plan = scan_workbook_color_matches(
        workbook_path=context["workbook_path"],
        selected_color_key=context["selected_color_key"],
    )
    scan_seconds = time.perf_counter() - scan_start
    _log(logger, "info", f"扫描耗时：{scan_seconds:.2f} 秒。")
    _log(logger, "info", f"匹配 sheet 数：{color_plan['matched_sheet_count']}")
    _log(logger, "info", f"同色单元格总数：{color_plan['matched_cell_count']}")
    _log(logger, "info", "清空范围：所有匹配 sheet")

    return {
        **context,
        **color_plan,
        "sheet_clear_plans": build_sheet_clear_plans(color_plan["sheet_plans"]),
        "scan_seconds": scan_seconds,
    }


def execute_clear_active_workbook_plan(plan: dict, logger=None) -> dict:
    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    try:
        excel = _get_active_excel(win32com.client, pythoncom, logger=logger)
        workbook = _find_open_workbook_by_path(excel, plan["workbook_path"])
        if workbook is None:
            raise RuntimeError("未找到当前活动工作簿对应的已打开文件，请确认原工作簿仍在当前 Excel 中打开。")

        if not _get_workbook_saved_state(workbook):
            raise ValueError("当前活动工作簿存在未保存修改，请先保存后再执行按颜色清空内容。")

        clear_result = clear_open_workbook_with_com(
            workbook,
            plan["sheet_clear_plans"],
            logger=logger,
        )
        _log(logger, "info", f"清空 sheet 数：{plan['matched_sheet_count']}")
        _log(logger, "info", f"清空单元格数：{clear_result['cleared_cell_count']}")
        _log(logger, "info", "按颜色清空内容完成：当前工作簿未自动保存，请检查后自行保存。")

        return {
            "target_workbook_name": workbook.Name,
            "workbook_path": plan["workbook_path"],
            "matched_sheet_count": plan["matched_sheet_count"],
            "matched_cell_count": plan["matched_cell_count"],
            "cleared_sheet_count": plan["matched_sheet_count"],
            "cleared_cell_count": clear_result["cleared_cell_count"],
            "skipped_merged_cell_count": clear_result["skipped_merged_cell_count"],
            "range_group_count": clear_result["range_group_count"],
            "current_color_text": plan["current_color_text"],
            "scan_seconds": plan["scan_seconds"],
        }
    finally:
        pythoncom.CoUninitialize()


def clear_multiple_workbooks_by_color(target_paths: list[str], skip_backup: bool = False, logger=None) -> dict:
    if not target_paths:
        raise ValueError("请选择至少一个目标工作簿。")

    context = prepare_active_fill_color_context(logger=logger)
    _log(logger, "info", "目标模式：多个工作簿")
    _log(logger, "info", f"当前工作簿：{context['workbook_name']}")
    _log(logger, "info", f"当前工作表：{context['selected_sheet_name']}")
    _log(logger, "info", f"当前单元格：{context['selected_cell_address']}")
    _log(logger, "info", f"当前颜色：{context['current_color_text']}")
    _log(logger, "info", f"目标文件数量：{len(target_paths)}")

    summary_records = []
    detail_records = []
    modified_file_count = 0
    skipped_file_count = 0
    failed_file_count = 0
    cleared_cell_total = 0
    output_dir = os.path.dirname(os.path.abspath(target_paths[0])) or os.getcwd()
    records = collect_multi_workbook_clear_records(target_paths, context["selected_color_key"], logger=logger)
    batch_backup_dir = apply_multi_workbook_backup_plan(records, output_dir, skip_backup=skip_backup)
    if batch_backup_dir:
        _log(logger, "info", f"批次备份文件夹：{batch_backup_dir}")
    elif skip_backup and get_multi_workbook_backup_candidates(records):
        _log(logger, "info", "用户已选择跳过备份，本次不会生成备份文件。")

    for record in records:
        if record["needs_modify"]:
            try:
                if record["save_mode"] == "com":
                    clear_result = clear_closed_workbook_with_com(
                        record["file_path"],
                        record["sheet_clear_plans"],
                        logger=logger,
                    )
                else:
                    clear_result = clear_workbook_file_with_openpyxl(
                        record["file_path"],
                        record["sheet_plans"],
                        detail_records=detail_records,
                        logger=logger,
                    )

                record["status"] = "成功"
                record["cleared_cell_count"] = clear_result["cleared_cell_count"]
                record["note"] = (
                    "已清空值和公式，保留格式。"
                    if clear_result["skipped_merged_cell_count"] == 0
                    else f"已清空值和公式，保留格式；跳过 {clear_result['skipped_merged_cell_count']} 个合并区域非左上角单元格。"
                )
                modified_file_count += 1
                cleared_cell_total += clear_result["cleared_cell_count"]
            except Exception as e:
                record["status"] = "失败"
                record["note"] = str(e)
                failed_file_count += 1
        elif record["status"] == "失败":
            failed_file_count += 1
        else:
            skipped_file_count += 1

        summary_records.append(
            build_clear_summary_log_record(
                record["file_path"],
                record["matched_sheet_count"],
                record["matched_cell_count"],
                record.get("cleared_cell_count", 0),
                record["save_mode"],
                record["backup_mode"],
                record["backup_path"],
                record["status"],
                record["note"],
            )
        )

    log_path = write_clear_by_color_log_workbook(output_dir, summary_records, detail_records)
    _log(logger, "info", f"处理日志：{log_path}")
    _log(logger, "info", f"成功处理文件数：{modified_file_count}")
    _log(logger, "info", f"跳过文件数：{skipped_file_count}")
    _log(logger, "info", f"失败文件数：{failed_file_count}")
    _log(logger, "info", f"清空单元格总数：{cleared_cell_total}")

    return {
        "current_color_text": context["current_color_text"],
        "target_file_count": len(target_paths),
        "processed_file_count": modified_file_count,
        "skipped_file_count": skipped_file_count,
        "failed_file_count": failed_file_count,
        "cleared_cell_count": cleared_cell_total,
        "log_path": log_path,
        "batch_backup_dir": batch_backup_dir,
        "summary_records": summary_records,
    }


def clear_loaded_workbook_cells(
    workbook,
    workbook_path: str,
    sheet_plans: list[dict],
    detail_records: list[dict] | None = None,
    logger=None,
) -> dict:
    cleared_cell_count = 0
    skipped_merged_cell_count = 0
    detail_records = detail_records if detail_records is not None else []

    for sheet_plan in sheet_plans:
        worksheet = workbook[sheet_plan["sheet_name"]]
        _log(logger, "info", f"正在清空工作表：{sheet_plan['sheet_name']}；匹配单元格 {len(sheet_plan['cells'])} 个。")
        for cell_address in sheet_plan["cells"]:
            cell = worksheet[cell_address]
            if isinstance(cell, MergedCell):
                skipped_merged_cell_count += 1
                detail_records.append(
                    build_clear_detail_log_record(
                        workbook_path,
                        sheet_plan["sheet_name"],
                        cell_address,
                        "跳过",
                        "合并区域非左上角单元格，未清空。",
                    )
                )
                continue
            cell.value = None
            cleared_cell_count += 1

    return {
        "cleared_cell_count": cleared_cell_count,
        "skipped_merged_cell_count": skipped_merged_cell_count,
    }


def clear_workbook_file_with_openpyxl(
    workbook_path: str,
    sheet_plans: list[dict],
    detail_records: list[dict] | None = None,
    logger=None,
) -> dict:
    workbook = None
    try:
        workbook = load_workbook(workbook_path, keep_vba=should_keep_vba_for_workbook(workbook_path))
        result = clear_loaded_workbook_cells(
            workbook,
            workbook_path,
            sheet_plans,
            detail_records=detail_records,
            logger=logger,
        )
        workbook.save(workbook_path)
        return {
            **result,
            "range_group_count": 0,
        }
    finally:
        if workbook is not None:
            workbook.close()


def clear_open_workbook_with_com(workbook, sheet_clear_plans: list[dict], logger=None) -> dict:
    cleared_cell_count = 0
    skipped_merged_cell_count = 0
    range_group_count = 0

    for sheet_plan in sheet_clear_plans:
        worksheet = _find_worksheet_by_name(workbook, sheet_plan["sheet_name"])
        if worksheet is None:
            raise RuntimeError(f"工作簿中不存在工作表：{sheet_plan['sheet_name']}。")

        _log(
            logger,
            "info",
            f"正在清空目标工作表：{sheet_plan['sheet_name']}；待清空单元格 {sheet_plan['matched_cell_count']} 个，批量区域 {sheet_plan['range_count']} 组。",
        )
        for range_address in sheet_plan["range_addresses"]:
            worksheet.Range(range_address).ClearContents()

        cleared_cell_count += sheet_plan["matched_cell_count"]
        range_group_count += sheet_plan["range_count"]

    return {
        "cleared_cell_count": cleared_cell_count,
        "skipped_merged_cell_count": skipped_merged_cell_count,
        "range_group_count": range_group_count,
    }


def clear_closed_workbook_with_com(workbook_path: str, sheet_clear_plans: list[dict], logger=None) -> dict:
    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(os.path.abspath(workbook_path), UpdateLinks=0)
        result = clear_open_workbook_with_com(workbook, sheet_clear_plans, logger=logger)
        workbook.Save()
        return result
    finally:
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()


def _load_workbook_for_scan(path: str):
    return load_workbook(path, read_only=True, data_only=False, keep_vba=should_keep_vba_for_workbook(path))


def _normalize_cell_address(cell_address: str) -> str:
    return str(cell_address).replace("$", "").strip()


def _find_open_workbook_by_path(excel, workbook_path: str):
    normalized_path = os.path.normcase(os.path.abspath(workbook_path))
    for workbook in excel.Workbooks:
        try:
            full_name = os.path.normcase(os.path.abspath(workbook.FullName))
        except Exception:
            continue
        if full_name == normalized_path:
            return workbook
    return None


def _get_workbook_saved_state(workbook) -> bool:
    saved = _get_com_property(workbook, "Saved")
    return bool(saved)

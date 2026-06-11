import os
import shutil
from copy import copy
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Font


DIRECTORY_SHEET_NAME = "多簿汇总目录"
LOG_SHEET_NAME = "多簿汇总日志"
SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm"}
XLS_EXTENSION = ".xls"
MAX_EXCEL_SHEET_NAME_LENGTH = 31
INVALID_SHEET_NAME_CHARS = set(r":\/?*[]")


def is_supported_openpyxl_workbook(path: str) -> bool:
    return os.path.splitext(str(path))[1].lower() in SUPPORTED_EXTENSIONS


def is_office_temp_file(path: str) -> bool:
    return os.path.basename(str(path)).startswith("~$")


def get_first_visible_sheet_name(workbook) -> str | None:
    for worksheet in workbook.worksheets:
        if worksheet.sheet_state == "visible":
            return worksheet.title
    return None


def resolve_source_sheet_name(workbook, requested_sheet_name: str | None) -> tuple[str | None, list[str], str]:
    available_sheet_names = [
        worksheet.title
        for worksheet in workbook.worksheets
        if worksheet.sheet_state == "visible"
    ]
    requested_name = (requested_sheet_name or "").strip()

    if not requested_name:
        first_visible_sheet_name = get_first_visible_sheet_name(workbook)
        if first_visible_sheet_name is None:
            return None, available_sheet_names, "没有可见工作表"
        return first_visible_sheet_name, available_sheet_names, "未指定源 Sheet 名，已使用第一个可见 sheet"

    normalized_requested_name = _normalize_sheet_lookup_name(requested_name)
    for sheet_name in available_sheet_names:
        if _normalize_sheet_lookup_name(sheet_name) == normalized_requested_name:
            return sheet_name, available_sheet_names, "匹配指定源 Sheet 名"

    return None, available_sheet_names, f"找不到指定源 Sheet：{requested_name}"


def build_unique_sheet_name(base_name: str, existing_names: set[str]) -> str:
    cleaned_base_name = _clean_sheet_name(base_name)
    normalized_existing_names = {_normalize_sheet_lookup_name(name) for name in existing_names}
    candidate = cleaned_base_name[:MAX_EXCEL_SHEET_NAME_LENGTH]

    if _normalize_sheet_lookup_name(candidate) not in normalized_existing_names:
        return candidate

    counter = 2
    while True:
        suffix = f"_{counter}"
        max_base_length = MAX_EXCEL_SHEET_NAME_LENGTH - len(suffix)
        candidate = f"{cleaned_base_name[:max_base_length]}{suffix}"
        if _normalize_sheet_lookup_name(candidate) not in normalized_existing_names:
            return candidate
        counter += 1


def copy_worksheet_basic(source_ws, target_wb, target_sheet_name: str, values_only: bool = False, cached_source_ws=None):
    target_ws = target_wb.create_sheet(title=target_sheet_name)
    formula_count = 0
    empty_formula_cache_count = 0

    for row in source_ws.iter_rows():
        for source_cell in row:
            target_cell = target_ws[source_cell.coordinate]
            source_value = source_cell.value
            is_formula = _is_formula_value(source_value)

            if is_formula:
                formula_count += 1

            if values_only and is_formula:
                cached_value = cached_source_ws[source_cell.coordinate].value if cached_source_ws is not None else None
                if cached_value is None:
                    target_cell.value = source_value
                    empty_formula_cache_count += 1
                else:
                    target_cell.value = cached_value
            else:
                target_cell.value = source_value

            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.fill = copy(source_cell.fill)
                target_cell.border = copy(source_cell.border)
                target_cell.alignment = copy(source_cell.alignment)
                target_cell.number_format = source_cell.number_format
                target_cell.protection = copy(source_cell.protection)

            if source_cell.hyperlink:
                hyperlink = copy(source_cell.hyperlink)
                hyperlink.ref = target_cell.coordinate
                target_cell._hyperlink = hyperlink

    _copy_dimensions(source_ws, target_ws)
    _copy_merged_cells(source_ws, target_ws)
    _copy_basic_sheet_settings(source_ws, target_ws)

    target_ws._workbook_merge_formula_count = formula_count
    target_ws._workbook_merge_empty_formula_cache_count = empty_formula_cache_count
    return target_ws


def create_target_backup(target_path: str) -> str:
    abs_target_path = os.path.abspath(target_path)
    if not os.path.exists(abs_target_path):
        raise FileNotFoundError(f"目标工作簿不存在：{abs_target_path}")

    directory = os.path.dirname(abs_target_path)
    file_stem, extension = os.path.splitext(os.path.basename(abs_target_path))
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_base_name = f"{file_stem}_多簿汇总备份_{timestamp}"

    counter = 1
    while True:
        suffix = "" if counter == 1 else f"_{counter}"
        backup_path = os.path.join(directory, f"{backup_base_name}{suffix}{extension}")
        if not os.path.exists(backup_path):
            shutil.copy2(abs_target_path, backup_path)
            return backup_path
        counter += 1


def merge_workbooks_to_existing_workbook(
    source_paths: list[str],
    target_path: str,
    requested_sheet_name: str | None,
    values_only: bool,
    logger=None,
) -> dict:
    if not source_paths:
        raise ValueError("请选择源 Excel 文件。")

    abs_target_path = os.path.abspath(target_path)
    if not is_supported_openpyxl_workbook(abs_target_path):
        raise ValueError("目标工作簿仅支持 .xlsx / .xlsm。")
    if not os.path.exists(abs_target_path):
        raise FileNotFoundError(f"目标工作簿不存在：{abs_target_path}")

    try:
        backup_path = create_target_backup(abs_target_path)
    except Exception as error:
        raise RuntimeError(
            "目标工作簿无法备份，请先关闭目标工作簿后重试，并确认目标文件所在目录可写。"
            f" 详细信息：{error}"
        ) from error
    _log(logger, "info", f"已生成目标工作簿备份：{backup_path}")

    target_wb = None
    records = []
    directory_rows = []
    success_count = 0
    skipped_count = 0
    failed_count = 0

    try:
        target_wb = load_workbook(abs_target_path, keep_vba=_is_xlsm(abs_target_path))
        _remove_fixed_output_sheets(target_wb)

        existing_names = set(target_wb.sheetnames)
        existing_names.update({DIRECTORY_SHEET_NAME, LOG_SHEET_NAME})

        for source_path in source_paths:
            record = _new_log_record(source_path, requested_sheet_name, values_only)
            records.append(record)
            abs_source_path = record["源文件路径"]

            if os.path.normcase(abs_source_path) == os.path.normcase(abs_target_path):
                skipped_count += 1
                _mark_record(record, "跳过", "源文件与目标工作簿相同，已跳过")
                _log(logger, "info", f"跳过目标工作簿本身：{abs_source_path}")
                continue

            if is_office_temp_file(abs_source_path):
                skipped_count += 1
                _mark_record(record, "跳过", "临时文件已跳过")
                _log(logger, "info", f"跳过临时文件：{abs_source_path}")
                continue

            source_ext = os.path.splitext(abs_source_path)[1].lower()
            if source_ext == XLS_EXTENSION:
                skipped_count += 1
                _mark_record(record, "跳过", "请另存为 xlsx/xlsm 后再处理")
                _log(logger, "info", f"跳过 .xls 文件：{abs_source_path}")
                continue
            if not is_supported_openpyxl_workbook(abs_source_path):
                skipped_count += 1
                _mark_record(record, "跳过", "不支持的文件类型")
                _log(logger, "info", f"跳过不支持文件：{abs_source_path}")
                continue
            if not os.path.exists(abs_source_path):
                failed_count += 1
                _mark_record(record, "失败", "源文件不存在")
                _log(logger, "error", f"源文件不存在：{abs_source_path}")
                continue

            source_wb = None
            cached_source_wb = None
            try:
                _log(logger, "info", f"正在导入源文件：{abs_source_path}")
                source_wb = load_workbook(abs_source_path, data_only=False, keep_vba=False)
                cached_source_wb = load_workbook(abs_source_path, data_only=True, keep_vba=False) if values_only else None

                actual_sheet_name, available_sheet_names, resolve_message = resolve_source_sheet_name(
                    source_wb,
                    requested_sheet_name,
                )
                record["源文件可用 Sheet 列表"] = "、".join(available_sheet_names)
                record["说明"] = resolve_message

                if actual_sheet_name is None:
                    skipped_count += 1
                    _mark_record(record, "跳过", resolve_message)
                    _log(logger, "info", f"跳过文件：{os.path.basename(abs_source_path)}，{resolve_message}")
                    continue

                source_ws = source_wb[actual_sheet_name]
                cached_source_ws = cached_source_wb[actual_sheet_name] if cached_source_wb is not None else None
                target_sheet_name = build_unique_sheet_name(os.path.splitext(os.path.basename(abs_source_path))[0], existing_names)
                target_ws = copy_worksheet_basic(
                    source_ws,
                    target_wb,
                    target_sheet_name,
                    values_only=values_only,
                    cached_source_ws=cached_source_ws,
                )
                existing_names.add(target_ws.title)

                formula_count = getattr(target_ws, "_workbook_merge_formula_count", 0)
                empty_formula_cache_count = getattr(target_ws, "_workbook_merge_empty_formula_cache_count", 0)
                description = resolve_message
                if values_only and empty_formula_cache_count:
                    description = f"{description}；公式缓存为空，已保留公式 {empty_formula_cache_count} 个"

                success_count += 1
                record.update(
                    {
                        "实际导入 Sheet 名": actual_sheet_name,
                        "目标 Sheet 名": target_ws.title,
                        "状态": "成功",
                        "说明": description,
                        "处理行数": source_ws.max_row,
                        "处理列数": source_ws.max_column,
                        "是否存在公式": "是" if formula_count else "否",
                    }
                )
                directory_rows.append(
                    {
                        "目标 Sheet 名": target_ws.title,
                        "来源文件名": os.path.basename(abs_source_path),
                        "来源文件路径": abs_source_path,
                        "来源 Sheet 名": actual_sheet_name,
                        "备注": description,
                    }
                )
                _log(logger, "info", f"成功导入：{os.path.basename(abs_source_path)} -> {target_ws.title}")
            except Exception as error:
                failed_count += 1
                _mark_record(record, "失败", f"处理失败：{error}")
                _log(logger, "error", f"处理失败：{abs_source_path}，{error}")
            finally:
                _close_workbook(cached_source_wb)
                _close_workbook(source_wb)

        _create_directory_sheet(target_wb, directory_rows)
        _create_log_sheet(target_wb, records)
        _disable_forced_recalculation(target_wb)

        try:
            target_wb.save(abs_target_path)
        except Exception as error:
            raise RuntimeError(
                "目标工作簿无法保存，请先关闭目标工作簿后重试，并确认文件可写。"
                f" 详细信息：{error}"
            ) from error

    finally:
        _close_workbook(target_wb)

    return {
        "target_path": abs_target_path,
        "backup_path": backup_path,
        "success_count": success_count,
        "skipped_count": skipped_count,
        "failed_count": failed_count,
        "directory_sheet_name": DIRECTORY_SHEET_NAME,
        "log_sheet_name": LOG_SHEET_NAME,
        "records": records,
    }


def _normalize_sheet_lookup_name(name: str) -> str:
    return str(name).strip().casefold()


def _clean_sheet_name(name: str) -> str:
    cleaned_name = "".join(
        char
        for char in str(name or "").strip().strip("'")
        if char not in INVALID_SHEET_NAME_CHARS
    )
    return cleaned_name or "Sheet"


def _is_formula_value(value) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _is_xlsm(path: str) -> bool:
    return os.path.splitext(str(path))[1].lower() == ".xlsm"


def _log(logger, level: str, message: str) -> None:
    if logger is not None:
        getattr(logger, level)(message)


def _close_workbook(workbook) -> None:
    if workbook is not None:
        try:
            workbook.close()
        except Exception:
            pass


def _copy_dimensions(source_ws, target_ws) -> None:
    for row_index, row_dimension in source_ws.row_dimensions.items():
        target_dimension = target_ws.row_dimensions[row_index]
        target_dimension.height = row_dimension.height
        target_dimension.hidden = row_dimension.hidden
        target_dimension.outlineLevel = row_dimension.outlineLevel
        target_dimension.collapsed = row_dimension.collapsed

    for column_key, column_dimension in source_ws.column_dimensions.items():
        target_dimension = target_ws.column_dimensions[column_key]
        target_dimension.width = column_dimension.width
        target_dimension.hidden = column_dimension.hidden
        target_dimension.bestFit = column_dimension.bestFit
        target_dimension.outlineLevel = column_dimension.outlineLevel
        target_dimension.collapsed = column_dimension.collapsed


def _copy_merged_cells(source_ws, target_ws) -> None:
    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))


def _copy_basic_sheet_settings(source_ws, target_ws) -> None:
    target_ws.freeze_panes = source_ws.freeze_panes
    target_ws.sheet_format.defaultColWidth = source_ws.sheet_format.defaultColWidth
    target_ws.sheet_format.defaultRowHeight = source_ws.sheet_format.defaultRowHeight

    for attribute_name in (
        "showGridLines",
        "showRowColHeaders",
        "showRuler",
        "showZeros",
        "rightToLeft",
        "view",
        "zoomScale",
        "zoomScaleNormal",
    ):
        if hasattr(source_ws.sheet_view, attribute_name):
            setattr(target_ws.sheet_view, attribute_name, getattr(source_ws.sheet_view, attribute_name))

    if source_ws.sheet_properties.tabColor is not None:
        target_ws.sheet_properties.tabColor = copy(source_ws.sheet_properties.tabColor)


def _remove_fixed_output_sheets(workbook) -> None:
    for sheet_name in (DIRECTORY_SHEET_NAME, LOG_SHEET_NAME):
        if sheet_name in workbook.sheetnames:
            workbook.remove(workbook[sheet_name])


def _new_log_record(source_path: str, requested_sheet_name: str | None, values_only: bool) -> dict:
    abs_source_path = os.path.abspath(source_path)
    return {
        "处理时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "源文件名": os.path.basename(abs_source_path),
        "源文件路径": abs_source_path,
        "指定源 Sheet 名": (requested_sheet_name or "").strip(),
        "实际导入 Sheet 名": "",
        "目标 Sheet 名": "",
        "状态": "",
        "说明": "",
        "源文件可用 Sheet 列表": "",
        "处理行数": "",
        "处理列数": "",
        "是否只复制值": "是" if values_only else "否",
        "是否存在公式": "否",
    }


def _mark_record(record: dict, status: str, description: str) -> None:
    record["状态"] = status
    record["说明"] = description


def _create_directory_sheet(workbook, directory_rows: list[dict]) -> None:
    worksheet = workbook.create_sheet(DIRECTORY_SHEET_NAME, 0)
    headers = ["序号", "目标 Sheet 名", "来源文件名", "来源文件路径", "来源 Sheet 名", "跳转链接", "备注"]
    _write_header_row(worksheet, headers)

    for row_index, row_data in enumerate(directory_rows, start=2):
        worksheet.cell(row=row_index, column=1, value=row_index - 1)
        worksheet.cell(row=row_index, column=2, value=row_data["目标 Sheet 名"])
        worksheet.cell(row=row_index, column=3, value=row_data["来源文件名"])
        worksheet.cell(row=row_index, column=4, value=row_data["来源文件路径"])
        worksheet.cell(row=row_index, column=5, value=row_data["来源 Sheet 名"])
        link_cell = worksheet.cell(row=row_index, column=6, value="打开")
        link_cell.hyperlink = _build_internal_sheet_link(row_data["目标 Sheet 名"])
        link_cell.style = "Hyperlink"
        worksheet.cell(row=row_index, column=7, value=row_data["备注"])

    _set_basic_table_widths(worksheet, [8, 24, 28, 56, 24, 12, 40])


def _create_log_sheet(workbook, records: list[dict]) -> None:
    worksheet = workbook.create_sheet(LOG_SHEET_NAME, 1)
    headers = [
        "处理时间",
        "源文件名",
        "源文件路径",
        "指定源 Sheet 名",
        "实际导入 Sheet 名",
        "目标 Sheet 名",
        "状态",
        "说明",
        "源文件可用 Sheet 列表",
        "处理行数",
        "处理列数",
        "是否只复制值",
        "是否存在公式",
    ]
    _write_header_row(worksheet, headers)

    for row_index, record in enumerate(records, start=2):
        for column_index, header in enumerate(headers, start=1):
            worksheet.cell(row=row_index, column=column_index, value=record.get(header, ""))

    _set_basic_table_widths(worksheet, [20, 28, 56, 20, 20, 20, 10, 42, 42, 10, 10, 12, 12])


def _write_header_row(worksheet, headers: list[str]) -> None:
    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=header)
        cell.font = Font(bold=True)
    worksheet.freeze_panes = "A2"


def _set_basic_table_widths(worksheet, widths: list[int]) -> None:
    for column_index, width in enumerate(widths, start=1):
        column_letter = worksheet.cell(row=1, column=column_index).column_letter
        worksheet.column_dimensions[column_letter].width = width


def _build_internal_sheet_link(sheet_name: str) -> str:
    escaped_sheet_name = sheet_name.replace("'", "''")
    return f"#'{escaped_sheet_name}'!A1"


def _disable_forced_recalculation(workbook) -> None:
    calculation = getattr(workbook, "calculation", None)
    if calculation is None:
        return
    if hasattr(calculation, "fullCalcOnLoad"):
        calculation.fullCalcOnLoad = False
    if hasattr(calculation, "forceFullCalc"):
        calculation.forceFullCalc = False

import os
import shutil
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from .link_replace_ops import XL_EXCEL_LINKS, normalize_com_link_sources


SUPPORTED_TEMPLATE_EXTENSIONS = {".xlsx", ".xlsm"}
SUPPORTED_TB_EXTENSIONS = {".xlsx", ".xlsm"}
OUTPUT_STEM_SUFFIX = "_报表"
LOG_WORKBOOK_NAME = "按模板多选TB生成报表_处理日志.xlsx"
LOG_HEADERS = [
    "序号",
    "TB 文件名",
    "TB 文件路径",
    "输出文件名",
    "输出文件路径",
    "被替换的旧链接",
    "新链接",
    "状态",
    "说明",
]


@dataclass
class TemplateTbReportRecord:
    index: int
    tb_path: str
    tb_name: str
    output_name: str = ""
    output_path: str = ""
    old_link_path: str = ""
    new_link_path: str = ""
    status: str = ""
    message: str = ""


@dataclass
class TemplateTbReportPlan:
    template_path: str
    template_suffix: str
    output_dir: str
    old_link_path: str
    records: list[TemplateTbReportRecord] = field(default_factory=list)


def is_office_temp_file(path: str) -> bool:
    return os.path.basename(str(path)).startswith("~$")


def is_supported_template_file(path: str) -> bool:
    filename = os.path.basename(str(path))
    if not filename or filename.startswith("~$"):
        return False
    return Path(filename).suffix.lower() in SUPPORTED_TEMPLATE_EXTENSIONS


def is_supported_tb_file(path: str) -> bool:
    filename = os.path.basename(str(path))
    if not filename or filename.startswith("~$"):
        return False
    return Path(filename).suffix.lower() in SUPPORTED_TB_EXTENSIONS


def build_output_filename(tb_path: str, template_suffix: str) -> str:
    stem = Path(str(tb_path)).stem
    suffix = template_suffix if template_suffix.startswith(".") else f".{template_suffix}"
    return f"{stem}{OUTPUT_STEM_SUFFIX}{suffix}"


def build_unique_output_path(
    output_dir: str,
    output_name: str,
    reserved_paths: set[str] | None = None,
) -> str:
    directory = Path(output_dir).resolve()
    reserved = reserved_paths or set()
    candidate = directory / output_name
    if not candidate.exists() and str(candidate) not in reserved:
        return str(candidate)

    stem = candidate.stem
    suffix = candidate.suffix
    counter = 2
    while True:
        next_candidate = directory / f"{stem}_{counter}{suffix}"
        if not next_candidate.exists() and str(next_candidate) not in reserved:
            return str(next_candidate)
        counter += 1


def build_tb_report_plan(
    template_path: str,
    tb_paths: list[str],
    output_dir: str,
    old_link_path: str,
) -> TemplateTbReportPlan:
    template_str = str(template_path or "").strip()
    if not template_str:
        raise ValueError("请先选择模板工作簿。")
    if not is_supported_template_file(template_str):
        raise ValueError("模板工作簿仅支持 .xlsx / .xlsm。")

    old_link_text = str(old_link_path or "").strip()
    if not old_link_text:
        raise ValueError("请先选择要替换的旧链接。")

    if not tb_paths:
        raise ValueError("请至少选择一个 TB 文件。")

    abs_template_path = os.path.abspath(template_str)
    abs_output_dir = os.path.abspath(str(output_dir or "").strip() or ".")
    template_suffix = Path(abs_template_path).suffix

    plan = TemplateTbReportPlan(
        template_path=abs_template_path,
        template_suffix=template_suffix,
        output_dir=abs_output_dir,
        old_link_path=old_link_text,
    )

    used_output_paths: set[str] = set()
    for index, raw_tb_path in enumerate(tb_paths, start=1):
        tb_text = str(raw_tb_path or "").strip()
        record = TemplateTbReportRecord(
            index=index,
            tb_path=tb_text,
            tb_name=os.path.basename(tb_text),
            old_link_path=old_link_text,
            new_link_path=tb_text,
        )

        if not tb_text:
            record.status = "跳过"
            record.message = "TB 文件路径为空"
            plan.records.append(record)
            continue
        if is_office_temp_file(tb_text):
            record.status = "跳过"
            record.message = "临时文件已跳过"
            plan.records.append(record)
            continue
        if not is_supported_tb_file(tb_text):
            record.status = "跳过"
            record.message = "不是支持的 TB 文件，仅支持 .xlsx / .xlsm"
            plan.records.append(record)
            continue

        abs_tb_path = os.path.abspath(tb_text)
        record.tb_path = abs_tb_path
        record.tb_name = os.path.basename(abs_tb_path)
        record.new_link_path = abs_tb_path

        output_name = build_output_filename(abs_tb_path, template_suffix)
        output_path = build_unique_output_path(abs_output_dir, output_name, used_output_paths)
        used_output_paths.add(output_path)

        record.output_name = os.path.basename(output_path)
        record.output_path = output_path
        record.status = "待执行"
        record.message = "待生成"
        plan.records.append(record)

    return plan


def summarize_tb_report_records(records: list[TemplateTbReportRecord]) -> dict:
    return {
        "total_count": len(records),
        "success_count": len([record for record in records if record.status == "成功"]),
        "skipped_count": len([record for record in records if record.status == "跳过"]),
        "failed_count": len([record for record in records if record.status == "失败"]),
    }


def write_tb_report_log_workbook(
    records: list[TemplateTbReportRecord],
    output_dir: str,
    template_path: str,
    old_link_path: str,
) -> str:
    abs_output_dir = os.path.abspath(output_dir)
    os.makedirs(abs_output_dir, exist_ok=True)
    log_path = _build_unique_log_path(abs_output_dir)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "处理日志"

    sheet.append([
        "模板路径",
        os.path.abspath(template_path),
        "被替换的旧链接",
        old_link_path,
        "处理时间",
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    ])
    for column in range(1, 7):
        sheet.cell(row=1, column=column).font = Font(bold=True)

    sheet.append([])
    sheet.append(LOG_HEADERS)
    for column in range(1, len(LOG_HEADERS) + 1):
        sheet.cell(row=3, column=column).font = Font(bold=True)

    for record in records:
        sheet.append([
            record.index,
            record.tb_name,
            record.tb_path,
            record.output_name,
            record.output_path,
            record.old_link_path,
            record.new_link_path,
            record.status,
            record.message,
        ])

    column_widths = {"A": 6, "B": 28, "C": 56, "D": 28, "E": 56, "F": 48, "G": 48, "H": 10, "I": 42}
    for column, width in column_widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A4"

    workbook.save(log_path)
    workbook.close()
    return log_path


def read_template_external_links(template_path: str) -> list[str]:
    abs_template_path = os.path.abspath(template_path)
    if not os.path.exists(abs_template_path):
        raise FileNotFoundError(f"模板工作簿不存在：{abs_template_path}")
    if not is_supported_template_file(abs_template_path):
        raise ValueError("模板工作簿仅支持 .xlsx / .xlsm。")

    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False
        workbook = excel.Workbooks.Open(
            abs_template_path,
            ReadOnly=True,
            UpdateLinks=0,
        )
        return normalize_com_link_sources(workbook.LinkSources(XL_EXCEL_LINKS))
    finally:
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()


def generate_reports_from_template_and_tb_files(
    template_path: str,
    tb_paths: list[str],
    output_dir: str,
    old_link_path: str,
    logger=None,
) -> dict:
    plan = build_tb_report_plan(template_path, tb_paths, output_dir, old_link_path)
    abs_template_path = plan.template_path
    if not os.path.exists(abs_template_path):
        raise FileNotFoundError(f"模板工作簿不存在：{abs_template_path}")
    os.makedirs(plan.output_dir, exist_ok=True)

    executable_records = [record for record in plan.records if record.status == "待执行"]

    if executable_records:
        import pythoncom
        import win32com.client

        pythoncom.CoInitialize()
        excel = None
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.AskToUpdateLinks = False

            for record in executable_records:
                workbook = None
                try:
                    if not os.path.exists(record.tb_path):
                        record.status = "失败"
                        record.message = "TB 文件不存在"
                        _log(logger, "error", f"TB 文件不存在，已跳过：{record.tb_path}")
                        continue

                    _log(logger, "info", f"正在生成报表 {record.index}：{record.output_path}")
                    shutil.copy2(abs_template_path, record.output_path)
                    workbook = excel.Workbooks.Open(
                        record.output_path,
                        ReadOnly=False,
                        UpdateLinks=0,
                    )
                    current_links = normalize_com_link_sources(workbook.LinkSources(XL_EXCEL_LINKS))
                    if not _link_exists(plan.old_link_path, current_links):
                        record.status = "失败"
                        record.message = "复制后的工作簿中未找到旧链接"
                        continue

                    workbook.ChangeLink(plan.old_link_path, record.new_link_path, XL_EXCEL_LINKS)
                    workbook.Save()
                    record.status = "成功"
                    record.message = "已生成并替换链接"
                except Exception as e:
                    record.status = "失败"
                    record.message = f"生成失败：{e}"
                    _log(logger, "error", f"生成失败：第 {record.index} 行 {record.tb_name}。详细信息：{e}")
                finally:
                    if workbook is not None:
                        try:
                            workbook.Close(SaveChanges=False)
                        except Exception:
                            pass
        finally:
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            pythoncom.CoUninitialize()

    log_path = write_tb_report_log_workbook(
        plan.records,
        plan.output_dir,
        abs_template_path,
        plan.old_link_path,
    )
    summary = summarize_tb_report_records(plan.records)
    return {
        "template_path": abs_template_path,
        "output_dir": plan.output_dir,
        "old_link_path": plan.old_link_path,
        "tb_file_count": len(plan.records),
        "log_path": log_path,
        "records": plan.records,
        **summary,
    }


def _build_unique_log_path(output_dir: str) -> str:
    base = Path(output_dir) / LOG_WORKBOOK_NAME
    if not base.exists():
        return str(base)
    stem = base.stem
    suffix = base.suffix
    counter = 2
    while True:
        candidate = base.with_name(f"{stem}_{counter}{suffix}")
        if not candidate.exists():
            return str(candidate)
        counter += 1


def _link_exists(target_link: str, current_links: list[str]) -> bool:
    return any(_same_link_path(link, target_link) for link in current_links)


def _same_link_path(left: str, right: str) -> bool:
    return os.path.normcase(str(left).strip()) == os.path.normcase(str(right).strip())


def _log(logger, level: str, message: str) -> None:
    if logger is None:
        return
    log_func = getattr(logger, level, None)
    if log_func:
        log_func(message)

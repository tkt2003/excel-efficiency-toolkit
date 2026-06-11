import os
import shutil
import tempfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


INSTRUCTION_SHEET_NAME = "使用说明"
SETTINGS_SHEET_NAME = "参数设置"
RULE_SHEET_NAME = "链接清单"
LOG_SHEET_NAME = "处理日志"

EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
XL_EXCEL_LINKS = 1

RULE_HEADERS = ["工作簿路径", "工作簿名", "原链接路径", "新链接路径", "是否执行", "状态", "说明"]
LOG_HEADERS = ["处理时间", "工作簿路径", "原链接路径", "新链接路径", "状态", "说明"]


@dataclass
class LinkReplaceSettings:
    auto_backup: bool = True
    save_target_workbook: bool = True
    skip_temp_files: bool = True
    only_excel_links: bool = True
    warnings: list[str] | None = None


@dataclass
class LinkReplaceRule:
    row_number: int
    workbook_path: str
    workbook_name: str
    old_link_path: str
    new_link_path: str
    execute_raw: str


@dataclass
class LinkReplaceAction:
    row_number: int
    workbook_path: str
    old_link_path: str
    new_link_path: str
    status: str
    message: str
    backup_path: str = ""


def is_office_temp_file(path: str) -> bool:
    return os.path.basename(str(path)).startswith("~$")


def is_excel_workbook_file(path: str) -> bool:
    filename = os.path.basename(str(path))
    if filename.startswith("~$"):
        return False
    return Path(filename).suffix.lower() in EXCEL_EXTENSIONS


def normalize_com_link_sources(link_sources) -> list[str]:
    if not link_sources:
        return []
    if isinstance(link_sources, str):
        return [link_sources]
    return [str(item) for item in link_sources if item]


def should_execute_rule(value: object) -> bool:
    text = _cell_to_text(value).casefold()
    if not text:
        return True
    return text not in {"否", "no", "n", "false", "0", "不执行", "跳过"}


def build_backup_path(workbook_path: str, timestamp: str | None = None) -> str:
    path = Path(workbook_path).resolve()
    stamp = timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = f"{path.stem}_批量换链接备份_{stamp}"
    counter = 1
    while True:
        suffix = "" if counter == 1 else f"_{counter}"
        candidate = path.with_name(f"{base_name}{suffix}{path.suffix}")
        if not candidate.exists():
            return str(candidate)
        counter += 1


def create_workbook_backup(workbook_path: str) -> str:
    if not os.path.exists(workbook_path):
        raise FileNotFoundError(f"目标工作簿不存在：{workbook_path}")
    backup_path = build_backup_path(workbook_path)
    shutil.copy2(workbook_path, backup_path)
    return backup_path


def create_link_replace_rule_workbook(link_records: list[dict], output_dir: str | None = None) -> str:
    target_dir = Path(output_dir or tempfile.gettempdir()).resolve()
    target_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = _get_unique_path(target_dir / f"批量换链接_临时规则表_{timestamp}.xlsx")

    workbook = Workbook()
    instruction_sheet = workbook.active
    instruction_sheet.title = INSTRUCTION_SHEET_NAME
    settings_sheet = workbook.create_sheet(SETTINGS_SHEET_NAME)
    rule_sheet = workbook.create_sheet(RULE_SHEET_NAME)
    log_sheet = workbook.create_sheet(LOG_SHEET_NAME)

    _write_instruction_sheet(instruction_sheet)
    _write_settings_sheet(settings_sheet)
    _write_rule_sheet(rule_sheet, link_records)
    _write_log_sheet(log_sheet)

    workbook.active = workbook.sheetnames.index(RULE_SHEET_NAME)
    workbook.save(output_path)
    workbook.close()
    return str(output_path)


def generate_temporary_link_replace_rule_workbook(source_paths: list[str], logger=None) -> dict:
    valid_paths = [str(Path(path).resolve()) for path in source_paths if is_excel_workbook_file(path)]
    if not valid_paths:
        raise ValueError("请选择至少一个 Excel 工作簿。")

    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    link_records: list[dict] = []
    read_success_count = 0

    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False

        for workbook_path in valid_paths:
            workbook = None
            try:
                _log(logger, "info", f"正在扫描外部链接：{workbook_path}")
                workbook = excel.Workbooks.Open(
                    workbook_path,
                    ReadOnly=True,
                    UpdateLinks=0,
                )
                read_success_count += 1
                links = normalize_com_link_sources(workbook.LinkSources(XL_EXCEL_LINKS))
                for link_path in links:
                    link_records.append(
                        {
                            "workbook_path": workbook_path,
                            "workbook_name": str(workbook.Name),
                            "old_link_path": link_path,
                        }
                    )
            except Exception as e:
                _log(logger, "error", f"扫描失败，已跳过：{workbook_path}。详细信息：{e}")
            finally:
                if workbook is not None:
                    try:
                        workbook.Close(SaveChanges=False)
                    except Exception:
                        pass
                    workbook = None

        if read_success_count == 0:
            raise RuntimeError("所有 Excel 工作簿都读取失败，未生成规则表。")

        rule_path = create_link_replace_rule_workbook(link_records)
        return {
            "source_file_count": len(valid_paths),
            "read_success_count": read_success_count,
            "link_count": len(link_records),
            "output_path": rule_path,
        }
    finally:
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()


def read_link_replace_settings(rule_workbook_path: str) -> LinkReplaceSettings:
    workbook = load_workbook(rule_workbook_path, data_only=True)
    warnings: list[str] = []
    try:
        sheet = workbook[SETTINGS_SHEET_NAME]
        values: dict[str, str] = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            key = _cell_to_text(row[0] if len(row) > 0 else None)
            value = _cell_to_text(row[1] if len(row) > 1 else None)
            if key:
                values[key] = value

        return LinkReplaceSettings(
            auto_backup=_read_yes_no_setting(values.get("执行前自动备份", ""), True, "执行前自动备份", warnings),
            save_target_workbook=_read_yes_no_setting(
                values.get("是否保存目标工作簿", ""),
                True,
                "是否保存目标工作簿",
                warnings,
            ),
            skip_temp_files=_read_yes_no_setting(values.get("跳过临时文件", ""), True, "跳过临时文件", warnings),
            only_excel_links=_read_yes_no_setting(
                values.get("仅处理 Excel 外部工作簿链接", ""),
                True,
                "仅处理 Excel 外部工作簿链接",
                warnings,
            ),
            warnings=warnings,
        )
    finally:
        workbook.close()


def read_link_replace_rules(rule_workbook_path: str) -> list[LinkReplaceRule]:
    workbook = load_workbook(rule_workbook_path, data_only=True)
    try:
        sheet = workbook[RULE_SHEET_NAME]
        rules: list[LinkReplaceRule] = []
        for row_number in range(2, sheet.max_row + 1):
            workbook_path = _cell_to_text(sheet.cell(row=row_number, column=1).value)
            workbook_name = _cell_to_text(sheet.cell(row=row_number, column=2).value)
            old_link_path = _cell_to_text(sheet.cell(row=row_number, column=3).value)
            new_link_path = _cell_to_text(sheet.cell(row=row_number, column=4).value)
            execute_raw = _cell_to_text(sheet.cell(row=row_number, column=5).value)
            if not any([workbook_path, workbook_name, old_link_path, new_link_path, execute_raw]):
                continue
            rules.append(
                LinkReplaceRule(
                    row_number=row_number,
                    workbook_path=workbook_path,
                    workbook_name=workbook_name,
                    old_link_path=old_link_path,
                    new_link_path=new_link_path,
                    execute_raw=execute_raw,
                )
            )
        return rules
    finally:
        workbook.close()


def build_link_replace_actions(
    rules: list[LinkReplaceRule],
    settings: LinkReplaceSettings | None = None,
) -> list[LinkReplaceAction]:
    settings = settings or LinkReplaceSettings()
    return [_build_action(rule, settings) for rule in rules]


def group_link_replace_actions_by_workbook_path(
    actions: list[LinkReplaceAction],
) -> dict[str, list[LinkReplaceAction]]:
    grouped: dict[str, list[LinkReplaceAction]] = {}
    for action in actions:
        key = _normalize_path_for_compare(action.workbook_path)
        grouped.setdefault(key, []).append(action)
    return grouped


def summarize_link_replace_actions(actions: list[LinkReplaceAction]) -> dict:
    return {
        "success_count": len([action for action in actions if action.status == "成功"]),
        "skipped_count": len([action for action in actions if action.status == "跳过"]),
        "failed_count": len([action for action in actions if action.status == "失败"]),
        "total_count": len(actions),
    }


def execute_link_replacement_from_rule_workbook(rule_workbook_path: str, logger=None) -> dict:
    settings = read_link_replace_settings(rule_workbook_path)
    for warning in settings.warnings or []:
        _log(logger, "info", warning)

    rules = read_link_replace_rules(rule_workbook_path)
    actions = build_link_replace_actions(rules, settings)
    execute_link_replace_actions_with_com(actions, settings, logger=logger)
    write_link_replace_results_to_workbook(rule_workbook_path, actions)
    summary = summarize_link_replace_actions(actions)
    summary["rule_count"] = len(rules)
    return summary


def execute_link_replace_actions_with_com(
    actions: list[LinkReplaceAction],
    settings: LinkReplaceSettings,
    logger=None,
) -> None:
    executable_actions = [action for action in actions if action.status == "待执行"]
    if not executable_actions:
        return

    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False

        grouped_actions = group_link_replace_actions_by_workbook_path(executable_actions)
        for actions_for_workbook in grouped_actions.values():
            workbook = None
            workbook_path = actions_for_workbook[0].workbook_path
            try:
                if not os.path.exists(workbook_path):
                    _mark_actions(actions_for_workbook, "失败", "工作簿不存在")
                    continue

                backup_path = ""
                if settings.auto_backup:
                    backup_path = create_workbook_backup(workbook_path)
                    _log(logger, "info", f"已生成备份：{backup_path}")

                _log(logger, "info", f"正在打开工作簿换链接：{workbook_path}")
                workbook = excel.Workbooks.Open(
                    os.path.abspath(workbook_path),
                    ReadOnly=False,
                    UpdateLinks=0,
                )
                current_links = normalize_com_link_sources(workbook.LinkSources(XL_EXCEL_LINKS))

                for action in actions_for_workbook:
                    action.backup_path = backup_path
                    try:
                        if not _link_exists(action.old_link_path, current_links):
                            action.status = "跳过"
                            action.message = "原链接不存在，已跳过"
                            continue
                        workbook.ChangeLink(action.old_link_path, action.new_link_path, XL_EXCEL_LINKS)
                        action.status = "成功"
                        action.message = "已替换"
                        current_links = [
                            action.new_link_path if _same_link_path(link, action.old_link_path) else link
                            for link in current_links
                        ]
                    except Exception as e:
                        action.status = "失败"
                        action.message = f"ChangeLink 失败：{e}"

                if settings.save_target_workbook and any(action.status == "成功" for action in actions_for_workbook):
                    workbook.Save()
                    _log(logger, "info", f"工作簿已保存：{workbook_path}")
                elif not settings.save_target_workbook:
                    _log(logger, "info", f"参数设置为不保存，已跳过保存：{workbook_path}")
            except Exception as e:
                _mark_actions(actions_for_workbook, "失败", f"工作簿处理失败：{e}")
                _log(logger, "error", f"工作簿处理失败：{workbook_path}。详细信息：{e}")
            finally:
                if workbook is not None:
                    try:
                        workbook.Close(SaveChanges=False)
                    except Exception:
                        pass
                    workbook = None
    finally:
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()


def write_link_replace_results_to_workbook(rule_workbook_path: str, actions: list[LinkReplaceAction]) -> None:
    workbook = load_workbook(rule_workbook_path)
    try:
        rule_sheet = workbook[RULE_SHEET_NAME]
        log_sheet = workbook[LOG_SHEET_NAME]
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for action in actions:
            rule_sheet.cell(row=action.row_number, column=6, value=action.status)
            rule_sheet.cell(row=action.row_number, column=7, value=action.message)
            log_sheet.append(
                [
                    now,
                    action.workbook_path,
                    action.old_link_path,
                    action.new_link_path,
                    action.status,
                    action.message,
                ]
            )

        workbook.save(rule_workbook_path)
    finally:
        workbook.close()


def _build_action(rule: LinkReplaceRule, settings: LinkReplaceSettings) -> LinkReplaceAction:
    if not rule.workbook_path:
        return _skip_action(rule, "工作簿路径为空")
    if settings.skip_temp_files and is_office_temp_file(rule.workbook_path):
        return _skip_action(rule, "临时文件已跳过")
    if not is_excel_workbook_file(rule.workbook_path):
        return _skip_action(rule, "不是支持的 Excel 工作簿")
    if not rule.old_link_path:
        return _skip_action(rule, "原链接路径为空")
    if not rule.new_link_path:
        return _skip_action(rule, "未填写新链接路径")
    if not should_execute_rule(rule.execute_raw):
        return _skip_action(rule, "是否执行为否，已跳过")
    return LinkReplaceAction(
        row_number=rule.row_number,
        workbook_path=str(Path(rule.workbook_path).resolve()),
        old_link_path=rule.old_link_path,
        new_link_path=rule.new_link_path,
        status="待执行",
        message="待替换",
    )


def _skip_action(rule: LinkReplaceRule, message: str) -> LinkReplaceAction:
    return LinkReplaceAction(
        row_number=rule.row_number,
        workbook_path=rule.workbook_path,
        old_link_path=rule.old_link_path,
        new_link_path=rule.new_link_path,
        status="跳过",
        message=message,
    )


def _mark_actions(actions: list[LinkReplaceAction], status: str, message: str) -> None:
    for action in actions:
        action.status = status
        action.message = message


def _link_exists(target_link: str, current_links: list[str]) -> bool:
    return any(_same_link_path(link, target_link) for link in current_links)


def _same_link_path(left: str, right: str) -> bool:
    return os.path.normcase(str(left).strip()) == os.path.normcase(str(right).strip())


def _write_instruction_sheet(sheet) -> None:
    rows = [
        ["批量换链接使用说明"],
        ["1. 本功能只处理 Excel 原生外部工作簿链接。"],
        ["2. 在【链接清单】D 列填写新链接路径。"],
        ["3. E 列为空或“是”表示执行，填写“否”表示跳过。"],
        ["4. 保存并关闭规则表后，回到工具台点击执行。"],
        ["5. 执行前默认自动备份目标工作簿，执行后回写状态和处理日志。"],
        ["6. 暂不处理 VBA、Power Query、数据透视表连接、ODBC/OLEDB 连接。"],
    ]
    _append_rows(sheet, rows)
    sheet["A1"].font = Font(bold=True)
    sheet.column_dimensions["A"].width = 92


def _write_settings_sheet(sheet) -> None:
    rows = [
        ["参数项", "参数值", "说明"],
        ["执行前自动备份", "是", "执行 ChangeLink 前复制一份目标工作簿备份。"],
        ["是否保存目标工作簿", "是", "成功执行后保存被修改的目标工作簿。"],
        ["跳过临时文件", "是", "跳过以 ~$ 开头的 Office 临时文件。"],
        ["仅处理 Excel 外部工作簿链接", "是", "当前版本固定只处理 Workbook.LinkSources 中的 Excel 链接。"],
    ]
    _append_rows(sheet, rows)
    _style_header(sheet, 1, len(rows[0]))
    _set_widths(sheet, {"A": 28, "B": 18, "C": 72})


def _write_rule_sheet(sheet, link_records: list[dict]) -> None:
    _append_rows(sheet, [RULE_HEADERS])
    _style_header(sheet, 1, len(RULE_HEADERS))
    fill = PatternFill(fill_type="solid", fgColor="FFF2CC")

    for record in link_records:
        workbook_path = _cell_to_text(record.get("workbook_path"))
        sheet.append(
            [
                workbook_path,
                _cell_to_text(record.get("workbook_name")) or Path(workbook_path).name,
                _cell_to_text(record.get("old_link_path")),
                "",
                "是",
                "",
                "",
            ]
        )

    for row in range(2, sheet.max_row + 1):
        sheet.cell(row=row, column=4).fill = fill
        sheet.cell(row=row, column=5).fill = fill

    sheet.freeze_panes = "A2"
    _set_widths(sheet, {"A": 58, "B": 24, "C": 58, "D": 58, "E": 12, "F": 12, "G": 42})


def _write_log_sheet(sheet) -> None:
    _append_rows(sheet, [LOG_HEADERS])
    _style_header(sheet, 1, len(LOG_HEADERS))
    sheet.freeze_panes = "A2"
    _set_widths(sheet, {"A": 20, "B": 58, "C": 58, "D": 58, "E": 12, "F": 42})


def _append_rows(sheet, rows: list[list[object]]) -> None:
    for row in rows:
        sheet.append(row)


def _style_header(sheet, row_number: int, column_count: int) -> None:
    for column in range(1, column_count + 1):
        sheet.cell(row=row_number, column=column).font = Font(bold=True)


def _set_widths(sheet, widths: dict[str, int]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _read_yes_no_setting(value: str, default: bool, setting_name: str, warnings: list[str]) -> bool:
    normalized = value.strip().casefold()
    if normalized in {"是", "yes", "y", "true", "1"}:
        return True
    if normalized in {"否", "no", "n", "false", "0"}:
        return False
    if normalized:
        default_text = "是" if default else "否"
        warnings.append(f"参数“{setting_name}”无效，已使用默认值“{default_text}”。")
    return default


def _cell_to_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _get_unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    counter = 2
    while True:
        candidate = path.with_name(f"{path.stem}_{counter}{path.suffix}")
        if not candidate.exists():
            return candidate
        counter += 1


def _normalize_path_for_compare(path: str) -> str:
    return os.path.normcase(os.path.abspath(path))


def _log(logger, level: str, message: str) -> None:
    if logger is None:
        return
    log_func = getattr(logger, level, None)
    if log_func:
        log_func(message)

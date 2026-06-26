import os
import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


MODE_COMMON = "common"
MODE_PER_FILE = "per_file"

COMMON_RULE_HEADERS = ["查找内容", "替换为", "是否启用"]
PER_FILE_RULE_HEADERS = ["Word文件路径", "查找内容", "替换为", "是否启用"]
SUPPORTED_WORD_EXTENSIONS = {".docx", ".doc", ".docm"}

WD_FIND_STOP = 0
WD_REPLACE_ALL = 2
TARGET_STORY_TYPES = {1, 6, 7, 8, 9, 10, 11}


@dataclass
class WordReplaceRule:
    row_number: int
    find_text: str
    replace_text: str
    enabled_raw: str = ""
    word_path: str = ""


@dataclass
class WordReplaceAction:
    file_path: str
    find_text: str
    replace_text: str
    rule_row_number: int
    status: str = "待预览"
    message: str = "待预览"
    estimated_count: int = 0
    replace_count: int = 0
    backup_path: str = ""
    backup_status: str = ""

    def to_preview_dict(self) -> dict:
        return {
            "file_path": self.file_path,
            "find_text": self.find_text,
            "replace_text": self.replace_text,
            "estimated_count": self.estimated_count,
            "status": self.status,
            "message": self.message,
            "rule_row_number": self.rule_row_number,
        }


@dataclass
class WordReplaceFileLog:
    file_path: str
    file_name: str
    rule_count: int
    replace_count: int
    status: str
    message: str = ""
    backup_path: str = ""
    backup_status: str = ""

    def to_dict(self) -> dict:
        return {
            "file_path": self.file_path,
            "file_name": self.file_name,
            "rule_count": self.rule_count,
            "replace_count": self.replace_count,
            "status": self.status,
            "message": self.message,
            "backup_path": self.backup_path,
            "backup_status": self.backup_status,
        }


def is_word_file(path: str) -> bool:
    filename = os.path.basename(str(path))
    if filename.startswith("~$"):
        return False
    return Path(filename).suffix.lower() in SUPPORTED_WORD_EXTENSIONS


def parse_rule_enabled(value: object) -> bool:
    text = _cell_to_text(value).casefold()
    if not text:
        return True
    if text in {"是", "yes", "y", "true", "1"}:
        return True
    if text in {"否", "no", "n", "false", "0"}:
        return False
    return True


def load_common_replace_rules(rule_workbook_path: str, sheet_name: str | None = None) -> list[WordReplaceRule]:
    workbook = load_workbook(rule_workbook_path, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        headers = _read_header_indexes(sheet, COMMON_RULE_HEADERS)
        rules: list[WordReplaceRule] = []
        for row_number in range(2, sheet.max_row + 1):
            find_text = _cell_to_text(sheet.cell(row=row_number, column=headers["查找内容"]).value)
            replace_text = _cell_to_text(sheet.cell(row=row_number, column=headers["替换为"]).value)
            enabled_raw = _cell_to_text(sheet.cell(row=row_number, column=headers["是否启用"]).value)
            if not any([find_text, replace_text, enabled_raw]):
                continue
            if not parse_rule_enabled(enabled_raw):
                continue
            if not find_text:
                raise ValueError(f"第 {row_number} 行查找内容为空。")
            rules.append(
                WordReplaceRule(
                    row_number=row_number,
                    find_text=find_text,
                    replace_text=replace_text,
                    enabled_raw=enabled_raw,
                )
            )
        return rules
    finally:
        workbook.close()


def build_common_replace_rules_from_entries(entries: list[tuple[object, object]]) -> list[WordReplaceRule]:
    rules: list[WordReplaceRule] = []
    for index, (find_value, replace_value) in enumerate(entries, start=1):
        find_text = _cell_to_text(find_value)
        replace_text = _cell_to_text(replace_value)
        if not find_text and not replace_text:
            continue
        if not find_text:
            raise ValueError(f"第 {index} 行查找内容为空，请修正。")
        rules.append(
            WordReplaceRule(
                row_number=index,
                find_text=find_text,
                replace_text=replace_text,
                enabled_raw="是",
            )
        )
    if not rules:
        raise ValueError("请至少填写一条替换规则。")
    return rules


def load_per_file_replace_rules(
    rule_workbook_path: str,
    sheet_name: str | None = None,
) -> dict[str, list[WordReplaceRule]]:
    workbook = load_workbook(rule_workbook_path, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        headers = _read_header_indexes(sheet, PER_FILE_RULE_HEADERS)
        rules_by_file: dict[str, list[WordReplaceRule]] = {}
        for row_number in range(2, sheet.max_row + 1):
            word_path = _cell_to_text(sheet.cell(row=row_number, column=headers["Word文件路径"]).value)
            find_text = _cell_to_text(sheet.cell(row=row_number, column=headers["查找内容"]).value)
            replace_text = _cell_to_text(sheet.cell(row=row_number, column=headers["替换为"]).value)
            enabled_raw = _cell_to_text(sheet.cell(row=row_number, column=headers["是否启用"]).value)
            if not any([word_path, find_text, replace_text, enabled_raw]):
                continue
            if not parse_rule_enabled(enabled_raw):
                continue
            if not word_path:
                raise ValueError(f"第 {row_number} 行 Word文件路径为空。")
            if not find_text:
                raise ValueError(f"第 {row_number} 行查找内容为空。")
            resolved_word_path = str(Path(word_path).resolve())
            rule = WordReplaceRule(
                row_number=row_number,
                find_text=find_text,
                replace_text=replace_text,
                enabled_raw=enabled_raw,
                word_path=resolved_word_path,
            )
            rules_by_file.setdefault(_normalize_path_key(resolved_word_path), []).append(rule)
        return rules_by_file
    finally:
        workbook.close()


def create_per_file_replace_rule_template(word_paths: list[str], output_dir: str | None = None) -> str:
    valid_paths = [str(Path(path).resolve()) for path in word_paths if is_word_file(path)]
    if not valid_paths:
        raise ValueError("请选择至少一个 Word 文件。")

    target_dir = Path(output_dir or Path(valid_paths[0]).parent).resolve()
    target_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = _get_unique_path(target_dir / f"Word批量替换规则_{timestamp}.xlsx")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "替换规则"
    sheet.append(PER_FILE_RULE_HEADERS)
    for path in valid_paths:
        sheet.append([path, "", "", "是"])

    for column in range(1, len(PER_FILE_RULE_HEADERS) + 1):
        sheet.cell(row=1, column=column).font = Font(bold=True)
    fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row=row, column=2).fill = fill
        sheet.cell(row=row, column=3).fill = fill
        sheet.cell(row=row, column=4).fill = fill
    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 72
    sheet.column_dimensions["B"].width = 24
    sheet.column_dimensions["C"].width = 24
    sheet.column_dimensions["D"].width = 14

    workbook.save(output_path)
    workbook.close()
    return str(output_path)


def build_word_replace_actions(
    word_paths: list[str],
    mode: str,
    common_rules: list[WordReplaceRule] | None = None,
    per_file_rules: dict[str, list[WordReplaceRule]] | None = None,
) -> list[WordReplaceAction]:
    selected_paths = [str(Path(path).resolve()) for path in word_paths if str(path).strip()]
    if mode not in {MODE_COMMON, MODE_PER_FILE}:
        raise ValueError("未知 Word 替换模式。")

    normalized_per_file_rules = _normalize_per_file_rules(per_file_rules or {})
    actions: list[WordReplaceAction] = []
    for file_path in selected_paths:
        if not is_word_file(file_path):
            actions.append(_skip_action(file_path, "不是支持的 Word 文件"))
            continue

        if mode == MODE_COMMON:
            rules = common_rules or []
            no_rule_message = "规则表没有可执行规则"
        else:
            rules = normalized_per_file_rules.get(_normalize_path_key(file_path), [])
            no_rule_message = "规则表未匹配该文件"

        if not rules:
            actions.append(_skip_action(file_path, no_rule_message))
            continue

        for rule in rules:
            if not rule.find_text:
                actions.append(_skip_action(file_path, "查找内容为空", rule))
                continue
            actions.append(
                WordReplaceAction(
                    file_path=file_path,
                    find_text=rule.find_text,
                    replace_text=rule.replace_text,
                    rule_row_number=rule.row_number,
                )
            )
    return actions


def preview_word_replace(
    word_paths: list[str],
    mode: str,
    common_rules: list[WordReplaceRule] | None = None,
    per_file_rules: dict[str, list[WordReplaceRule]] | None = None,
    logger=None,
) -> dict:
    actions = build_word_replace_actions(
        word_paths=word_paths,
        mode=mode,
        common_rules=common_rules,
        per_file_rules=per_file_rules,
    )
    executable_actions = [action for action in actions if action.status == "待预览"]
    if not executable_actions:
        return _build_preview_result(mode, actions)

    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    word = None
    document = None
    try:
        word = win32com.client.DispatchEx("Word.Application")
        word.Visible = False
        word.DisplayAlerts = 0

        for file_path, file_actions in group_actions_by_file(executable_actions).items():
            document = None
            try:
                _log(logger, "info", f"正在预览 Word 替换：{file_path}")
                document = _open_word_document(word, file_path, read_only=True)
                for action in file_actions:
                    try:
                        estimated_count = _count_occurrences(document, action.find_text)
                        action.estimated_count = estimated_count
                        if estimated_count <= 0:
                            action.status = "跳过"
                            action.message = "预计替换次数为 0"
                            continue

                        action.status = "待执行"
                        action.message = "预览完成"
                        _replace_all(document, action.find_text, action.replace_text)
                    except Exception as e:
                        action.status = "失败"
                        action.message = f"预览失败：{e}"
            except Exception as e:
                _mark_actions(file_actions, "失败", f"Word 文件打开失败：{e}")
                _log(logger, "error", f"Word 文件预览失败：{file_path}。详细信息：{e}")
            finally:
                if document is not None:
                    try:
                        document.Close(SaveChanges=False)
                    except Exception:
                        pass
                    document = None
    finally:
        if document is not None:
            try:
                document.Close(SaveChanges=False)
            except Exception:
                pass
        if word is not None:
            try:
                word.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()

    return _build_preview_result(mode, actions)


def execute_word_replace(preview_actions: list[WordReplaceAction], logger=None) -> dict:
    if not preview_actions:
        raise ValueError("请先预览替换计划。")

    file_logs: list[WordReplaceFileLog] = []
    grouped_all_actions = group_actions_by_file(preview_actions)
    executable_by_file = {
        file_path: [action for action in actions if action.status == "待执行" and action.estimated_count > 0]
        for file_path, actions in grouped_all_actions.items()
    }
    executable_by_file = {file_path: actions for file_path, actions in executable_by_file.items() if actions}

    for file_path, actions in grouped_all_actions.items():
        if file_path not in executable_by_file:
            file_logs.append(
                WordReplaceFileLog(
                    file_path=file_path,
                    file_name=Path(file_path).name,
                    rule_count=len(actions),
                    replace_count=0,
                    status="跳过",
                    message=_join_action_messages(actions) or "没有需要执行的替换",
                    backup_status="无需备份",
                )
            )

    if not executable_by_file:
        return _build_execution_result(preview_actions, file_logs)

    import pythoncom
    import win32com.client

    batch_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    pythoncom.CoInitialize()
    word = None
    document = None
    try:
        word = win32com.client.DispatchEx("Word.Application")
        word.Visible = False
        word.DisplayAlerts = 0

        for file_path, file_actions in executable_by_file.items():
            document = None
            file_log = WordReplaceFileLog(
                file_path=file_path,
                file_name=Path(file_path).name,
                rule_count=len(file_actions),
                replace_count=0,
                status="失败",
                backup_status="未备份",
            )
            try:
                backup_path = create_word_backup(file_path, batch_timestamp=batch_timestamp)
                file_log.backup_path = backup_path
                file_log.backup_status = "已备份"
                for action in file_actions:
                    action.backup_path = backup_path
                    action.backup_status = "已备份"
                _log(logger, "info", f"已生成 Word 备份：{backup_path}")
            except Exception as e:
                message = f"备份失败：{e}"
                _mark_actions(file_actions, "失败", message)
                for action in file_actions:
                    action.backup_status = message
                file_log.message = message
                file_log.backup_status = message
                file_logs.append(file_log)
                _log(logger, "error", f"Word 备份失败，已跳过：{file_path}。详细信息：{e}")
                continue

            try:
                _log(logger, "info", f"正在执行 Word 替换：{file_path}")
                document = _open_word_document(word, file_path, read_only=False)
                file_replace_count = 0
                has_failed_rule = False
                has_success_rule = False
                for action in file_actions:
                    try:
                        replace_count = _count_occurrences(document, action.find_text)
                        action.replace_count = replace_count
                        if replace_count <= 0:
                            action.status = "跳过"
                            action.message = "执行时未找到内容"
                            continue
                        _replace_all(document, action.find_text, action.replace_text)
                        action.status = "成功"
                        action.message = "已替换"
                        has_success_rule = True
                        file_replace_count += replace_count
                    except Exception as e:
                        action.status = "失败"
                        action.message = f"替换失败：{e}"
                        has_failed_rule = True

                if has_success_rule:
                    try:
                        document.Save()
                    except Exception as e:
                        _mark_success_actions_failed(file_actions, f"保存失败：{e}")
                        file_log.status = "失败"
                        file_log.message = f"保存失败：{e}"
                        file_log.replace_count = 0
                        file_logs.append(file_log)
                        _log(logger, "error", f"Word 文件保存失败：{file_path}。详细信息：{e}")
                        continue

                file_log.replace_count = file_replace_count
                if has_failed_rule:
                    file_log.status = "失败"
                    file_log.message = "部分规则替换失败"
                elif has_success_rule:
                    file_log.status = "成功"
                    file_log.message = "已保存"
                else:
                    file_log.status = "跳过"
                    file_log.message = "执行时未找到需要替换的内容"
                file_logs.append(file_log)
            except Exception as e:
                _mark_actions(file_actions, "失败", f"Word 文件处理失败：{e}")
                file_log.status = "失败"
                file_log.message = f"Word 文件处理失败：{e}"
                file_logs.append(file_log)
                _log(logger, "error", f"Word 文件处理失败：{file_path}。详细信息：{e}")
            finally:
                if document is not None:
                    try:
                        document.Close(SaveChanges=False)
                    except Exception:
                        pass
                    document = None
    finally:
        if document is not None:
            try:
                document.Close(SaveChanges=False)
            except Exception:
                pass
        if word is not None:
            try:
                word.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()

    return _build_execution_result(preview_actions, file_logs)


def build_word_backup_path(word_path: str, timestamp: str | None = None) -> str:
    path = Path(word_path).resolve()
    stamp = timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_dir = path.parent / "老头表格助手备份" / f"word_replace_{stamp}"
    candidate = backup_dir / path.name
    counter = 1
    while candidate.exists():
        counter += 1
        candidate = backup_dir / f"{path.stem}_{counter}{path.suffix}"
    return str(candidate)


def create_word_backup(word_path: str, batch_timestamp: str | None = None) -> str:
    if not os.path.exists(word_path):
        raise FileNotFoundError(f"Word 文件不存在：{word_path}")
    backup_path = build_word_backup_path(word_path, timestamp=batch_timestamp)
    Path(backup_path).parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(word_path, backup_path)
    return backup_path


def group_actions_by_file(actions: list[WordReplaceAction]) -> dict[str, list[WordReplaceAction]]:
    grouped: dict[str, list[WordReplaceAction]] = {}
    for action in actions:
        grouped.setdefault(str(Path(action.file_path).resolve()), []).append(action)
    return grouped


def has_empty_replacement(rules: list[WordReplaceRule] | dict[str, list[WordReplaceRule]]) -> bool:
    if isinstance(rules, dict):
        all_rules = [rule for file_rules in rules.values() for rule in file_rules]
    else:
        all_rules = rules
    return any(rule.replace_text == "" for rule in all_rules)


def _build_preview_result(mode: str, actions: list[WordReplaceAction]) -> dict:
    return {
        "mode": mode,
        "actions": actions,
        "items": [action.to_preview_dict() for action in actions],
        "summary": {
            "file_count": len(group_actions_by_file(actions)),
            "rule_count": len(actions),
            "pending_count": len([action for action in actions if action.status == "待执行"]),
            "skipped_count": len([action for action in actions if action.status == "跳过"]),
            "failed_count": len([action for action in actions if action.status == "失败"]),
            "estimated_replace_count": sum(action.estimated_count for action in actions),
        },
    }


def _build_execution_result(actions: list[WordReplaceAction], file_logs: list[WordReplaceFileLog]) -> dict:
    return {
        "actions": actions,
        "items": [action.to_preview_dict() for action in actions],
        "file_logs": [file_log.to_dict() for file_log in file_logs],
        "summary": {
            "file_count": len(file_logs),
            "success_file_count": len([item for item in file_logs if item.status == "成功"]),
            "skipped_file_count": len([item for item in file_logs if item.status == "跳过"]),
            "failed_file_count": len([item for item in file_logs if item.status == "失败"]),
            "rule_count": len(actions),
            "success_rule_count": len([action for action in actions if action.status == "成功"]),
            "skipped_rule_count": len([action for action in actions if action.status == "跳过"]),
            "failed_rule_count": len([action for action in actions if action.status == "失败"]),
            "replace_count": sum(action.replace_count for action in actions),
        },
    }


def _open_word_document(word, file_path: str, read_only: bool):
    return word.Documents.Open(
        FileName=os.path.abspath(file_path),
        ConfirmConversions=False,
        ReadOnly=read_only,
        AddToRecentFiles=False,
    )


def _count_occurrences(document, find_text: str) -> int:
    count = 0
    for story_range in _iter_target_story_ranges(document):
        count += str(story_range.Text).count(find_text)
    return count


def _replace_all(document, find_text: str, replace_text: str) -> None:
    for story_range in _iter_target_story_ranges(document):
        search_range = story_range.Duplicate
        find = search_range.Find
        find.ClearFormatting()
        find.Replacement.ClearFormatting()
        find.Execute(
            FindText=find_text,
            MatchCase=False,
            MatchWholeWord=False,
            MatchWildcards=False,
            MatchSoundsLike=False,
            MatchAllWordForms=False,
            Forward=True,
            Wrap=WD_FIND_STOP,
            Format=False,
            ReplaceWith=replace_text,
            Replace=WD_REPLACE_ALL,
        )


def _iter_target_story_ranges(document):
    seen = set()
    for story_type in TARGET_STORY_TYPES:
        try:
            current = document.StoryRanges(story_type)
        except Exception:
            continue
        while current is not None:
            try:
                current_story_type = int(current.StoryType)
            except Exception:
                current_story_type = story_type
            if current_story_type in TARGET_STORY_TYPES:
                key = _story_range_key(current, current_story_type)
                if key not in seen:
                    seen.add(key)
                    yield current
            try:
                next_range = current.NextStoryRange
            except Exception:
                next_range = None
            if next_range is not None:
                try:
                    next_story_type = int(next_range.StoryType)
                except Exception:
                    next_story_type = story_type
                if _story_range_key(next_range, next_story_type) in seen:
                    next_range = None
            current = next_range


def _story_range_key(story_range, story_type: int) -> tuple:
    try:
        return (story_type, int(story_range.Start), int(story_range.End))
    except Exception:
        return (story_type, id(story_range))


def _read_header_indexes(sheet, required_headers: list[str]) -> dict[str, int]:
    header_indexes = {}
    for cell in sheet[1]:
        header = _cell_to_text(cell.value)
        if header:
            header_indexes[header] = cell.column
    missing_headers = [header for header in required_headers if header not in header_indexes]
    if missing_headers:
        raise ValueError(f"规则表缺少表头：{'、'.join(missing_headers)}")
    return header_indexes


def _normalize_per_file_rules(
    per_file_rules: dict[str, list[WordReplaceRule]]
) -> dict[str, list[WordReplaceRule]]:
    return {_normalize_path_key(path): rules for path, rules in per_file_rules.items()}


def _normalize_path_key(path: str) -> str:
    return os.path.normcase(str(Path(path).resolve()))


def _skip_action(
    file_path: str,
    message: str,
    rule: WordReplaceRule | None = None,
) -> WordReplaceAction:
    return WordReplaceAction(
        file_path=str(Path(file_path).resolve()),
        find_text=rule.find_text if rule else "",
        replace_text=rule.replace_text if rule else "",
        rule_row_number=rule.row_number if rule else 0,
        status="跳过",
        message=message,
    )


def _mark_actions(actions: list[WordReplaceAction], status: str, message: str) -> None:
    for action in actions:
        action.status = status
        action.message = message


def _mark_success_actions_failed(actions: list[WordReplaceAction], message: str) -> None:
    for action in actions:
        if action.status == "成功":
            action.status = "失败"
            action.message = message
            action.replace_count = 0


def _join_action_messages(actions: list[WordReplaceAction]) -> str:
    messages = []
    for action in actions:
        if action.message and action.message not in messages:
            messages.append(action.message)
    return "；".join(messages)


def _cell_to_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _get_unique_path(path: Path) -> Path:
    if not path.exists():
        return path

    counter = 2
    while True:
        candidate = path.with_name(f"{path.stem}_{counter}{path.suffix}")
        if not candidate.exists():
            return candidate
        counter += 1


def _log(logger, level: str, message: str) -> None:
    if logger is None:
        return
    getattr(logger, level)(message)
